    # core/views.py
from functools import wraps
import os
from urllib.parse import urlparse
from django.shortcuts import render, redirect
from django.contrib import messages
from django.views.decorators.csrf import csrf_exempt
from django.db import connection, transaction
from django.db.utils import IntegrityError
from .url_encryption import encrypt_id, decrypt_id
from django.http import HttpResponseForbidden, HttpResponseRedirect, JsonResponse, HttpResponse
from django.views.decorators.cache import cache_page
from django.views.decorators.http import require_POST
from django.core.files.storage import FileSystemStorage
from django.views.decorators.cache import cache_control 
from django.views.decorators.clickjacking import xframe_options_exempt
from django.contrib.auth.hashers import check_password
from django import forms
from django.urls import reverse
import hashlib
from django.utils import timezone
import logging

logger = logging.getLogger(__name__)
import time
import logging
import secrets
import base64
from django.db.models import Q, F
from datetime import datetime
from .decorators import custom_login_required
from .email_templates_views import ensure_email_templates_for_school

from .utils import (
    safe_int, safe_float, safe_strptime, safe_json_obj,
    get_context, validate_uploaded_file, bytes_to_data_uri,
    _get_custom_session_info,
    SESSION_COOKIE_NAME, OTP_COOKIE_NAME, OTP_COOKIE_MAX_AGE,
    ERP_DEFAULT_LOGO_STATIC,
    ALLOWED_IMAGE_TYPES, ALLOWED_DOCUMENT_TYPES, MAX_FILE_SIZE
)
from .branding_utils import get_branding_title
from .user_views import (
    user_list, load_more_users, create_user, update_user,
    user_soft_delete, user_restore, get_user_password, validate_user_api
)

import json
import threading
import re
from datetime import datetime
from django.utils import timezone
from django.contrib.auth.hashers import make_password
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger

from .auth_utils import generate_and_store_otp, verify_otp
import csv
from mail.utils import send_email_by_code
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import StringIO
from .pdf_generator import generate_pdf_from_template, generate_employee_job_letter_pdf
from mail.utils import send_email_by_code
from .email_queue import send_admission_emails_async_v2, email_queue
from .database_email_queue import send_admission_emails_async_database, database_email_queue
from .email_tracking_models import EmailTracking, EmailTrackingManager
import pyodbc
from django.core.serializers.json import DjangoJSONEncoder
import numpy as np

# --------------------------------------------------------------------------
# Generic validation and execution helpers (remaining)
# --------------------------------------------------------------------------

@custom_login_required
def school_dropdown_api(request):
    """
    Global API to get school dropdown values.
    Returns: JSON list of schools with {id: encrypted_id, name: [Code] Name}
    """
    try:
        from .utils import get_school_dropdown
        raw_schools = get_school_dropdown()
        
        schools = []
        for s in raw_schools:
            schools.append({
                'id': encrypt_id(s['SchoolID']),
                'name': s['DisplayName'],
                'raw_id': s['SchoolID'] # Optional: For internal use if needed, but safer to stick to encrypted
            })
            
        return JsonResponse({'status': 'SUCCESS', 'data': schools})
    except Exception as e:
        logger.error(f"Error in school_dropdown_api: {e}")
        return JsonResponse({'status': 'ERROR', 'message': str(e)}, status=500)

# Global logger
logger = logging.getLogger(__name__)



def send_employee_registration_email_async(email_data):
    """Queue employee registration email using database email tracking system"""
    try:
        logger.info("Queuing employee registration email")
        
        # Get school information
        school_id = email_data.get('school_id')
        
        # Ensure email templates exist for this school
        ensure_email_templates_for_school(school_id)
        school_name = "ShikshaWave School"  # Default fallback
        
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT "SchoolName", "SchoolLogo" FROM "SchoolMaster" WHERE "SchoolID" = %s
            """, [school_id])
            school_result = cursor.fetchone()
            if school_result:
                school_name = school_result[0]
                # Handle school logo - convert binary to base64 if it exists
                if school_result[1]:
                    import base64
                    try:
                        # Convert binary logo to base64 data URL
                        logo_base64 = base64.b64encode(school_result[1]).decode('utf-8')
                        school_logo = f"data:image/png;base64,{logo_base64}"
                    except Exception as e:
                        logger.warning(f"Failed to convert school logo to base64: {str(e)}")
                        school_logo = ""
                else:
                    school_logo = ""
        
        # Get school rules dynamically based on profile/role
        school_rules = []
        profile_id = email_data.get('profile_id')
        
        # Map profile_id to role name for rules lookup and position
        profile_role_mapping = {
            1: 'Super Admin',
            2: 'School Admin', 
            3: 'Teacher',
            4: 'Student',
            5: 'Parent',
            6: 'Driver',
            7: 'Librarian',
            8: 'Accountant'
        }
        
        # Get role name for position and rules
        role_name = profile_role_mapping.get(profile_id, 'Staff') if profile_id else 'Staff'
        
        if profile_id:
            
            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT "Title", "Description" 
                    FROM "SchoolRulesInstructions" 
                    WHERE "AppliesTo" = %s AND "SchoolID" = %s AND "IsDeleted" = FALSE
                """, [role_name, school_id])
                school_rules = [{'Title': row[0], 'Description': row[1]} for row in cursor.fetchall()]
                
                logger.info(f"Found {len(school_rules)} rules for profile")
        else:
            logger.warning("No profile_id provided, using default Teacher rules")
            # Fallback to Teacher rules if no profile_id
            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT "Title", "Description" 
                    FROM "SchoolRulesInstructions" 
                    WHERE "AppliesTo" = 'Teacher' AND "SchoolID" = %s AND "IsDeleted" = FALSE
                """, [school_id])
                school_rules = [{'Title': row[0], 'Description': row[1]} for row in cursor.fetchall()]
        
        # Get school-specific terms & conditions from TermsConditions table
        school_terms_conditions = []
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT "Title", "Description", "Category", "DisplayOrder"
                FROM "TermsConditions" 
                WHERE "SchoolId" = %s AND "IsActive" = TRUE
                ORDER BY 
                    CASE "Category" 
                        WHEN 'Employment' THEN 1
                        WHEN 'General' THEN 2
                        WHEN 'Financial' THEN 3
                        WHEN 'Academic' THEN 4
                        ELSE 5
                    END,
                    "DisplayOrder"
            """, [school_id])
            school_terms_conditions = [{'Title': row[0], 'Description': row[1], 'Category': row[2], 'DisplayOrder': row[3]} for row in cursor.fetchall()]
            
            logger.info(f"Found {len(school_terms_conditions)} terms & conditions")
        
        # Extract employee photo from document_components if available
        employee_photo = ""
        document_components = email_data.get('document_components', [])
        for doc in document_components:
            if doc.get('DocumentType') == 'Employee Passport Photo':
                file_content = doc.get('FileContent')
                file_extension = doc.get('FileExtension', 'png')
                if file_content:
                    employee_photo = f"data:image/{file_extension};base64,{file_content}"
                break
        
        # Get salary component details and format breakdown
        salary_breakdown = ""
        salary_components_with_details = []
        if email_data.get('salary_components'):
            with connection.cursor() as cursor:
                for component in email_data.get('salary_components', []):
                    cursor.execute("""
                        SELECT "ComponentName", "ComponentType" 
                        FROM "SalaryComponentMaster" 
                        WHERE "ComponentID" = %s AND "SchoolID" = %s
                    """, [component.get('ComponentID'), school_id])
                    result = cursor.fetchone()
                    if result:
                        component_detail = {
                            'ComponentID': component.get('ComponentID'),
                            'ComponentName': result[0],
                            'ComponentType': result[1],
                            'Amount': component.get('Amount', 0)
                        }
                        salary_components_with_details.append(component_detail)
            
            # Format as plain text for email template
            salary_breakdown = ""
            for component in salary_components_with_details:
                salary_breakdown += f"• {component.get('ComponentName', 'N/A')}: ₹{component.get('Amount', 0):.2f} ({component.get('ComponentType', 'N/A')})\n"
        
        # Format school rules as plain text
        rules_html = ""
        if school_rules:
            for rule in school_rules:
                rules_html += f"• {rule['Title']}: {rule['Description']}\n"
        else:
            rules_html = "Standard school policies and procedures apply."
        
        # Get earnings and deductions for context
        earnings = [c for c in salary_components_with_details if c.get('ComponentType') == 'Earning']
        deductions = [c for c in salary_components_with_details if c.get('ComponentType') == 'Deduction']
        
        # Calculate net salary
        net_salary = sum(c.get('Amount', 0) for c in earnings) - sum(c.get('Amount', 0) for c in deductions)

        # Get preferred template for job letter
        template_file = 'employee_job_letter.html' # Default
        try:
            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT "TemplateFile" 
                    FROM "TemplateSettings" 
                    WHERE "SchoolID" = %s AND "TemplateType" = 'EmployeeJobLetter' 
                    AND "IsActive" = TRUE AND "IsDeleted" = FALSE
                """, [school_id])
                preference_row = cursor.fetchone()
                if preference_row and preference_row[0]:
                    # Use the stored path
                    template_file = preference_row[0]
                    logger.info(f"Using preferred job letter template: {template_file}")
                else:
                    logger.info("Using default job letter template: employee_job_letter.html")
        except Exception as e:
            logger.error(f"Error fetching template preference: {e}")

        # Fetch Subject Specializations for Teachers
        specializations_text = ""
        if role_name and 'Teacher' in role_name:
            try:
                with connection.cursor() as cursor:
                    cursor.execute("""
                        SELECT string_agg(ssm."SpecializationName", ', ')
                        FROM "EmployeeSpecialization" es
                        JOIN "SubjectSpecializationMaster" ssm ON es."SpecializationID" = ssm."SpecializationID"
                        WHERE es."EmployeeID" = (SELECT "EmployeeID" FROM "EmployeeMaster" WHERE "EmployeeCode" = %s AND "SchoolID" = %s AND "IsDeleted" = FALSE)
                    """, [email_data.get('employee_code'), school_id])
                    row = cursor.fetchone()
                    if row and row[0]:
                        specializations_text = row[0]
            except Exception as e:
                logger.error(f"Error fetching specializations: {e}")

        # Generate Dynamic Greeting Message (consistent with Proc_EmployeeJobLetter_Get)
        greeting_message = f"We are delighted to welcome you to the team. We believe your commitment to excellence aligns perfectly with our vision."
        dept_name = "General"
        if role_name:
            if 'Teacher' in role_name:
                dept_name = "Academic"
                spec_suffix = f" (Specialization: {specializations_text})" if specializations_text else ""
                greeting_message = (
                    f"We are delighted to formally welcome you to {school_name}. Your appointment as {role_name}{spec_suffix} "
                    f"in the {dept_name.lower()} department reflects your skills and dedication. We are confident that your "
                    f"pedagogical expertise and commitment to student development will be a valuable asset to our academic community."
                )
            elif 'Accountant' in role_name:
                dept_name = "Accounts"
                greeting_message = (
                    f"We are delighted to formally welcome you to {school_name}. Your appointment as {role_name} "
                    f"in the finance department is confirmed. Your financial acumen and attention to detail will "
                    f"play a vital role in maintaining the integrity of our institution's financial operations."
                )
            elif 'Principal' in role_name or 'Director' in role_name:
                dept_name = "Leadership"
                greeting_message = (
                    f"It is our privilege to welcome you to the leadership team of {school_name}. Your appointment as {role_name} "
                    f"marks a new chapter in our pursuit of excellence. Your vision, leadership, and experience will be instrumental "
                    f"in guiding our institution toward greater success."
                )
            elif 'Admin' in role_name or 'Manager' in role_name:
                dept_name = "Administration"
                greeting_message = (
                    f"We are delighted to formally welcome you to {school_name}. Your appointment as {role_name} "
                    f"in the {dept_name.lower()} department is confirmed. We look forward to your support in strengthening "
                    f"administrative processes and ensuring the smooth functioning of our institution."
                )

        # Prepare login info for custom templates
        login_info = {
            "Title": "System Login Information",
            "Options": [
                {"Label": "Employee Code", "Value": email_data.get('employee_code')},
                {"Label": "Email Address", "Value": email_data.get('email')},
                {"Label": "Phone Number", "Value": email_data.get('mobile_no')}
            ],
            "Disclaimer": "Please keep these credentials secure and change your password after first login.",
            "OptionsHeader": "Login Options - You can use any of the following as your Username:"
        }

        # Prepare placeholders for email body
        placeholders = {
            'employee_name': email_data.get('employee_name'),
            'employee_code': email_data.get('employee_code'),
            'position': role_name,
            'date_of_joining': email_data.get('date_of_joining'),
            'email': email_data.get('email'),
            'mobile_no': email_data.get('mobile_no'),
            'username': email_data.get('username'),
            'password': email_data.get('password'),
            'salary_breakdown': salary_breakdown,
            'school_rules': rules_html,
            'school_name': school_name,
            'user_photo': employee_photo,
            'greeting_message': greeting_message,
            'specialization': specializations_text,
        }

        # Prepare attachment details for job letter PDF
        attachment_details = {
            'job_letter': {
                'filename': f"Job_Letter_{email_data.get('employee_code')}.pdf",
                'type': 'application/pdf',
                'template': template_file,
                'context': {
                    'employee': {
                        'employee_name': email_data.get('employee_name'),
                        'full_name': email_data.get('employee_name'), # for template compatibility
                        'employee_code': email_data.get('employee_code'),
                        'position': role_name,
                        'date_of_joining': email_data.get('date_of_joining'),
                        'email': email_data.get('email'),
                        'mobile_no': email_data.get('mobile_no'),
                        'username': email_data.get('username'),
                        'password': email_data.get('password'),
                        'school_name': school_name,
                        'school_logo': school_logo,
                        'user_photo': employee_photo,
                        'salary_components': salary_components_with_details,
                        'earnings': earnings,
                         'deductions': deductions,
                         'school_rules': school_rules if school_rules else school_terms_conditions,
                         'school_terms_conditions': school_terms_conditions,
                        'current_date': timezone.now().strftime('%d-%b-%Y'),
                        'net_salary': net_salary,
                        'greeting_message': greeting_message,
                        'subject_specialization': specializations_text
                    },
                    'login_info': login_info
                }
            }
        }
        
        # Queue email using database email tracking system
        email_task = EmailTrackingManager.create_email_task(
            email_code='EMPLOYEE_REGISTRATION_CONFIRMATION',
            to_email=email_data.get('email'),
            placeholders=placeholders,
            school_id=school_id,
            priority=5,  # Normal priority
            student_code=email_data.get('employee_code'),  # Store employee code in StudentCode column
            has_attachments=True,
            attachment_details=attachment_details
        )
        
        logger.info("Employee registration email queued successfully")
        
    except Exception as e:
        logger.error(f"Failed to queue employee registration email: {str(e)}", exc_info=True)

def send_admission_emails_async(email_data):
    """Send admission emails asynchronously in background thread"""
    def email_worker():
        try:
            logger.info("Background email worker started")
            
            # Prepare placeholders for emails
            placeholders_ack = {
                'student_name': email_data.get('student_name'),
                'student_code': email_data.get('student_code'),
                'admission_class': email_data.get('admission_class'),
                'admission_date': email_data.get('admission_date'),
                'school_name': email_data.get('school_name'),
            }
            placeholders_rcpt = {
                'student_name': email_data.get('student_name'),
                'student_code': email_data.get('student_code'),
                'receipt_number': email_data.get('payment_receipt', {}).get('receipt_number'),
                'amount_paid': email_data.get('payment_receipt', {}).get('amount_paid'),
                'payment_mode': email_data.get('payment_receipt', {}).get('payment_mode'),
                'payment_date': email_data.get('payment_receipt', {}).get('payment_date'),
                'school_name': email_data.get('school_name'),
            }

            # Try to generate PDFs and send emails with attachments
            try:
                # Generate PDFs
                logger.info("Background: Generating acknowledgment PDF...")
                # Get complete data from stored procedure
                ack_data = {}
                student_code = email_data.get('student_code')
                if student_code:
                    with connection.cursor() as cursor:
                        cursor.execute("SELECT * FROM proc_student_acknowledgment_get(%s)", [student_code])
                        row = cursor.fetchone()
                        if row:
                            columns = [col[0] for col in cursor.description]
                            ack_data = dict(zip(columns, row))
                            if ack_data.get('StudentCode'):
                                ack_data['student_code'] = ack_data['StudentCode']
                            if ack_data.get('school_logo'):
                                ack_data['school_logo'] = _bytes_to_data_uri(ack_data['school_logo'])
                        ack_data['instructions'] = []
                        if cursor.nextset():
                            ack_data['instructions'] = [{'title': r[0], 'text': r[1]} for r in cursor.fetchall()]
                        ack_data['documents'] = []
                        if cursor.nextset():
                            ack_data['documents'] = [{'type': r[0], 'name': r[1]} for r in cursor.fetchall()]
                        ack_data['fees'] = []
                        if cursor.nextset():
                            ack_data['fees'] = [{'name': r[0], 'amount': float(r[1] or 0), 'discount': float(r[2] or 0), 'final': float(r[3] or 0)} for r in cursor.fetchall()]
                ack_pdf = generate_pdf_from_template('admission_acknowledgment.html', {'acknowledgment': ack_data})
                
                logger.info("Background: Generating receipt PDF...")
                # Get selected template from database
                receipt_template = 'core/document_templates/payment_receipt/payment_success.html'
                school_id = email_data.get('school_id')
                if school_id:
                    with connection.cursor() as cursor:
                        cursor.execute('SELECT "TemplateFile" FROM "TemplateSettings" WHERE "SchoolID" = %s AND "TemplateType" = \'PaymentReceipt\' AND "IsActive" = TRUE AND "IsDeleted" = FALSE', [school_id])
                        row = cursor.fetchone()
                        if row and row[0]:
                            receipt_template = row[0]
                rcpt_pdf = generate_pdf_from_template(receipt_template, {'payment_receipt': email_data.get('payment_receipt')})

                # Send acknowledgment email with PDF
                logger.info("Background: Sending acknowledgment email...")
                send_email_by_code(
                    code='ADMISSION_ACKNOWLEDGMENT',
                    to_emails=email_data.get('email'),
                    placeholders=placeholders_ack,
                    school_id=email_data.get('school_id'),
                    attachments=[(f"Acknowledgment-{email_data.get('student_code')}.pdf", ack_pdf, 'application/pdf')]
                )
                
                # Send payment receipt email with PDF
                logger.info("Background: Sending receipt email...")
                send_email_by_code(
                    code='PAYMENT_RECEIPT',
                    to_emails=email_data.get('email'),
                    placeholders=placeholders_rcpt,
                    school_id=email_data.get('school_id'),
                    attachments=[(f"Receipt-{email_data.get('student_code')}.pdf", rcpt_pdf, 'application/pdf')]
                )
                
                logger.info("Background: Emails with PDFs sent successfully")
                
            except Exception as pdf_error:
                logger.warning(f"Background: PDF generation failed, sending emails without attachments: {str(pdf_error)}")
                
                # Send emails without PDF attachments as fallback
                send_email_by_code(
                    code='ADMISSION_ACKNOWLEDGMENT',
                    to_emails=email_data.get('email'),
                    placeholders=placeholders_ack,
                    school_id=email_data.get('school_id')
                )
                
                send_email_by_code(
                    code='PAYMENT_RECEIPT',
                    to_emails=email_data.get('email'),
                    placeholders=placeholders_rcpt,
                    school_id=email_data.get('school_id')
                )
                
                logger.info("Background: Emails sent without PDFs")
                
        except Exception as email_error:
            logger.error(f"Background: Email sending failed: {str(email_error)[:100]}")
    
    # Start the email worker in a separate thread
    email_thread = threading.Thread(target=email_worker, daemon=True)
    email_thread.start()
    logger.info("Background email thread started")


from .models import (
    ProfileMaster,
    UserMaster,
    SchoolMaster,
    MenuMaster,
    ProfileMenuMapping
)

# -----------------
# Config / constants
# -----------------
# --------------------------------------------------------------------------
# Core logic helpers (moved to utils.py)
# --------------------------------------------------------------------------
# Utility helpers
# -----------------
def _get_client_ip(request):
    x_forwarded = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded:
        return x_forwarded.split(',')[0].strip()
    return request.META.get('REMOTE_ADDR')

def safe_log_data(data_dict):
    """Remove sensitive fields before logging"""
    if not isinstance(data_dict, dict):
        return str(data_dict)[:200]
    sensitive_fields = ['password', 'token', 'aadhaar', 'ssn', 'credit_card', 'PasswordHash', 
                       'studentAadhaarNumber', 'fatherAadhaar', 'motherAadhaar', 'new_password',
                       'confirm_password', 'studentPassword', 'student_password']
    safe_dict = data_dict.copy()
    for field in sensitive_fields:
        if field in safe_dict:
            safe_dict[field] = '***REDACTED***'
    return safe_dict


def is_safe_url(url, allowed_hosts=None):
    """Validate redirect URL to prevent open redirect vulnerabilities (CWE-601)"""
    if not url:
        return False
    url = url.strip()
    # Allow relative URLs starting with /
    if url.startswith('/'):
        return not url.startswith('//')
    # Parse absolute URLs
    try:
        parsed = urlparse(url)
        # Reject URLs with schemes other than http/https
        if parsed.scheme and parsed.scheme not in ['http', 'https']:
            return False
        # If no host, it's relative - safe
        if not parsed.netloc:
            return True
        # Check against allowed hosts
        if allowed_hosts:
            return parsed.netloc in allowed_hosts
        return False
    except:
        return False

@custom_login_required
def timetable_view(request):
    return render(request, 'core/timetable.html')

# -----------------
# Custom session helpers (DB-backed)
# -----------------
def _create_custom_session(response, user_row, request, timeout_minutes=None, login_type=None):
    """
    Insert a row into user_sessions and set the session cookie.
    Accepts either:
      - tuple from SELECT in order: [UserID, FullName, ProfileID, ProfileName, SchoolID, SchoolName, SchoolLogo, UserPhoto]
      - dict with keys: user_id, user_name, profile_id, profile_name, school_id, school_name, school_logo, user_photo
    timeout_minutes: User's SessionTimeoutMinutes from UserMaster (None = no timeout, default 60)
    login_type: How the user logged in ('Password', 'OTP', 'FaceID')
    """
    token = secrets.token_urlsafe(48)

    try:
        # Our SELECT order: indices 0,1,2,3,4,5,6,7
        user_id = user_row[0]
        user_name = user_row[1]
        profile_id = user_row[2]
        profile_name = user_row[3]
        school_id = user_row[4]
        school_name = user_row[5]
        school_logo = user_row[6]
        user_photo = user_row[7]
    except Exception:
        if isinstance(user_row, dict):
            user_id = user_row.get('user_id')
            user_name = user_row.get('user_name')
            profile_id = user_row.get('profile_id')
            profile_name = user_row.get('profile_name')
            school_id = user_row.get('school_id')
            school_name = user_row.get('school_name')
            school_logo = user_row.get('school_logo')
            user_photo = user_row.get('user_photo')
        else:
            user_id = user_row[0] if user_row else None
            user_name = profile_id = profile_name = school_id = school_name = school_logo = user_photo = None

    ip = _get_client_ip(request)
    ua = request.META.get("HTTP_USER_AGENT", "")[:2000]
    
    # Calculate expiry based on user's timeout setting (default 60 minutes if not set)
    if timeout_minutes is None:
        timeout_minutes = 60  # Default 1 hour
    
    lifetime_seconds = timeout_minutes * 60
    
    # Default login type if not specified
    if login_type is None:
        login_type = 'By UserId and Password'  # Default assumption

    with connection.cursor() as cursor:
        cursor.execute("""
            INSERT INTO "user_sessions"
            ("user_id", "session_token", "profile_id", "profile_name", "school_id", "school_name", "ip_address", "user_agent", "created_at", "last_activity", "expires_at", "LoginType")
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP + (%s || ' second')::INTERVAL, %s)
        """, [user_id, token, profile_id, profile_name, school_id, school_name, ip, ua, lifetime_seconds, login_type])

    response.set_cookie(
        SESSION_COOKIE_NAME,
        token,
        max_age=lifetime_seconds,
        httponly=True,
        samesite='Lax',
        secure=False  # set True behind HTTPS
    )
    return token


def _touch_custom_session(token):
    if not token:
        return
    with connection.cursor() as cursor:
        cursor.execute("UPDATE \"user_sessions\" SET \"last_activity\" = CURRENT_TIMESTAMP WHERE \"session_token\" = %s", [token])
        logger.debug(f"Session touched for token: {token}")

def _destroy_custom_session_by_token(token):
    from django.db import transaction
    if not token:
        logger.warning("No token provided to destroy session")
        return
    
    logger.info(f"Attempting to update LogoutTime for token: {token}")
    
    with transaction.atomic():
        with connection.cursor() as cursor:
            # Check if session exists
            cursor.execute("SELECT \"session_token\", \"user_id\", \"LogoutTime\" FROM \"user_sessions\" WHERE \"session_token\" = %s", [token])
            result = cursor.fetchone()
            logger.info(f"Session found before update: {result}")
            
            # Update LogoutTime
            cursor.execute("UPDATE \"user_sessions\" SET \"LogoutTime\" = CURRENT_TIMESTAMP WHERE \"session_token\" = %s", [token])
            rows = cursor.rowcount
            logger.info(f"UPDATE affected {rows} rows")
            
            # Verify update
            cursor.execute("SELECT \"session_token\", \"user_id\", \"LogoutTime\" FROM \"user_sessions\" WHERE \"session_token\" = %s", [token])
            result_after = cursor.fetchone()
            logger.info(f"Session after update: {result_after}")

        rows_affected = cursor.rowcount
        logger.info(f"Session destroyed - Rows affected: {rows_affected}")

def _trigger_login_alert(request, user_row):
    """Send security notification on successful login"""
    try:
        from mail.utils import send_email_by_code
        
        # user_row indices depend on which query was used
        if isinstance(user_row, dict):
            u_name = user_row.get('user_name')
            email = user_row.get('email')
            profile = user_row.get('profile_name')
            s_name = user_row.get('school_name')
            s_logo = user_row.get('school_logo')
            s_id = user_row.get('school_id')
        else:
             # Assuming standard login query results (tuples vary slightly)
             # Index mapping for common login success tuples: 
             # 0:UserID, 1:UserName, 2:ProfileID, 3:ProfileName, 4:SchoolID, 5:SchoolName
             u_name = user_row[1]
             # For password login, we have more fields
             email = user_row[13] if len(user_row) > 13 else None
             profile = user_row[3]
             s_name = user_row[5]
             s_logo = user_row[7] if len(user_row) > 7 else None
             s_id = user_row[4]

        if not email:
            # Fallback fetch for email
            u_id = user_row[0] if not isinstance(user_row, dict) else user_row.get('user_id')
            with connection.cursor() as cursor:
                cursor.execute('SELECT "Email" FROM "UserMaster" WHERE "UserID" = %s', [u_id])
                res = cursor.fetchone()
                if res: email = res[0]

        if email:
            # Enforce ShikshaWave Branding for security emails
            # Optimization: Use static URL instead of dynamic Base64 to keep email size small (prevents clipping)
            shikshawave_logo = request.build_absolute_uri('/static/images/ShikshaWave_Logo.png')
            
            placeholders = {
                'user_name': u_name,
                'login_id': email,
                'profile': profile,
                'school_name': 'ShikshaWave',
                'school_logo': shikshawave_logo,
                'header_title': 'New Login Alert',
                'ip_address': _get_client_ip(request),
                'browser': request.META.get('HTTP_USER_AGENT', 'Unknown'),
                'timestamp': timezone.now().strftime('%Y-%m-%d %H:%M:%S')
            }
            send_email_by_code('NEW_LOGIN_ALERT', [email], placeholders=placeholders, school_id=s_id)
            logger.info(f"Login alert sent to {email}")
    except Exception as e:
        logger.error(f"Failed to trigger login alert: {e}")

# -----------------
# Decorators (strict permission)
# -----------------

def strict_permission_required(menu_id_or_url, action='view'):
    """
    Enforce ProfileMenuMapping permissions on a per-view basis.
    Usage: @strict_permission_required('/students', action='add') OR @strict_permission_required(12, 'edit')
    """
    action_col = {
        'view': 'CanView',
        'add': 'CanAdd',
        'edit': 'CanEdit',
        'delete': 'CanDelete'
    }.get(action, 'CanView')

    def decorator(view_func):
        @wraps(view_func)
        def _wrapped(request, *args, **kwargs):
            sess = _get_custom_session_info(request)
            if not sess:
                logger.warning("Access denied: No valid session")
                messages.error(request, "Please login to continue.")
                return redirect('login')

            profile_id = sess.get('profile_id')
            if profile_id == 1:
                request.custom_user = sess
                _touch_custom_session(sess.get('session_token'))
                return view_func(request, *args, **kwargs)

            with connection.cursor() as cursor:
                if isinstance(menu_id_or_url, int):
                    cursor.execute(f"""
                        SELECT 1
                        FROM "ProfileMenuMapping" pmm
                        JOIN "MenuMaster" m ON m."MenuID" = pmm."MenuID" AND m."IsActive" = TRUE AND m."IsDeleted" IS NOT TRUE
                        WHERE pmm."ProfileID" = %s AND pmm."{action_col}" = TRUE AND pmm."IsDeleted" IS NOT TRUE AND pmm."MenuID" = %s
                    """, [profile_id, menu_id_or_url])
                else:
                    cursor.execute(f"""
                        SELECT 1
                        FROM "ProfileMenuMapping" pmm
                        JOIN "MenuMaster" m ON m."MenuID" = pmm."MenuID" AND m."IsActive" = TRUE AND m."IsDeleted" IS NOT TRUE
                        WHERE pmm."ProfileID" = %s AND pmm."{action_col}" = TRUE AND pmm."IsDeleted" IS NOT TRUE 
                        AND (m."MenuURL" = %s OR m."MenuName" = %s)
                    """, [profile_id, menu_id_or_url, menu_id_or_url])

                allowed = cursor.fetchone()

            if not allowed:
                logger.warning(f"Permission denied for action: {action}")
                return HttpResponseForbidden("You do not have permission to access this resource.")

            request.custom_user = sess
            _touch_custom_session(sess.get('session_token'))
            return view_func(request, *args, **kwargs)
        return _wrapped
    return decorator

# -----------------
# Contact / Home
# -----------------
@require_POST
def contact_submit(request):
    email = request.POST.get('email', '').strip()
    message = request.POST.get('message', '').strip()
    ip_address = request.POST.get('ip_address') or _get_client_ip(request)
    browser_info = request.POST.get('browser_info') or request.META.get('HTTP_USER_AGENT', '')
    logger.info(f"Incoming contact submit request: {email}")
    location = request.POST.get('location') or ''

    if not email or not message:
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'status': 0, 'message': 'Email and message are required.'})
        messages.error(request, "Email and message are required.")
        return redirect('home')

    try:
        with connection.cursor() as cursor:
            cursor.execute("SELECT * FROM sp_InsertContactUs(%s, %s, %s, %s, %s)", 
                         [email, message, ip_address, browser_info, location])
            row = cursor.fetchone()
            if row:
                status = int(row[0] or 0)
                out_msg = row[1] or ''
                logger.info(f"DB Response: status={status}, msg={out_msg}")
            else:
                status = 0
                out_msg = 'No response from DB procedure.'
                logger.warning("DB Response: No row returned")
    except Exception as e:
        status = 0
        out_msg = str(e)
        logger.error(f"Contact submission failed: {str(e)}", exc_info=True)

    if status == 1:
        # Send Email Notifications
        try:
            placeholders = {
                'email': email,
                'message': message,
                'ip_address': ip_address,
                'browser_info': browser_info,
                'location': location
            }
            
            # 1. Notify Admin
            admin_email = 'shikshawaves@gmail.com'
            send_email_by_code(
                code='CONTACT_US_NOTIFICATION',
                to_emails=admin_email,
                placeholders=placeholders,
                is_async=True
            )
            
            # 2. Acknowledge User
            send_email_by_code(
                code='CONTACT_US_ACKNOWLEDGMENT',
                to_emails=email,
                placeholders=placeholders,
                is_async=True
            )
            logger.info(f"Contact submission emails queued for {email}")
            
        except Exception as ee:
            logger.error(f"Error triggering contact emails: {ee}")

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return JsonResponse({'status': status, 'message': out_msg})

    if status == 1:
        messages.success(request, out_msg)
    elif status == 2:
        messages.warning(request, out_msg)
    else:
        messages.error(request, out_msg)
    return redirect('home')

def home_view(request):
    return render(request, 'home.html')

# -----------------
# Login / OTP / Session endpoints
# -----------------
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def login_view(request):
    """
    Handle login via password or OTP. Redirects to dashboard on success.
    """
    session_info = _get_custom_session_info(request)
    if session_info:
        logger.info("User already logged in, redirecting to dashboard")
        return redirect('dashboard')
    else:
        logger.info("No valid session found, showing login page")

    if request.method == "POST":
        login_type = request.POST.get('login_type', 'password').strip().lower()
        identifier = (request.POST.get('identifier') or
                      request.POST.get('username') or
                      request.POST.get('user_code') or "").strip()
        next_url = request.POST.get('next', 'dashboard')
        # Validate redirect URL to prevent open redirects (CWE-601)
        if not is_safe_url(next_url):
            next_url = 'dashboard'
        logger.debug(f"Login attempt: type={login_type}")

        if not identifier:
            logger.warning("Login failed: Missing identifier")
            messages.error(request, "Username/UserCode is required.")
            return render(request, "core/login.html", {'next': next_url, 'identifier': identifier})

        try:
            if login_type == 'password':
                password = request.POST.get('password', '')
                if not password:
                    logger.warning("Login failed: Missing password")
                    messages.error(request, "Password is required.")
                    return render(request, "core/login.html", {'next': next_url, 'identifier': identifier})
                with connection.cursor() as cursor:
                    cursor.execute("""
                        SELECT 
                            u."UserID",
                            u."UserName" AS FullName,
                            u."ProfileID",
                            p."ProfileName",
                            u."SchoolID",
                            s."SchoolName",
                            u."UserPhoto",
                            s."SchoolLogo",
                            u."PasswordHash",
                            u."DarkTheme",
                            u."SessionTimeoutMinutes",
                            u."FailedLoginAttempts",
                            u."BlockedUntil",
                            u."Email"
                        FROM "UserMaster" u
                        INNER JOIN "ProfileMaster" p ON u."ProfileID" = p."ProfileID"
                        LEFT JOIN "SchoolMaster" s ON u."SchoolID" = s."SchoolID"
                        WHERE (u."UserName" = %s OR u."UserCode" = %s OR u."Email" = %s)
                          AND u."IsActive" = TRUE
                          AND u."IsDeleted" IS NOT TRUE
                    """, [identifier, identifier, identifier])
                    row = cursor.fetchone()
                    logger.debug(f"Password login query result: {row}")

                if row:
                    (db_userid, fullname, profileid, profilename, schoolid, schoolname, 
                     userphoto, schoollogo, stored_hash, dark_theme, timeout_minutes,
                     failed_attempts, blocked_until, user_email) = row

                    now = timezone.now()
                    show_reset_link = False

                    # 1. Check if currently blocked
                    if blocked_until and blocked_until > now:
                        logger.warning(f"Login blocked for {identifier} until {blocked_until}")
                        messages.error(request, "Your account is blocked for 24 hours. For urgent access, please contact the Admin.")
                        return render(request, "core/login.html", {'next': next_url})

                    # 2. Check if block just expired
                    if blocked_until and blocked_until <= now and failed_attempts >= 3:
                        messages.warning(request, "Your previous block has expired. We strongly recommend resetting your password if you've forgotten it.")
                        show_reset_link = True

                    # ✅ Check both hashing schemes
                    valid_password = False
                    try:
                        if stored_hash.startswith("pbkdf2_sha256$"):
                            valid_password = check_password(password, stored_hash)
                        else:
                            # Legacy SHA-256 hash support
                            valid_password = (hashlib.sha256(password.encode()).hexdigest() == stored_hash)
                    except Exception as e:
                        logger.error(f"Password check failed: {e}", exc_info=True)

                    if valid_password:
                        # Reset strikes on success
                        with connection.cursor() as cursor:
                            cursor.execute("""
                                UPDATE "UserMaster" SET "FailedLoginAttempts" = 0, "BlockedUntil" = NULL 
                                WHERE "UserID" = %s
                            """, [db_userid])

                        resp = redirect(next_url)
                        # ✅ Save user in session (exclude UserPhoto and SchoolLogo)
                        request.session['UserId'] = db_userid
                        request.session['UserName'] = fullname
                        request.session['ProfileID'] = profileid
                        request.session['ProfileName'] = profilename
                        request.session['SchoolID'] = schoolid
                        request.session['SchoolName'] = schoolname
                        request.session['dark_mode'] = dark_theme == 'Yes' if dark_theme else False
                        request.session['session_timeout_minutes'] = timeout_minutes if timeout_minutes is not None else 60

                        token = _create_custom_session(resp, row[:11], request, timeout_minutes=timeout_minutes, login_type='By UserId and Password')
                        _trigger_login_alert(request, row)
                        logger.info(f"User logged in successfully (password) with timeout: {timeout_minutes} minutes")
                        return resp
                    
                    else:
                        # Password failure logic
                        new_failed_attempts = failed_attempts + 1
                        new_blocked_until = None
                        msg = "Invalid username or password."
                        msg_type = "error"

                        if new_failed_attempts == 2:
                            msg = "Invalid credentials. If you've forgotten your password, you can reset it below."
                            show_reset_link = True
                        elif new_failed_attempts >= 3:
                            new_blocked_until = now + timezone.timedelta(hours=24)
                            msg = "Your account has been blocked for 24 hours due to multiple failed attempts. For urgent access, please contact the Admin."
                            show_reset_link = False # Requirement: "they will not get pop for reset password" when blocked

                            # Send block notification email
                            if user_email:
                                try:
                                    # Determine Branding
                                    branding_name = get_branding_title(profileid, schoolname)
                                    
                                    placeholders = {
                                        'user_name': fullname,
                                        'login_id': identifier,
                                        'timestamp': now.strftime('%Y-%m-%d %H:%M:%S'),
                                        'ip_address': request.META.get('REMOTE_ADDR', 'Unknown'),
                                        'browser': request.META.get('HTTP_USER_AGENT', 'Unknown'),
                                        'profile': profilename,
                                        'school_name': branding_name,
                                        'school_logo': None
                                    }
                                    send_email_by_code(
                                        'ACCOUNT_BLOCKED', 
                                        [user_email], 
                                        placeholders=placeholders,
                                        school_id=schoolid
                                    )
                                except Exception as ee:
                                    logger.error(f"Failed to send block notification: {ee}")

                        with connection.cursor() as cursor:
                            cursor.execute("""
                                UPDATE "UserMaster" 
                                SET "FailedLoginAttempts" = %s, 
                                    "BlockedUntil" = %s,
                                    "LastFailedLogin" = %s
                                WHERE "UserID" = %s
                            """, [new_failed_attempts, new_blocked_until, now, db_userid])
                        
                        if msg_type == "error":
                            messages.error(request, msg)
                        else:
                            messages.warning(request, msg)
                        
                        return render(request, "core/login.html", {
                            'next': next_url,
                            'show_reset_link': show_reset_link,
                            'identifier': identifier
                        })

                logger.warning("Invalid credentials")
                messages.error(request, "Invalid username or password.")

            elif login_type == 'otp':
                with connection.cursor() as cursor:
                    cursor.execute("""
                        SELECT u."UserID", u."Email", u."DarkTheme", u."SessionTimeoutMinutes"
                        FROM "UserMaster" u
                        WHERE (u."UserName" = %s OR u."UserCode" = %s OR u."Email" = %s)
                          AND u."IsActive" = TRUE
                          AND u."IsDeleted" IS NOT TRUE
                    """, [identifier, identifier, identifier])
                    exists = cursor.fetchone()
                    logger.debug(f"OTP account check: {exists}")

                if not exists:
                    logger.warning("Account not found")
                    messages.error(request, "Account not found.")
                    return render(request, "core/login.html", {'next': next_url, 'identifier': identifier})

                if not exists[1]:
                    logger.warning("No email configured")
                    messages.error(request, "No email address associated with this account.")
                    return render(request, "core/login.html", {'next': next_url, 'identifier': identifier})

                try:
                    generate_and_store_otp(identifier=identifier, purpose='login', request=request)
                    
                    # Clear any existing error messages before showing success
                    list(messages.get_messages(request))
                    
                    resp = redirect('verify_otp')
                    resp.set_cookie(OTP_COOKIE_NAME, identifier, max_age=OTP_COOKIE_MAX_AGE, samesite='Lax', secure=False, httponly=True)
                    # Validate next_url before setting cookie (CWE-601, CWE-614)
                    safe_next = next_url if is_safe_url(next_url) else 'dashboard'
                    resp.set_cookie(f"{OTP_COOKIE_NAME}_next", safe_next, max_age=OTP_COOKIE_MAX_AGE, samesite='Lax', secure=False, httponly=True)
                    
                    # Store preferences for OTP flow
                    if exists[2]:
                        request.session['dark_mode_temp'] = exists[2] == 'Yes'
                    if exists[3] is not None:
                        request.session['session_timeout_temp'] = exists[3]
                    
                    logger.info("OTP sent, redirecting to verify_otp")
                    messages.success(request, f"OTP sent to your email.")
                    return resp
                except Exception as e:
                    logger.error(f"OTP generation failed: {str(e)[:100]}", exc_info=True)
                    messages.error(request, f"Failed to send OTP: {str(e)}")

            elif login_type == 'faceid':
                # Redirect to secure face authentication
                from .face_auth_views import FaceAuthenticationView
                
                # Check if this is a secure face auth request
                if 'face_descriptor' in request.POST:
                    # Handle secure face authentication
                    try:
                        import json
                        face_descriptor_raw = request.POST.get('face_descriptor', '[]')
                        
                        face_descriptor = json.loads(face_descriptor_raw)
                        
                        logger.debug(f"Login view descriptors received: face_len={len(face_descriptor)}")
                        
                        if len(face_descriptor) != 128:
                            logger.warning(f"Invalid face descriptor length: {len(face_descriptor)}")
                            messages.error(request, f"Invalid face data ({len(face_descriptor)} dims). Please try again.")
                            return render(request, "core/login.html", {'next': next_url, 'identifier': identifier})
                        
                        from .face_recognition_service import FaceRecognitionService
                        face_service = FaceRecognitionService()
                        
                        auth_result = face_service.authenticate_face(identifier, face_descriptor, request)
                        
                        if auth_result and auth_result.get('success'):
                            # Authentication successful
                            user_data = auth_result['user_data']
                            similarity = auth_result['similarity']
                            
                            # Set session data
                            request.session['UserId'] = user_data[0]
                            request.session['UserID'] = user_data[0]
                            request.session['UserName'] = user_data[1]
                            request.session['ProfileId'] = user_data[2]
                            request.session['ProfileID'] = user_data[2]
                            request.session['ProfileName'] = user_data[3]
                            request.session['SchoolId'] = user_data[4]
                            request.session['SchoolID'] = user_data[4]
                            request.session['SchoolName'] = user_data[5]
                            request.session['dark_mode'] = user_data[8] == 'Yes' if user_data[8] else False
                            request.session['session_timeout_minutes'] = user_data[9] if user_data[9] is not None else 60
                            
                            resp = redirect(next_url)
                            _create_custom_session(resp, user_data, request, timeout_minutes=user_data[9], login_type='By FaceId')
                            _trigger_login_alert(request, user_data)
                            
                            logger.info(f"Secure face authentication successful for user {user_data[1]} (ID: {user_data[0]}) - Similarity: {similarity:.2f}%")
                            messages.success(request, f'Welcome back, {user_data[1]}! Face authentication successful ({similarity:.2f}% match).')
                            return resp
                        else:
                            logger.warning("Secure face authentication failed")
                            messages.error(request, "Face authentication failed. Please try again or use alternative login method.")
                            
                    except Exception as e:
                        logger.error(f"Secure face authentication error: {e}")
                        messages.error(request, "Face authentication service unavailable. Please try again.")
                else:
                    # Legacy face ID authentication (deprecated but maintained for compatibility)
                    face_photo = request.POST.get('face_photo', '')
                    match_percentage = request.POST.get('match_percentage', '0')
                    
                    if not identifier:
                        logger.warning("Face ID login failed: No identifier provided")
                        messages.error(request, "Please enter your email or username for Face ID authentication.")
                        return render(request, "core/login.html", {'next': next_url})
                    
                    if not face_photo:
                        logger.warning("Face ID login failed: No face photo provided")
                        messages.error(request, "Face ID authentication failed - no photo captured.")
                        return render(request, "core/login.html", {'next': next_url})
                    
                    try:
                        match_percentage = float(match_percentage)
                    except (ValueError, TypeError):
                        match_percentage = 0
                    
                    # Check if match percentage meets threshold (80%)
                    if match_percentage < 80:
                        logger.warning("Face ID login failed: Match percentage too low")
                        messages.error(request, f"Face match too low: {match_percentage:.1f}%. Need 80% or higher for authentication.")
                        return render(request, "core/login.html", {'next': next_url})
                    
                    # Find user by identifier
                    with connection.cursor() as cursor:
                        cursor.execute("""
                            SELECT 
                                u."UserID",
                                u."UserName" AS FullName,
                                u."ProfileID",
                                p."ProfileName",
                                u."SchoolID",
                                s."SchoolName",
                                u."UserPhoto",
                                s."SchoolLogo",
                                u."DarkTheme",
                                u."SessionTimeoutMinutes"
                            FROM "UserMaster" u
                            INNER JOIN "ProfileMaster" p ON u."ProfileID" = p."ProfileID"
                            LEFT JOIN "SchoolMaster" s ON u."SchoolID" = s."SchoolID"
                            WHERE (u."UserName" = %s OR u."UserCode" = %s OR u."Email" = %s)
                              AND u."IsActive" = TRUE
                              AND u."IsDeleted" IS NOT TRUE
                        """, [identifier, identifier, identifier])
                        
                        row = cursor.fetchone()
                        
                        if row:
                            # Create session for Face ID authenticated user
                            resp = redirect(next_url)
                            _create_custom_session(resp, row, request, timeout_minutes=row[9], login_type='By FaceId')
                            _trigger_login_alert(request, row)
                            
                            # Set session data
                            request.session['UserId'] = row[0]
                            request.session['UserName'] = row[1]
                            request.session['ProfileID'] = row[2]
                            request.session['ProfileName'] = row[3]
                            request.session['SchoolID'] = row[4]
                            request.session['SchoolName'] = row[5]
                            request.session['dark_mode'] = row[8] == 'Yes' if row[8] else False
                            request.session['session_timeout_minutes'] = row[9] if row[9] is not None else 60
                            
                            logger.info(f"Legacy Face ID login successful for user {row[1]} (ID: {row[0]}) - Match: {match_percentage:.1f}%")
                            messages.success(request, f"Welcome back, {row[1]}! Face ID authentication successful ({match_percentage:.1f}% match).")
                            return resp
                        else:
                            logger.warning("Face ID login failed: User not found")
                            messages.error(request, "User not found. Please check your email or username.")
                
                return render(request, "core/login.html", {'next': next_url})

            else:
                logger.warning(f"Unsupported login type: {login_type}")
                messages.error(request, "Unsupported login type.")
        except Exception as e:
            logger.error(f"Login failed: {str(e)}", exc_info=True)
            messages.error(request, f"Login service unavailable: {str(e)}")

    logger.debug(f"Rendering login page, next={request.GET.get('next', 'dashboard')}")
    return render(request, "core/login.html", {
        'next': request.GET.get('next', 'dashboard')
    })


def verify_otp_view(request):
    """
    Verify OTP and create session on success.
    """
    identifier = request.COOKIES.get(OTP_COOKIE_NAME)
    next_url = request.COOKIES.get(f"{OTP_COOKIE_NAME}_next", 'dashboard')
    # Validate redirect URL (CWE-601)
    if not is_safe_url(next_url):
        next_url = 'dashboard'

    if not identifier:
        logger.warning("OTP verification failed: No identifier cookie")
        messages.error(request, "Session expired. Please log in again.")
        return redirect('login')

    if request.method == "POST":
        entered_otp = request.POST.get('otp', '').strip()
        if not entered_otp:
            logger.warning("OTP verification failed: Missing OTP")
            messages.error(request, "OTP is required.")
            return render(request, "core/verify_otp.html", {'next': next_url})

        try:
            is_valid, user_data = verify_otp(identifier, entered_otp, purpose='login')
            if is_valid:
                # Ensure we have complete user_data
                if not isinstance(user_data, dict) or not user_data.get('user_id'):
                    logger.warning(f"Invalid user_data from verify_otp: {user_data}")
                    with connection.cursor() as cursor:
                        cursor.execute("""
                            SELECT 
                                u."UserID",
                                u."UserName" AS FullName,
                                u."ProfileID",
                                p."ProfileName",
                                u."SchoolID",
                                s."SchoolName",
                                s."SchoolLogo",
                                u."UserPhoto",
                                u."DarkTheme",
                                u."SessionTimeoutMinutes"
                            FROM "UserMaster" u
                            INNER JOIN "ProfileMaster" p ON u."ProfileID" = p."ProfileID"
                            LEFT JOIN "SchoolMaster" s ON u."SchoolID" = s."SchoolID"
                            WHERE (u."UserName" = %s OR u."UserCode" = %s OR u."Email" = %s)
                              AND u."IsActive" = TRUE
                              AND u."IsDeleted" IS NOT TRUE
                        """, [identifier, identifier, identifier])
                        row = cursor.fetchone()
                        logger.debug(f"OTP fallback query result: {row}")

                    if not row:
                        logger.error(f"Account not found for {identifier}")
                        messages.error(request, "Account not found.")
                        return redirect('login')

                    # Extract user data including dark theme preference
                    db_userid, fullname, profileid, profilename, schoolid, schoolname, schoollogo, userphoto, dark_theme, timeout_minutes = row
                    
                    # Set session variables
                    request.session['UserId'] = db_userid
                    request.session['UserName'] = fullname
                    request.session['ProfileID'] = profileid
                    request.session['ProfileName'] = profilename
                    request.session['SchoolID'] = schoolid
                    request.session['SchoolName'] = schoolname
                    request.session['dark_mode'] = dark_theme == 'Yes' if dark_theme else False
                    request.session['session_timeout_minutes'] = timeout_minutes if timeout_minutes is not None else 60
                    
                    resp = redirect(next_url)
                    token = _create_custom_session(resp, row, request, timeout_minutes=timeout_minutes, login_type='By OTP')  # ? session includes user_id (created_by)
                    _trigger_login_alert(request, row)
                else:
                    # If user_data is already a dict with complete information
                    user_id = user_data.get('user_id')
                    
                    # Load dark mode preference and timeout from database
                    with connection.cursor() as cursor:
                        cursor.execute(
                            "SELECT \"DarkTheme\", \"SessionTimeoutMinutes\" FROM \"UserMaster\" WHERE \"UserID\" = %s",
                            [user_id]
                        )
                        result = cursor.fetchone()
                        dark_theme = result[0] if result else 'No'
                        timeout_minutes = result[1] if result and len(result) > 1 else 60
                    
                    # Set dark mode preference and timeout in session
                    request.session['dark_mode'] = dark_theme == 'Yes' if dark_theme else False
                    request.session['session_timeout_minutes'] = timeout_minutes if timeout_minutes is not None else 60
                    
                    resp = redirect(next_url)
                    token = _create_custom_session(resp, user_data, request, timeout_minutes=timeout_minutes, login_type='By OTP')  # ? also includes user_id
                    _trigger_login_alert(request, user_data)

                logger.info(f"User {identifier} logged in successfully (OTP), session_token={token}")
                resp.delete_cookie(OTP_COOKIE_NAME)
                resp.delete_cookie(f"{OTP_COOKIE_NAME}_next")
                return resp

            logger.warning(f"Invalid OTP for {identifier}")
            messages.error(request, "Invalid or expired OTP.")
        except Exception as e:
            logger.error(f"OTP verification failed: {str(e)}", exc_info=True)
            messages.error(request, f"Verification service unavailable: {str(e)}")

    logger.debug(f"Rendering OTP verification page, next={next_url}")
    return render(request, "core/verify_otp.html", {'next': next_url})


# -----------------
# Dashboard (role + menu aware)
# -----------------
# Dashboard views moved to dashboard_views.py for better code organization

def _fetch_user_menus(profile_id):
    """
    Returns {'flat': [...], 'tree': [...]}
    Each item: {'id','name','url','icon','parent_id','order','can_add','can_edit','can_delete'}
    Only menus with CanView=1 and active/not deleted are returned.
    """
    with connection.cursor() as cursor:
        cursor.execute('SELECT * FROM "Proc_User_Menu_List_Get"(%s)', [profile_id])
        rows = cursor.fetchall()
        logger.debug(f"Fetched menus for profile {profile_id}: {rows}")

    flat = []
    by_parent = {}
    for r in rows:
        item = {
            'id': r[0],
            'name': r[1],
            'url': r[2] or '#',
            'icon': r[3] or '',
            'parent_id': r[4],
            'order': r[5],
            'can_add': bool(r[6]),
            'can_edit': bool(r[7]),
            'can_delete': bool(r[8]),
        }
        flat.append(item)
        by_parent.setdefault(item['parent_id'], []).append(item)

    def build_tree(parent_id=None):
        children = by_parent.get(parent_id, [])
        children.sort(key=lambda x: (x['order'], x['name']))
        for c in children:
            c['children'] = build_tree(c['id'])
        return children

    tree = build_tree(None)
    logger.debug(f"Menu tree for profile {profile_id}: {tree}")
    return {'flat': flat, 'tree': tree}

# -----------------
# Add user (strictly enforce permissions if logged-in)
# -----------------
def add_user_view(request):
    """
    Behavior:
      - If a user is logged in, enforce role-based create permission (profile_id in [1,2]).
      - If no user is logged in, allow public creation (for first install).
    """
    current = _get_custom_session_info(request)  # may be None
    if current:
        if current.get('profile_id') not in [1, 2]:
            logger.warning(f"Unauthorized access attempt by profile {current.get('profile_id')}")
            return HttpResponseForbidden("You don't have permission to create users")

    if request.method == "POST":
        try:
            next_url = request.POST.get('next', 'dashboard')
            # Validate redirect URL (CWE-601)
            if not is_safe_url(next_url):
                next_url = 'dashboard'
            
            form_data = {
                'user_code': request.POST.get("user_code", "").strip(),
                'user_name': request.POST.get("user_name", "").strip(),
                'password': request.POST.get("password", "").strip(),
                'confirm_password': request.POST.get("confirm_password", "").strip(),
                'email': request.POST.get("email", "").strip().lower(),
                'phone': request.POST.get("phone", "").strip(),
                'profile_id': safe_int(request.POST.get("profile_id", 0)),
                'school_id': request.POST.get("school_id", "").strip() or None,
                'next': next_url
            }

            errors = []
            if not all([form_data['user_code'], form_data['user_name'], form_data['password'], form_data['confirm_password']]):
                errors.append("All required fields must be filled.")
            if form_data['password'] != form_data['confirm_password']:
                errors.append("Passwords do not match.")
            if len(form_data['password']) < 8:
                errors.append("Password must be at least 8 characters long.")
            if form_data['profile_id'] == 0:
                errors.append("Please select a profile type.")
            if form_data['profile_id'] in [2, 3, 4, 5] and not form_data['school_id']:
                errors.append("School ID is required for this profile type.")

            if errors:
                for error in errors:
                    messages.error(request, error)
                return render(request, "core/add_user.html", {'form_data': form_data})

            password_hash = make_password(form_data['password'])

            if form_data['profile_id'] == 1:
                form_data['school_id'] = None

            created_by = current.get('user_id') if current else None

            with transaction.atomic():
                with connection.cursor() as cursor:
                    cursor.execute("""
                        INSERT INTO UserMaster 
                        (UserCode, UserName, PasswordHash, Email, Phone, ProfileID, SchoolID, CreatedBy)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                    """, [
                        form_data['user_code'],
                        form_data['user_name'],
                        password_hash,
                        form_data['email'],
                        form_data['phone'],
                        form_data['profile_id'],
                        form_data['school_id'],
                        created_by
                    ])

            messages.success(request, f"User {form_data['user_name']} created successfully!")
            logger.info(f"User {form_data['user_name']} created by {created_by or 'anonymous'}")
            return redirect(form_data['next'])

        except IntegrityError as e:
            logger.error(f"User creation integrity error: {str(e)}")
            messages.error(request, "User already exists or invalid data")
        except Exception as e:
            logger.error(f"Unexpected error during user creation: {str(e)}", exc_info=True)
            messages.error(request, "An error occurred while creating the user")

    return render(request, "core/add_user.html", {'next': request.GET.get('next', 'dashboard')})

# -----------------

# AJAX: resend OTP
# -----------------
@require_POST
def resend_otp(request):
    identifier = request.POST.get('identifier') or request.POST.get('username') or request.POST.get('user_code')
    if not identifier:
        logger.warning("Resend OTP failed: Missing identifier")
        return JsonResponse({'success': False, 'message': 'Identifier required'})

    try:
        generate_and_store_otp(identifier=identifier, purpose='login', request=request)
        
        # Clear any existing error messages after successful OTP generation
        list(messages.get_messages(request))
        
        logger.info(f"OTP resent for {identifier}")
        return JsonResponse({'success': True})
    except Exception as e:
        logger.error(f"OTP resend error: {str(e)}", exc_info=True)
        return JsonResponse({'success': False, 'message': str(e)})

def placeholder_view(request):
    return HttpResponse("<h1>Coming Soon</h1><p>This page is under construction.</p>")

# -----------------
# Example: Settings (requires login)
# -----------------
@custom_login_required
def update_school_settings(request):
    if request.method == 'POST':
        try:
            school_name = request.POST.get('school_name')
            school_logo = request.FILES.get('school_logo')
            theme_color = request.POST.get('theme_color', '#4361ee')

            with connection.cursor() as cursor:
                cursor.execute("""
                    UPDATE SchoolSettings 
                    SET school_name = %s, 
                        theme_color = %s
                    WHERE id = 1
                """, [school_name, theme_color])

                if school_logo:
                    fs = FileSystemStorage(location='media/school/')
                    filename = fs.save(school_logo.name, school_logo)
                    cursor.execute("""
                        UPDATE SchoolSettings 
                        SET logo_path = %s 
                        WHERE id = 1
                    """, [filename])

            messages.success(request, "Settings updated successfully!")
            logger.info("School settings updated")
            return redirect('dashboard')

        except Exception as e:
            logger.error(f"Failed to update settings: {str(e)}", exc_info=True)
            messages.error(request, "Failed to update settings")

    return redirect('dashboard')

# Logout
# -----------------
def logout_view(request):
    token = request.COOKIES.get(SESSION_COOKIE_NAME)
    logger.info(f"Logout request - Token found: {bool(token)}")
    
    request.session.flush()
    
    if token:
        _destroy_custom_session_by_token(token)
        logger.info("Session destroyed from database")

    resp = redirect('login')
    resp.delete_cookie(SESSION_COOKIE_NAME)
    logger.info("Session cookie deleted from response")

    # ✅ prevent browser from caching after logout
    resp['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    resp['Pragma'] = 'no-cache'
    resp['Expires'] = '0'

    # messages.info(request, "You have been logged out successfully")  # Removed logout message
    logger.info("User logged out successfully")
    return resp

#---Dark Mode Toggle
@custom_login_required
def set_dark_mode(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            request.session['dark_mode'] = data.get('dark_mode', False)
            return JsonResponse({'status': 'success'})
        except json.JSONDecodeError:
            return JsonResponse({'status': 'error'}, status=400)
    return JsonResponse({'status': 'error'}, status=405)

@custom_login_required
# User management views moved to user_views.py


# User Detail View

@custom_login_required
def schools_create(request):

    if request.method == 'POST':
        # Extract form data
        school_name = request.POST.get('school_name')
        registration_number = request.POST.get('registration_number') or None
        address = request.POST.get('address') or None
        district = request.POST.get('district')
        state = request.POST.get('state')
        country = request.POST.get('country')
        pincode = request.POST.get('pincode') or None
        phone = request.POST.get('phone') or None
        email = request.POST.get('email') or None
        website = request.POST.get('website') or None
        created_by = request.session.get('UserId')
        
        logo_data = None
        if 'logo' in request.FILES:
            file = request.FILES['logo']
            is_valid, message = validate_uploaded_file(file, ALLOWED_IMAGE_TYPES)
            if not is_valid:
                messages.error(request, message)
                return render(request, 'create_school.html', get_context(request))
            logo_data = file.read()
        
        board_id = request.POST.get('board') or None
        medium_id = request.POST.get('medium') or None
        principal_name = request.POST.get('principal_name') or None
        principal_contact_mail = request.POST.get('principal_contact_mail') or None
        principal_contact_phone = request.POST.get('principal_contact_phone') or None
        director_name = request.POST.get('director_name') or None
        director_contact_phone = request.POST.get('director_contact_phone') or None
        director_contact_email = request.POST.get('director_contact_email') or None
        establish_date = request.POST.get('establish_date') or None

        # Validate created_by
        if not created_by:
            logger.warning("No UserId found in session for school creation")
            messages.error(request, "Session expired. Please login again.")
            return redirect('login')

        # Validate school name
        if not school_name or len(school_name) > 255:
            logger.warning(f"Invalid school name: {school_name}")
            messages.error(request, 'School name is required and must be 255 characters or less.')
            return render(request, 'create_school.html', get_context(request))

        # Convert district, state, country, board_id, medium_id to integers
        try:
            district = int(district) if district else None
            state = int(state) if state else None
            country = int(country) if country else None
            board_id = int(board_id) if board_id else None
            medium_id = int(medium_id) if medium_id else None
        except (ValueError, TypeError) as e:
            logger.warning(f"Invalid geographic/board/medium selection: {str(e)}")
            messages.error(request, 'Invalid geographic, board, or medium selection.')
            return render(request, 'create_school.html', get_context(request))

        # Validate phone numbers
        phone_fields = [
            ('phone', phone),
            ('principal_contact_phone', principal_contact_phone),
            ('director_contact_phone', director_contact_phone)
        ]
        for field_name, value in phone_fields:
            if value:
                # Remove dashes and check if it's exactly 10 digits
                clean_value = value.replace('-', '').replace(' ', '')
                if not (clean_value.isdigit() and len(clean_value) == 10):
                    logger.warning(f"Invalid {field_name}: {value}")
                    messages.error(request, f'{field_name.replace("_", " ").title()} must be exactly 10 digits.')
                    return render(request, 'create_school.html', get_context(request))
                # Store the clean value for database
                if field_name == 'phone':
                    phone = clean_value
                elif field_name == 'principal_contact_phone':
                    principal_contact_phone = clean_value
                elif field_name == 'director_contact_phone':
                    director_contact_phone = clean_value

        # Validate emails
        email_regex = r'^[^\s@]+@[^\s@]+\.[^\s@]+$'
        email_fields = [
            ('email', email),
            ('principal_contact_mail', principal_contact_mail),
            ('director_contact_email', director_contact_email)
        ]
        for field_name, value in email_fields:
            if value and not re.match(email_regex, value):
                logger.warning(f"Invalid {field_name}: {value}")
                messages.error(request, f'{field_name.replace("_", " ").title()} must be a valid email address.')
                return render(request, 'create_school.html', get_context(request))

        # Validate pincode
        if pincode and not (pincode.isdigit() and len(pincode) == 6):
            logger.warning(f"Invalid pincode: {pincode}")
            messages.error(request, 'Pincode must be exactly 6 digits.')
            return render(request, 'create_school.html', get_context(request))

        # Validate website
        if website:
            try:
                parsed_url = urlparse(website)
                if not all([parsed_url.scheme, parsed_url.netloc]):
                    raise ValueError("Invalid URL")
            except:
                logger.warning(f"Invalid website URL: {website}")
                messages.error(request, 'Please enter a valid URL (e.g., https://example.com).')
                return render(request, 'create_school.html', get_context(request))

        # Validate establish date
        if establish_date:
            try:
                input_date = safe_strptime(establish_date, '%Y-%m-%d').date()
                if input_date > datetime.now().date():
                    logger.warning(f"Future establish date: {establish_date}")
                    messages.error(request, 'Establish date cannot be in the future.')
                    return render(request, 'create_school.html', get_context(request))
            except ValueError as e:
                logger.warning(f"Invalid establish date format: {establish_date}")
                messages.error(request, 'Invalid establish date format.')
                return render(request, 'create_school.html', get_context(request))

        # Check for duplicate school
        with connection.cursor() as cursor:
            query = """
                SELECT COUNT(*) FROM "SchoolMaster" 
                WHERE ("SchoolName" = %s OR "RegistrationNumber" = %s) AND "IsDeleted" = FALSE
            """
            params = [school_name, registration_number or '']
            cursor.execute(query, params)
            count = cursor.fetchone()[0]
            if count > 0:
                logger.warning(f"Duplicate school detected: {school_name}, {registration_number}")
                messages.error(request, 'A school with this name or registration number already exists.')
                return render(request, 'create_school.html', get_context(request))

        # Call stored procedure
        try:
            with connection.cursor() as cursor:
                cursor.execute(
                    """
                    SELECT * FROM "Proc_CreateSchool_Set"(
                        %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                        %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s
                    )
                    """,
                    [
                        school_name,
                        registration_number,
                        address,
                        district,
                        state,
                        country,
                        pincode,
                        phone,
                        email,
                        website,
                        created_by,
                        logo_data,
                        board_id,
                        medium_id,
                        principal_name,
                        principal_contact_mail,
                        principal_contact_phone,
                        director_name,
                        director_contact_phone,
                        director_contact_email,
                        establish_date,
                    ]
                )
                result = cursor.fetchone()
                if result:
                    school_code, status = result
                else:
                    school_code, status = None, 'Error: No response from stored procedure.'
                    logger.error("No response from Proc_CreateSchool_Set")
        except Exception as e:
            logger.error(f"Error executing stored procedure: {str(e)}")
            messages.error(request, f'Error creating school: {str(e)}')
            return render(request, 'create_school.html', get_context(request))

        # Handle the result
        if school_code is not None:
            logger.info(f"School created successfully: {school_code}, {status}")
            messages.success(request, f"School '{school_name}' created successfully: {status}")
            return redirect('dashboard')
        else:
            logger.error(f"Failed to create school: {status}")
            messages.error(request, status)
            return render(request, 'create_school.html', get_context(request))

    # GET request
    context = get_context(request)
    return render(request, 'create_school.html', context)


def api_boards(request):
    """Get all boards - Public API"""
    with connection.cursor() as cursor:
        cursor.execute('SELECT "BoardID", "BoardName" FROM "Board_Master" WHERE "IsDeleted" = FALSE ORDER BY "BoardName"')
        boards = [{'id': row[0], 'name': row[1]} for row in cursor.fetchall()]
    return JsonResponse(boards, safe=False)

def api_mediums(request):
    """Get all mediums - Public API"""
    with connection.cursor() as cursor:
        cursor.execute('SELECT "MediumID", "MediumName" FROM "Medium_Master" WHERE "IsDeleted" = FALSE ORDER BY "MediumName"')
        mediums = [{'id': row[0], 'name': row[1]} for row in cursor.fetchall()]
    return JsonResponse(mediums, safe=False)


# Universal Geography API Views
@cache_page(60 * 15)  # Cache for 15 minutes
def api_countries(request):
    """Get all countries"""
    with connection.cursor() as cursor:
        cursor.execute(
            'SELECT "Geog_Id", "Geog_Name" FROM "Geographical_Master" '
            'WHERE "Geog_Type" = \'Country\' AND "IsDeleted" = FALSE '
            'ORDER BY CASE WHEN "Geog_Name" = \'India\' THEN 0 ELSE 1 END, "Geog_Name"'
        )
        countries = [{'id': row[0], 'name': row[1]} for row in cursor.fetchall()]
    return JsonResponse(countries, safe=False)

@cache_page(60 * 15)  # Cache for 15 minutes
def api_states(request):
    """Get states by country ID"""
    country_id = request.GET.get('country_id')
    if not country_id:
        return JsonResponse({'error': 'Country ID is required'}, status=400)
    
    with connection.cursor() as cursor:
        cursor.execute(
            'SELECT "Geog_Id", "Geog_Name" FROM "Geographical_Master" '
            'WHERE "Geog_Type" = \'State\' AND "Geog_Parent_Id" = %s AND "IsDeleted" = FALSE '
            'ORDER BY "Geog_Name"',
            [country_id]
        )
        states = [{'id': row[0], 'name': row[1]} for row in cursor.fetchall()]
    return JsonResponse(states, safe=False)

@cache_page(60 * 15)  # Cache for 15 minutes
def api_districts(request):
    """Get districts by state ID"""
    state_id = request.GET.get('state_id')
    if not state_id:
        return JsonResponse({'error': 'State ID is required'}, status=400)
    
    with connection.cursor() as cursor:
        cursor.execute(
            'SELECT "Geog_Id", "Geog_Name" FROM "Geographical_Master" '
            'WHERE "Geog_Type" = \'District\' AND "Geog_Parent_Id" = %s AND "IsDeleted" = FALSE '
            'ORDER BY "Geog_Name"',
            [state_id]
        )
        districts = [{'id': row[0], 'name': row[1]} for row in cursor.fetchall()]
    return JsonResponse(districts, safe=False)


# API endpoints moved to student_promote_views.py

# API endpoint moved to dashboard_views.py

@require_POST

# -- Aadhaar duplicate check API
@require_POST
@custom_login_required
def check_aadhaar_duplicate(request):
    """Check if a given Aadhaar number already exists for any student.
    Expects JSON body: { "aadhaar": "XXXXXXXXXXXX" }
    Returns: { exists: bool, message: str }
    """
    try:
        body = json.loads(request.body.decode('utf-8')) if request.body else {}
        aadhaar = (body.get('aadhaar') or '').strip()
        if not aadhaar:
            return JsonResponse({
                'exists': False,
                'message': 'Aadhaar not provided'
            }, status=400)

        # Ensure only digits
        aadhaar_digits = ''.join(ch for ch in aadhaar if ch.isdigit())
        if len(aadhaar_digits) != 12:
            return JsonResponse({
                'exists': False,
                'message': 'Invalid Aadhaar number'
            }, status=400)

        exists = False
        with connection.cursor() as cursor:
            # Try Student table first (expected schema)
            try:
                cursor.execute(
                    'SELECT COUNT(1) FROM "StudentMaster" WHERE COALESCE("IsDeleted", FALSE) = FALSE AND "StudentAadhaar" = %s',
                    [aadhaar_digits]
                )
                count = cursor.fetchone()[0]
                exists = count > 0
            except Exception:
                # Fallback: check in UserMaster if Student schema differs
                try:
                    cursor.execute(
                        'SELECT COUNT(1) FROM "UserMaster" WHERE COALESCE("IsDeleted", FALSE) = FALSE AND "Aadhaar" = %s',
                        [aadhaar_digits]
                    )
                    count = cursor.fetchone()[0]
                    exists = count > 0
                except Exception:
                    exists = False

        if exists:
            return JsonResponse({
                'exists': True,
                'message': 'A student with this Aadhaar already exists.'
            })
        return JsonResponse({
            'exists': False,
            'message': 'Aadhaar is available.'
        })
    except json.JSONDecodeError:
        return JsonResponse({
            'exists': False,
            'message': 'Invalid request payload'
        }, status=400)
    except Exception as e:
        logger.error(f"Aadhaar duplicate check failed: {str(e)}", exc_info=True)
        return JsonResponse({
            'exists': False,
            'message': 'Service unavailable'
        }, status=500)

#--School List View

@custom_login_required
def schools_list(request):
    # Get user context for header
    context = get_context(request)
    
    schools = []
    page_number = safe_int(request.GET.get('page', 1))
    page_size = safe_int(request.GET.get('per_page', 10))  # Support configurable page size

    # Search and filter parameters
    search_params = {
        'school_code': request.GET.get('school_code', ''),
        'school_name': request.GET.get('school_name', ''),
        'registration_number': request.GET.get('registration_number', ''),
        'created_at': request.GET.get('created_at', ''),
        'to_date': request.GET.get('to_date', ''),
        'phone': request.GET.get('phone', ''),
        'email': request.GET.get('email', ''),
        'principal_name': request.GET.get('principal_name', ''),
        'director_name': request.GET.get('director_name', ''),
        'board': request.GET.get('board', ''),
        'medium': request.GET.get('medium', ''),
        'status': request.GET.get('status', ''),
        'country': request.GET.get('country', ''),
        'state': request.GET.get('state', ''),
        'district': request.GET.get('district', ''),
        'pincode': request.GET.get('pincode', ''),
        'show_deleted': request.GET.get('show_deleted', '0'),
    }

    # Convert string IDs to integers where applicable
    try:
        board = int(search_params['board']) if search_params['board'] else None
        medium = int(search_params['medium']) if search_params['medium'] else None
        country = int(search_params['country']) if search_params['country'] else None
        state = int(search_params['state']) if search_params['state'] else None
        district = int(search_params['district']) if search_params['district'] else None
    except ValueError:
        board = medium = country = state = district = None

    # Convert date strings to datetime objects
    from_date = None
    to_date = None
    try:
        if search_params['created_at']:
            from_date = safe_strptime(search_params['created_at'], '%Y-%m-%d').date()
        if search_params['to_date']:
            to_date = safe_strptime(search_params['to_date'], '%Y-%m-%d').date()
    except ValueError:
        search_params['created_at'] = ''
        search_params['to_date'] = ''

    sort_column = request.GET.get('sort_column', 'SchoolID')
    sort_direction = request.GET.get('sort_direction', 'DESC')

    # Get session data for filtering
    user_id = request.session.get('UserId')
    school_id = request.session.get('SchoolID')

    # If no userId, force login
    if not user_id:
        messages.error(request, "Session expired. Please login again.")
        return redirect('login')

    # Fetch profile_id if not in session
    profile_id = request.session.get('ProfileID')
    if not profile_id:
        try:
            with connection.cursor() as cursor:
                cursor.execute(
                    "SELECT \"ProfileID\" FROM \"UserMaster\" WHERE \"UserID\" = %s AND \"IsDeleted\" = FALSE",
                    [user_id]
                )
                result = cursor.fetchone()
                if result:
                    profile_id = result[0]
                    request.session['ProfileID'] = profile_id
                else:
                    messages.error(request, "Invalid user data.")
                    return redirect('login')
        except Exception as e:
            logger.error(f"Error fetching user profile: {e}")
            messages.error(request, "Error fetching user data.")
            return redirect('login')

    # Fetch user photo
    try:
        with connection.cursor() as cursor:
            cursor.execute(
                "SELECT \"UserPhoto\" FROM \"UserMaster\" WHERE \"UserID\" = %s AND \"IsDeleted\" = FALSE",
                [user_id]
            )
            photo_data = cursor.fetchone()
            if photo_data and photo_data[0]:
                user_photo_src = f"data:image/jpeg;base64,{base64.b64encode(photo_data[0]).decode('utf-8')}"
    except Exception as e:
        logger.error(f"Error fetching user photo: {e}")

    # Fetch school logo if exists
    if school_id:
        try:
            with connection.cursor() as cursor:
                cursor.execute(
                    "SELECT SchoolLogo FROM SchoolMaster WHERE SchoolID = %s AND IsDeleted = FALSE",
                    [school_id]
                )
                logo_data = cursor.fetchone()
                if logo_data and logo_data[0]:
                    school_logo_src = f"data:image/jpeg;base64,{base64.b64encode(logo_data[0]).decode('utf-8')}"
        except Exception as e:
            logger.error(f"Error fetching school logo: {e}")

    # Fetch board and medium options
    boards = []
    mediums = []
    try:
        with connection.cursor() as cursor:
            cursor.execute("SELECT BoardID, BoardName FROM Board_Master WHERE IsDeleted = FALSE")
            boards = [{'BoardID': row[0], 'BoardName': row[1]} for row in cursor.fetchall()]
            cursor.execute("SELECT MediumID, MediumName FROM Medium_Master WHERE IsDeleted = FALSE")
            mediums = [{'MediumID': row[0], 'MediumName': row[1]} for row in cursor.fetchall()]
    except Exception as e:
        logger.error(f"Error fetching boards or mediums: {e}")

    # Fetch schools
    total_schools = 0
    active_schools = 0
    deleted_schools = 0
    try:
        with connection.cursor() as cursor:
            cursor.execute(
                """
                EXEC Proc_SchoolList_get 
                    @SearchCreatedAt = %s,
                    @SearchToDate = %s,
                    @SearchSchoolCode = %s,
                    @SearchSchoolName = %s,
                    @SearchCountry = %s,
                    @SearchState = %s,
                    @SearchDistrict = %s,
                    @SearchPhone = %s,
                    @SearchEmail = %s,
                    @SearchBoard = %s,
                    @SearchMedium = %s,
                    @SearchPincode = %s,
                    @SearchRegistrationNumber = %s,
                    @SearchPrincipalName = %s,
                    @SearchDirectorName = %s,
                    @SearchStatus = %s,
                    @ShowDeleted = %s,
                    @SortColumn = %s,
                    @SortDirection = %s,
                    @PageNumber = %s,
                    @PageSize = %s,
                    @UserId = %s
                """,
                [
                    from_date,
                    to_date,
                    search_params['school_code'] or None,
                    search_params['school_name'] or None,
                    country,
                    state,
                    district,
                    search_params['phone'] or None,
                    search_params['email'] or None,
                    board,
                    medium,
                    search_params['pincode'] or None,
                    search_params['registration_number'] or None,
                    search_params['principal_name'] or None,
                    search_params['director_name'] or None,
                    search_params['status'] or None,
                    bool(int(search_params['show_deleted'])),
                    sort_column,
                    sort_direction,
                    page_number,
                    page_size,
                    user_id
                ]
            )
            columns = [col[0] for col in cursor.description]
            rows = cursor.fetchall()
            schools = [dict(zip(columns, row)) for row in rows]
            if schools:
                total_schools = schools[0].get('TotalCount', 0)
                active_schools = schools[0].get('ActiveSchools', 0)
                deleted_schools = schools[0].get('DeletedSchools', 0)

        # Convert SchoolLogo to base64
        for school in schools:
            if school.get('SchoolLogo'):
                school['SchoolLogo'] = f"data:image/jpeg;base64,{base64.b64encode(school['SchoolLogo']).decode('utf-8')}"
    except Exception as e:
        logger.error(f"Error fetching schools: {e}")
        messages.error(request, "An error occurred while fetching schools.")
        schools = []

    # Paginate results
    paginator = Paginator(schools, page_size)
    page_obj = paginator.get_page(page_number)

    # Update context with schools list specific data
    import json
    context.update({
        'page_obj': page_obj,
        'total_schools': total_schools,
        'active_schools': active_schools,
        'deleted_schools': deleted_schools,
        'boards': boards,
        'mediums': mediums,
        'boards_json': json.dumps(boards),
        'mediums_json': json.dumps(mediums),
        'search_params': search_params,
        'sort_column': sort_column,
        'sort_direction': sort_direction,
        'page_size': page_size,
    })

    return render(request, 'schools_list.html', context)

# School Delete and Restore Views

@custom_login_required
def school_soft_delete(request, school_id):
    """Soft delete a school using the stored procedure"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method'})
    
    try:
        user_id = request.session.get('UserId')
        if not user_id:
            return JsonResponse({'success': False, 'message': 'Session expired. Please login again.'})
        
        with connection.cursor() as cursor:
            cursor.execute(
                'SELECT "Proc_SoftDeleteSchool"(%s, %s, %s)',
                [school_id, user_id, 'DELETE']
            )
            
        return JsonResponse({'success': True, 'message': 'School deleted successfully.'})
        
    except Exception as e:
        logger.error(f"Error deleting school {school_id}: {e}")
        return JsonResponse({'success': False, 'message': 'An error occurred while deleting the school.'})


@custom_login_required
def school_restore(request, school_id):
    """Restore a soft deleted school using the stored procedure"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method'})
    
    try:
        user_id = request.session.get('UserId')
        if not user_id:
            return JsonResponse({'success': False, 'message': 'Session expired. Please login again.'})
        
        with connection.cursor() as cursor:
            cursor.execute(
                'SELECT "Proc_SoftDeleteSchool"(%s, %s, %s)',
                [school_id, user_id, 'Restore']
            )
            
        return JsonResponse({'success': True, 'message': 'School restored successfully.'})
        
    except Exception as e:
        logger.error(f"Error restoring school {school_id}: {e}")
        return JsonResponse({'success': False, 'message': 'An error occurred while restoring the school.'})

# School Update Views

@custom_login_required
def school_update(request, school_id):
    """View to display school update form with pre-filled data"""
    # Get user context for header
    context = get_context(request)
    
    # Fetch school details
    school_data = None
    try:
        with connection.cursor() as cursor:
            cursor.execute("EXEC Proc_GetSchoolDetails_ByID @SchoolID = %s", [school_id])
            columns = [col[0] for col in cursor.description]
            rows = cursor.fetchall()
            if rows:
                school_data = dict(zip(columns, rows[0]))
                
                # Check if there was an error
                if school_data.get('Status') == 'Error':
                    messages.error(request, school_data.get('ErrorMessage', 'Error retrieving school details'))
                    return redirect('schools_list')
                
                # Convert SchoolLogo to base64 if exists
                if school_data.get('SchoolLogo'):
                    school_data['SchoolLogo'] = f"data:image/jpeg;base64,{base64.b64encode(school_data['SchoolLogo']).decode('utf-8')}"
    except Exception as e:
        logger.error(f"Error fetching school details: {e}")
        messages.error(request, "An error occurred while fetching school details.")
        return redirect('schools_list')
    
    if not school_data:
        messages.error(request, "School not found.")
        return redirect('schools_list')
    
    # Dropdown data will be loaded via JavaScript API calls
    # This matches the approach used in the create school page
    
    # Update context with school data
    context.update({
        'school_data': school_data,
        'school_id': school_id,
    })
    
    return render(request, 'update_school.html', context)

@custom_login_required
def school_update_submit(request, school_id):
    """View to handle school update form submission"""
    if request.method != 'POST':
        return redirect('school_update', school_id=school_id)
    
    # Get form data
    school_name = request.POST.get('school_name', '').strip()
    registration_number = request.POST.get('registration_number', '').strip()
    address = request.POST.get('address', '').strip()
    district = request.POST.get('district', '').strip()
    state = request.POST.get('state', '').strip()
    country = request.POST.get('country', '').strip()
    pincode = request.POST.get('pincode', '').strip()
    phone = request.POST.get('phone', '').strip()
    email = request.POST.get('email', '').strip()
    website = request.POST.get('website', '').strip()
    board_id = request.POST.get('board', '').strip()
    medium_id = request.POST.get('medium', '').strip()
    principal_name = request.POST.get('principal_name', '').strip()
    principal_contact_mail = request.POST.get('principal_contact_mail', '').strip()
    principal_contact_phone = request.POST.get('principal_contact_phone', '').strip()
    director_name = request.POST.get('director_name', '').strip()
    director_contact_phone = request.POST.get('director_contact_phone', '').strip()
    director_contact_email = request.POST.get('director_contact_email', '').strip()
    establish_date = request.POST.get('establish_date', '').strip()
    
    # Debug logging
    logger.info(f"Form data received for school {school_id}:")
    logger.info(f"  School Name: {school_name}")
    logger.info(f"  Country: {country}")
    logger.info(f"  State: {state}")
    logger.info(f"  District: {district}")
    logger.info(f"  Board ID: {board_id}")
    logger.info(f"  Medium ID: {medium_id}")
    
    # Get user ID for UpdatedBy
    user_id = request.session.get('UserId')
    if not user_id:
        messages.error(request, 'User session expired. Please login again.')
        return redirect('login')
    
    # Validate required fields
    if not school_name:
        messages.error(request, 'School name is required.')
        return render(request, 'update_school.html', get_context(request))
    
    # Convert district, state, country, board_id, medium_id to integers
    try:
        district = int(district) if district else None
        state = int(state) if state else None
        country = int(country) if country else None
        board_id = int(board_id) if board_id else None
        medium_id = int(medium_id) if medium_id else None
    except (ValueError, TypeError) as e:
        logger.warning(f"Invalid geographic/board/medium selection: {str(e)}")
        messages.error(request, 'Invalid geographic, board, or medium selection.')
        return render(request, 'update_school.html', get_context(request))
    
    # Validate phone numbers
    phone_fields = [
        ('phone', phone),
        ('principal_contact_phone', principal_contact_phone),
        ('director_contact_phone', director_contact_phone)
    ]
    for field_name, value in phone_fields:
        if value:
            logger.info(f"Validating {field_name}: '{value}' (type: {type(value)})")
            
            # Convert to string and remove all non-digit characters except +
            clean_value = str(value).strip()
            
            # Remove all formatting characters
            import re
            clean_value = re.sub(r'[^\d+]', '', clean_value)
            
            # Remove country codes
            if clean_value.startswith('+91'):
                clean_value = clean_value[3:]
            elif clean_value.startswith('91') and len(clean_value) > 10:
                clean_value = clean_value[2:]
            elif clean_value.startswith('0') and len(clean_value) == 11:
                clean_value = clean_value[1:]
            
            logger.info(f"Cleaned {field_name}: '{clean_value}' (length: {len(clean_value)})")
            
            # Validate phone number
            is_valid = False
            
            if len(clean_value) == 10 and clean_value.isdigit():
                # Valid 10-digit number
                if clean_value[0] in '6789':
                    is_valid = True
                else:
                    logger.warning(f"Phone number doesn't start with 6-9: '{clean_value}'")
                    # Allow it but log a warning
                    is_valid = True
            elif len(clean_value) == 9 and clean_value.isdigit():
                # 9-digit number - likely incomplete, allow but warn
                logger.warning(f"Incomplete phone number (9 digits): '{clean_value}' - allowing but should be completed")
                is_valid = True
            elif len(clean_value) < 9:
                # Too short - invalid
                is_valid = False
            else:
                # Too long or invalid format
                is_valid = False
            
            if not is_valid:
                logger.warning(f"Invalid {field_name}: '{value}' -> cleaned: '{clean_value}' (length: {len(clean_value)})")
                messages.error(request, f'{field_name.replace("_", " ").title()} must be a valid Indian mobile number (10 digits starting with 6, 7, 8, or 9). Received: "{value}" (cleaned: "{clean_value}")')
                return render(request, 'update_school.html', get_context(request))
            
            logger.info(f"Valid {field_name}: '{clean_value}'")
            
            # Store the clean value for database
            if field_name == 'phone':
                phone = clean_value
            elif field_name == 'principal_contact_phone':
                principal_contact_phone = clean_value
            elif field_name == 'director_contact_phone':
                director_contact_phone = clean_value
    
    # Validate emails
    email_regex = r'^[^\s@]+@[^\s@]+\.[^\s@]+$'
    email_fields = [
        ('email', email),
        ('principal_contact_mail', principal_contact_mail),
        ('director_contact_email', director_contact_email)
    ]
    for field_name, value in email_fields:
        if value and not re.match(email_regex, value):
            logger.warning(f"Invalid {field_name}: {value}")
            messages.error(request, f'{field_name.replace("_", " ").title()} must be a valid email address.')
            return render(request, 'update_school.html', get_context(request))
    
    # Validate pincode
    if pincode and not (pincode.isdigit() and len(pincode) == 6):
        messages.error(request, 'Pincode must be exactly 6 digits.')
        return render(request, 'update_school.html', get_context(request))
    
    # Handle school logo upload
    school_logo = None
    if 'school_logo' in request.FILES:
        logo_file = request.FILES['school_logo']
        is_valid, message = validate_uploaded_file(logo_file, ALLOWED_IMAGE_TYPES)
        if not is_valid:
            messages.error(request, message)
            return render(request, 'update_school.html', get_context(request))
        school_logo = logo_file.read()
    
    # Convert establish_date to date object
    establish_date_obj = None
    if establish_date:
        try:
            establish_date_obj = safe_strptime(establish_date, '%Y-%m-%d').date()
        except ValueError:
            messages.error(request, 'Invalid establish date format.')
            return render(request, 'update_school.html', get_context(request))
    
    # Call stored procedure to update school
    try:
        logger.info(f"Calling stored procedure Proc_UpdateSchool_Set for school {school_id}")
        with connection.cursor() as cursor:
            cursor.execute(
                """
                DECLARE @Status NVARCHAR(50);
                DECLARE @ErrorMessage NVARCHAR(500);
                EXEC Proc_UpdateSchool_Set 
                    @SchoolID = %s,
                    @SchoolName = %s, 
                    @RegistrationNumber = %s, 
                    @Address = %s, 
                    @District = %s, 
                    @State = %s, 
                    @Country = %s, 
                    @Pincode = %s, 
                    @Phone = %s, 
                    @Email = %s, 
                    @Website = %s, 
                    @UpdatedBy = %s, 
                    @SchoolLogo = %s, 
                    @BoardID = %s, 
                    @MediumID = %s, 
                    @PrincipalName = %s, 
                    @PrincipalContactMail = %s, 
                    @PrincipalContactPhone = %s, 
                    @DirectorName = %s, 
                    @DirectorContactPhone = %s, 
                    @DirectorContactEmail = %s, 
                    @EstablishDate = %s, 
                    @Status = @Status OUTPUT, 
                    @ErrorMessage = @ErrorMessage OUTPUT;
                SELECT @Status AS Status, @ErrorMessage AS ErrorMessage;
                """,
                [
                    school_id,
                    school_name,
                    registration_number or None,
                    address or None,
                    district,
                    state,
                    country,
                    pincode or None,
                    phone or None,
                    email or None,
                    website or None,
                    user_id,
                    school_logo,
                    board_id,
                    medium_id,
                    principal_name or None,
                    principal_contact_mail or None,
                    principal_contact_phone or None,
                    director_name or None,
                    director_contact_phone or None,
                    director_contact_email or None,
                    establish_date_obj,
                ]
            )
            result = cursor.fetchone()
            status = result[0] if result else 'Error'
            error_message = result[1] if result else 'Unknown error occurred'
            
            logger.info(f"Stored procedure result: Status={status}, ErrorMessage={error_message}")
            
            if status == 'Success':
                logger.info(f"School {school_id} update successful, clearing messages and setting success message")
                # Clear any existing messages
                list(messages.get_messages(request))
                messages.success(request, 'School updated successfully!')
                logger.info(f"Success message set, redirecting to schools list")
                # Redirect with unique parameter to break back button history
                # Create redirect URL with parameters
                redirect_url = reverse('schools_list') + f'?updated={school_id}&t={int(time.time())}'
                response = HttpResponseRedirect(redirect_url)
                
                # Add cache control headers to prevent back button issues
                response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
                response['Pragma'] = 'no-cache'
                response['Expires'] = '0'
                return response
            else:
                messages.error(request, f'Error updating school: {error_message}')
                context = get_context(request)
                # Fetch school data for error display
                try:
                    with connection.cursor() as cursor:
                        cursor.execute("EXEC Proc_GetSchoolDetails_ByID @SchoolID = %s", [school_id])
                        columns = [col[0] for col in cursor.description]
                        rows = cursor.fetchall()
                        if rows:
                            school_data = dict(zip(columns, rows[0]))
                            if school_data.get('SchoolLogo'):
                                school_data['SchoolLogo'] = f"data:image/jpeg;base64,{base64.b64encode(school_data['SchoolLogo']).decode('utf-8')}"
                            context['school_data'] = school_data
                except Exception as e:
                    logger.error(f"Error fetching school data for error display: {e}")
                return render(request, 'update_school.html', context)
                
    except Exception as e:
        logger.error(f"Error updating school: {e}")
        # Only show error message if we haven't already shown a success message
        if not any(msg.tags == 'success' for msg in messages.get_messages(request)):
            messages.error(request, 'An error occurred while updating the school.')
            context = get_context(request)
            # Fetch school data for error display
            try:
                with connection.cursor() as cursor:
                    cursor.execute("EXEC Proc_GetSchoolDetails_ByID @SchoolID = %s", [school_id])
                    columns = [col[0] for col in cursor.description]
                    rows = cursor.fetchall()
                    if rows:
                        school_data = dict(zip(columns, rows[0]))
                        if school_data.get('SchoolLogo'):
                            school_data['SchoolLogo'] = f"data:image/jpeg;base64,{base64.b64encode(school_data['SchoolLogo']).decode('utf-8')}"
                        context['school_data'] = school_data
            except Exception as fetch_error:
                logger.error(f"Error fetching school data for error display: {fetch_error}")
            return render(request, 'update_school.html', context)

@custom_login_required
def load_more_schools(request):
    user_id = request.session.get('UserId')
    if not user_id:
        return JsonResponse({'success': False, 'message': 'Session expired'}, status=401)

    page_number = safe_int(request.GET.get('page', 1))
    page_size = safe_int(request.GET.get('per_page', 10))

    # Search and filter parameters
    search_params = {
        'school_code': request.GET.get('school_code', ''),
        'school_name': request.GET.get('school_name', ''),
        'registration_number': request.GET.get('registration_number', ''),
        'created_at': request.GET.get('created_at', ''),
        'to_date': request.GET.get('to_date', ''),
        'phone': request.GET.get('phone', ''),
        'email': request.GET.get('email', ''),
        'principal_name': request.GET.get('principal_name', ''),
        'director_name': request.GET.get('director_name', ''),
        'board': request.GET.get('board', ''),
        'medium': request.GET.get('medium', ''),
        'status': request.GET.get('status', ''),
        'country': request.GET.get('country', ''),
        'state': request.GET.get('state', ''),
        'district': request.GET.get('district', ''),
        'pincode': request.GET.get('pincode', ''),
    }

    # Convert string IDs to integers
    try:
        board = int(search_params['board']) if search_params['board'] else None
        medium = int(search_params['medium']) if search_params['medium'] else None
        country = int(search_params['country']) if search_params['country'] else None
        state = int(search_params['state']) if search_params['state'] else None
        district = int(search_params['district']) if search_params['district'] else None
    except ValueError:
        board = medium = country = state = district = None

    # Convert date strings
    from_date = None
    to_date = None
    try:
        if search_params['created_at']:
            from_date = safe_strptime(search_params['created_at'], '%Y-%m-%d').date()
        if search_params['to_date']:
            to_date = safe_strptime(search_params['to_date'], '%Y-%m-%d').date()
    except ValueError:
        search_params['created_at'] = ''
        search_params['to_date'] = ''

    sort_column = request.GET.get('sort_column', 'SchoolID')
    sort_direction = request.GET.get('sort_direction', 'DESC')

    schools = []
    total_schools = 0
    active_schools = 0
    try:
        with connection.cursor() as cursor:
            cursor.execute(
                """
                EXEC Proc_SchoolList_get 
                    @SearchCreatedAt = %s,
                    @SearchToDate = %s,
                    @SearchSchoolCode = %s,
                    @SearchSchoolName = %s,
                    @SearchCountry = %s,
                    @SearchState = %s,
                    @SearchDistrict = %s,
                    @SearchPhone = %s,
                    @SearchEmail = %s,
                    @SearchBoard = %s,
                    @SearchMedium = %s,
                    @SearchPincode = %s,
                    @SearchRegistrationNumber = %s,
                    @SearchPrincipalName = %s,
                    @SearchDirectorName = %s,
                    @SearchStatus = %s,
                    @SortColumn = %s,
                    @SortDirection = %s,
                    @PageNumber = %s,
                    @PageSize = %s,
                    @UserId = %s
                """,
                [
                    from_date,
                    to_date,
                    search_params['school_code'] or None,
                    search_params['school_name'] or None,
                    country,
                    state,
                    district,
                    search_params['phone'] or None,
                    search_params['email'] or None,
                    board,
                    medium,
                    search_params['pincode'] or None,
                    search_params['registration_number'] or None,
                    search_params['principal_name'] or None,
                    search_params['director_name'] or None,
                    search_params['status'] or None,
                    sort_column,
                    sort_direction,
                    page_number,
                    page_size,
                    user_id
                ]
            )
            columns = [col[0] for col in cursor.description]
            rows = cursor.fetchall()
            schools = [dict(zip(columns, row)) for row in rows]
            if schools:
                total_schools = schools[0].get('TotalCount', 0)
                active_schools = schools[0].get('ActiveSchools', 0)
                deleted_schools = schools[0].get('DeletedSchools', 0)

        # Convert SchoolLogo to base64
        for school in schools:
            if school.get('SchoolLogo'):
                school['SchoolLogo'] = f"data:image/jpeg;base64,{base64.b64encode(school['SchoolLogo']).decode('utf-8')}"
    except Exception as e:
        logger.error(f"Error fetching schools: {e}")
        return JsonResponse({'success': False, 'message': 'Error fetching schools'}, status=500)

    # Prepare JSON response
    response_data = {
        'success': True,
        'schools': schools,
        'total_count': total_schools,
        'active_schools': active_schools,
        'start_index': (page_number - 1) * page_size + 1,
        'end_index': min(page_number * page_size, total_schools),
        'has_next': page_number * page_size < total_schools,
        'has_previous': page_number > 1
    }

    return JsonResponse(response_data)

@custom_login_required
def export_schools(request):
    user_id = request.session.get('UserId')
    if not user_id:
        messages.error(request, "Session expired. Please login again.")
        return redirect('login')

    # Search and filter parameters
    search_params = {
        'school_code': request.GET.get('school_code', ''),
        'school_name': request.GET.get('school_name', ''),
        'registration_number': request.GET.get('registration_number', ''),
        'created_at': request.GET.get('created_at', ''),
        'to_date': request.GET.get('to_date', ''),
        'phone': request.GET.get('phone', ''),
        'email': request.GET.get('email', ''),
        'principal_name': request.GET.get('principal_name', ''),
        'director_name': request.GET.get('director_name', ''),
        'board': request.GET.get('board', ''),
        'medium': request.GET.get('medium', ''),
        'status': request.GET.get('status', ''),
        'country': request.GET.get('country', ''),
        'state': request.GET.get('state', ''),
        'district': request.GET.get('district', ''),
        'pincode': request.GET.get('pincode', ''),
    }

    # Convert string IDs to integers
    try:
        board = int(search_params['board']) if search_params['board'] else None
        medium = int(search_params['medium']) if search_params['medium'] else None
        country = int(search_params['country']) if search_params['country'] else None
        state = int(search_params['state']) if search_params['state'] else None
        district = int(search_params['district']) if search_params['district'] else None
    except ValueError:
        board = medium = country = state = district = None

    # Convert date strings
    from_date = None
    to_date = None
    try:
        if search_params['created_at']:
            from_date = safe_strptime(search_params['created_at'], '%Y-%m-%d').date()
        if search_params['to_date']:
            to_date = safe_strptime(search_params['to_date'], '%Y-%m-%d').date()
    except ValueError:
        search_params['created_at'] = ''
        search_params['to_date'] = ''

    schools = []
    try:
        with connection.cursor() as cursor:
            cursor.execute(
                """
                EXEC Proc_SchoolList_get 
                    @SearchCreatedAt = %s,
                    @SearchToDate = %s,
                    @SearchSchoolCode = %s,
                    @SearchSchoolName = %s,
                    @SearchCountry = %s,
                    @SearchState = %s,
                    @SearchDistrict = %s,
                    @SearchPhone = %s,
                    @SearchEmail = %s,
                    @SearchBoard = %s,
                    @SearchMedium = %s,
                    @SearchPincode = %s,
                    @SearchRegistrationNumber = %s,
                    @SearchPrincipalName = %s,
                    @SearchDirectorName = %s,
                    @SearchStatus = %s,
                    @SortColumn = %s,
                    @SortDirection = %s,
                    @PageNumber = %s,
                    @PageSize = %s,
                    @UserId = %s
                """,
                [
                    from_date,
                    to_date,
                    search_params['school_code'] or None,
                    search_params['school_name'] or None,
                    country,
                    state,
                    district,
                    search_params['phone'] or None,
                    search_params['email'] or None,
                    board,
                    medium,
                    search_params['pincode'] or None,
                    search_params['registration_number'] or None,
                    search_params['principal_name'] or None,
                    search_params['director_name'] or None,
                    search_params['status'] or None,
                    'SchoolID',
                    'ASC',
                    1,
                    10000,  # Fetch all records for export
                    user_id
                ]
            )
            columns = [col[0] for col in cursor.description]
            rows = cursor.fetchall()
            schools = [dict(zip(columns, row)) for row in rows]
    except Exception as e:
        logger.error(f"Error fetching schools for export: {e}")
        messages.error(request, "An error occurred while exporting schools.")
        return redirect('schools_list')

    # Create CSV response
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="schools_export.csv"'

    writer = csv.writer(response)
    # Write header
    headers = [
        'SchoolID', 'SchoolCode', 'SchoolName', 'RegistrationNumber', 'Address', 
        'DistrictName', 'StateName', 'CountryName', 'Pincode', 'Phone', 'Email', 
        'Website', 'BoardName', 'MediumName', 'PrincipalName', 'PrincipalContactMail', 
        'PrincipalContactPhone', 'DirectorName', 'DirectorContactPhone', 
        'DirectorContactEmail', 'EstablishDate', 'CreatedAt', 'STATUS'
    ]
    writer.writerow(headers)

    # Write data
    for school in schools:
        writer.writerow([
            school.get('SchoolID', ''),
            school.get('SchoolCode', ''),
            school.get('SchoolName', ''),
            school.get('RegistrationNumber', 'N/A'),
            school.get('Address', 'N/A'),
            school.get('DistrictName', 'N/A'),
            school.get('StateName', 'N/A'),
            school.get('CountryName', 'N/A'),
            school.get('Pincode', 'N/A'),
            school.get('Phone', 'N/A'),
            school.get('Email', 'N/A'),
            school.get('Website', 'N/A'),
            school.get('BoardName', 'N/A'),
            school.get('MediumName', 'N/A'),
            school.get('PrincipalName', 'N/A'),
            school.get('PrincipalContactMail', 'N/A'),
            school.get('PrincipalContactPhone', 'N/A'),
            school.get('DirectorName', 'N/A'),
            school.get('DirectorContactPhone', 'N/A'),
            school.get('DirectorContactEmail', 'N/A'),
            school.get('EstablishDate', 'N/A') if not school.get('EstablishDate') else school['EstablishDate'].strftime('%Y-%m-%d'),
            school.get('CreatedAt', 'N/A') if not school.get('CreatedAt') else school['CreatedAt'].strftime('%Y-%m-%d'),
            school.get('STATUS', 'N/A')
        ])

    return response

#--Dark Mode Toggle View


def toggle_dark_mode(request):
    if request.method == 'POST':
        try:
            # Parse JSON data from request body
            data = json.loads(request.body)
            is_dark_mode = data.get('dark_mode', False)
            
            user_id = request.session.get('UserId')
            logger.info(f"Dark mode toggle request - UserID: {user_id}, Dark Mode: {is_dark_mode}")
            
            if user_id:
                # Update session first
                request.session['dark_mode'] = is_dark_mode
                logger.info(f"Session updated for UserID {user_id}: {is_dark_mode}")
                
                # Update database using the new procedure
                with connection.cursor() as cursor:
                    cursor.execute(
                        'SELECT * FROM "Proc_User_Theme_Save"(%s, %s)',
                        [user_id, is_dark_mode]
                    )
                    result = cursor.fetchone()
                    
                    if result:
                        status = result[0]
                        message = result[1]
                        logger.info(f"Theme procedure result - Status: {status}, Message: {message}")
                        
                        if status == 'SUCCESS':
                            return JsonResponse({
                                'success': True, 
                                'user_id': user_id,
                                'dark_mode': is_dark_mode
                            })
                        else:
                            logger.error(f"Theme procedure failed: {message}")
                            return JsonResponse({'success': False, 'error': message})
                    else:
                        logger.error("Theme procedure returned no result")
                        return JsonResponse({'success': False, 'error': 'No response from database'})
            else:
                logger.warning("Dark mode toggle attempted without user login")
                return JsonResponse({'success': False, 'error': 'User not logged in'})\
        
        except Exception as e:
            logger.error(f"Error in dark mode toggle: {str(e)}")
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})






#-- Get Header level info for all views 
#-- Student Admission View

@custom_login_required
def student_admission(request):
    """
    Render student admission form and handle form submission using merged stored procedure
    """
    # Check if user is logged in and get user ID from session
    user_id = request.session.get('UserId')
    school_id = request.session.get('SchoolID')  # Get school ID from session
    
    # Debug session data
    logger.info(f"Session data - UserId: {user_id}, SchoolID: {school_id}")
    logger.info(f"Session keys: {list(request.session.keys())}")
    
    if not user_id:
        messages.error(request, "Please login to access admission form")
        return redirect('login')
    
    # Load dark mode preference
    dark_mode = request.session.get('dark_mode', False)
    
    # Handle Super Admin overriding school choice via URL
    profile_id = request.custom_user.get('profile_id') if hasattr(request, 'custom_user') else request.session.get('ProfileID')
    if str(profile_id) == '1':
        get_school = request.GET.get('school_id')
        if get_school:
            try:
                school_id = int(get_school)
            except (ValueError, TypeError):
                pass

    # Debug TO FILE
    try:
        with open('debug_api.txt', 'a') as f:
            f.write(f"\n[student_admission VIEW] Time: {datetime.now()}\n")
            f.write(f"ProfileID: {profile_id}\n")
            f.write(f"SchoolID (Final): {school_id}\n")
            f.write(f"GET param full: {request.GET}\n")
    except: pass

    # Fetch context data for header
    context = get_context(request)
    
    # Fetch admission-related fee types using stored procedure
    admission_fee_types = []
    try:
        with connection.cursor() as cursor:
            cursor.execute("EXEC Proc_Admission_Fee_Types_Get @SchoolID = %s", [school_id])
            admission_fee_types = cursor.fetchall()
            logger.info(f"Loaded {len(admission_fee_types)} admission fee types for School ID: {school_id}")
    except Exception as e:
        logger.error(f"Error fetching admission fee types for School ID {school_id}: {str(e)}")
        messages.warning(request, "Could not load admission fee types. Please contact administrator.")
    
    # Monthly fee types will be loaded via AJAX when class is selected
    monthly_fee_types = []
    
    context.update({
        'dark_mode': dark_mode,
        'user_id': user_id,
        'school_id': school_id,
        'admission_fee_types': admission_fee_types,
        'monthly_fee_types': monthly_fee_types,
    })
    
    # Handle form submission
    if request.method == 'POST':
        try:
            # Extract and clean form data
            full_name = request.POST.get('fullName')
            gender = request.POST.get('gender')
            date_of_birth = request.POST.get('dateOfBirth')
            age = request.POST.get('age')
            blood_group = request.POST.get('bloodGroup')
            category = request.POST.get('category')
            religion = request.POST.get('religion')
            nationality = request.POST.get('nationality')
            mother_tongue = request.POST.get('motherTongue')
            present_address = request.POST.get('presentAddress')
            permanent_address = request.POST.get('permanentAddress')
            parent_mobile = request.POST.get('parentMobile', '').replace('-', '')  # Clean mobile number
            alternate_number = request.POST.get('alternateNumber', '').replace('-', '')
            email = request.POST.get('email')
            father_name = request.POST.get('fatherName')
            father_occupation = request.POST.get('fatherOccupation')
            father_qualification = request.POST.get('fatherQualification')
            father_aadhaar = request.POST.get('fatherAadhaar', '').replace('-', '')  # Clean Aadhaar
            father_mobile = request.POST.get('fatherMobile', '').replace('-', '')
            mother_name = request.POST.get('motherName')
            mother_occupation = request.POST.get('motherOccupation')
            mother_qualification = request.POST.get('motherQualification')
            mother_aadhaar = request.POST.get('motherAadhaar', '').replace('-', '')
            mother_mobile = request.POST.get('motherMobile', '').replace('-', '')
            guardian_name = request.POST.get('guardianName')
            guardian_relation = request.POST.get('guardianRelation')
            guardian_mobile = request.POST.get('guardianMobile', '').replace('-', '')
            last_school = request.POST.get('lastSchool')
            last_class = request.POST.get('lastClass')
            tc_number = request.POST.get('tcNumber')
            medium_of_instruction = request.POST.get('medium')
            academic_year_id = request.POST.get('academicYear')
            admission_class = request.POST.get('admissionClass')
            section = request.POST.get('section')
            stream = request.POST.get('stream')
            mode_of_admission = request.POST.get('mode')
            admission_date = request.POST.get('admissionDate')
            father_sign = request.POST.get('fatherSign')
            mother_sign = request.POST.get('motherSign')
            guardian_sign = request.POST.get('guardianSign')
            student_sign = request.POST.get('studentSign')
            declaration_date = request.POST.get('declarationDate')
            principal_approval = request.POST.get('principalApproval')
            student_aadhaar = request.POST.get('studentAadhaarNumber', '').replace('-', '')  # Clean Aadhaar
            student_password = request.POST.get('studentPassword')
            
            # Extract location IDs
            country_id = request.POST.get('country_id')
            state_id = request.POST.get('state_id')
            district_id = request.POST.get('district_id')
            
            # If manual text entries were used instead of dropdowns
            if not country_id and request.POST.get('country_text'):
                # For now, we'll set to NULL. You might want to implement logic to handle manual entries
                country_id = None
                
            if not state_id and request.POST.get('state_text'):
                state_id = None
                
            if not district_id and request.POST.get('district_text'):
                district_id = None
            
            # Stored procedure handles all validation - no need for extra queries
            
            # Hash the password using Django's make_password
            student_password_hash = make_password(student_password) if student_password else None
            
            # Extract fee details - BOTH admission fees AND class-based monthly fees
            fees_data = []
            total_fees_amount = 0
            logger.info(f"Processing {len(admission_fee_types)} admission fee types")
            
            # Process admission fees
            for fee_type in admission_fee_types:
                try:
                    fee_type_id = int(fee_type[0])
                    fee_type_name = fee_type[2]  # FeeTypeName
                    default_amount = float(fee_type[3])  # DefaultAmount
                    discount_str = request.POST.get(f'discount_{fee_type_id}', '').strip()
                    discount_percentage = float(discount_str) if discount_str and discount_str.replace('.', '', 1).isdigit() else 0.0
                    
                    logger.info(f"Processing admission fee type {fee_type_id} ({fee_type_name}): amount={default_amount}, discount_str='{discount_str}', discount_percentage={discount_percentage}")
                    
                    if discount_percentage < 0:
                        discount_percentage = 0.0
                    elif discount_percentage > 100:
                        discount_percentage = 100.0

                    final_amount = default_amount
                    if discount_percentage > 0:
                        final_amount = default_amount - (default_amount * discount_percentage / 100)

                    fee_data = {
                        'feeTypeId': fee_type_id,
                        'feeTypeName': fee_type_name,
                        'amount': round(default_amount, 2),
                        'discountPercentage': round(discount_percentage, 2),
                        'finalAmount': round(final_amount, 2)
                    }
                    
                    fees_data.append(fee_data)
                    total_fees_amount += round(final_amount, 2)
                    
                    logger.info(f"Added admission fee: {fee_data}")
                
                except (ValueError, TypeError) as e:
                    logger.warning(f"Error processing admission fee type {fee_type[0]}: {str(e)}")
                    continue
            
            # Process class-based monthly fees from form
            monthly_fee_ids = request.POST.getlist('monthly_fee_id[]')
            monthly_fee_names = request.POST.getlist('monthly_fee_name[]')
            monthly_fee_amounts = request.POST.getlist('monthly_fee_amount[]')
            monthly_fee_discounts = request.POST.getlist('monthly_fee_discount[]')
            
            logger.info(f"Processing {len(monthly_fee_ids)} class-based monthly fees")
            
            for i in range(len(monthly_fee_ids)):
                try:
                    fee_type_id = int(monthly_fee_ids[i])
                    fee_type_name = monthly_fee_names[i]
                    default_amount = float(monthly_fee_amounts[i])
                    discount_percentage = float(monthly_fee_discounts[i]) if monthly_fee_discounts[i] else 0.0
                    
                    if discount_percentage < 0:
                        discount_percentage = 0.0
                    elif discount_percentage > 100:
                        discount_percentage = 100.0
                    
                    final_amount = default_amount
                    if discount_percentage > 0:
                        final_amount = default_amount - (default_amount * discount_percentage / 100)
                    
                    fee_data = {
                        'feeTypeId': fee_type_id,
                        'feeTypeName': fee_type_name,
                        'amount': round(default_amount, 2),
                        'discountPercentage': round(discount_percentage, 2),
                        'finalAmount': round(final_amount, 2)
                    }
                    
                    fees_data.append(fee_data)
                    total_fees_amount += round(final_amount, 2)
                    
                    logger.info(f"Added monthly fee: {fee_data}")
                    
                except (ValueError, TypeError, IndexError) as e:
                    logger.warning(f"Error processing monthly fee at index {i}: {str(e)}")
                    continue
            fees_json = json.dumps({'fees': fees_data}) if fees_data else None
            logger.info(f"Fees JSON being sent to procedure: {fees_json}")
            logger.info(f"Number of fees: {len(fees_data) if fees_data else 0}")
            if fees_data:
                for i, fee in enumerate(fees_data):
                    logger.info(f"Fee {i+1}: TypeId={fee.get('feeTypeId')}, Amount={fee.get('amount')}, Discount={fee.get('discountPercentage')}%, Final={fee.get('finalAmount')}")
            else:
                logger.warning("No fees_data found - this might be the issue!")
                logger.info(f"admission_fee_types count: {len(admission_fee_types) if admission_fee_types else 0}")
                logger.info(f"admission_fee_types: {admission_fee_types}")

            # Process document rows
            documents = []
            doc_types = request.POST.getlist('docType[]')
            doc_files = request.FILES.getlist('docFile[]')
            for doc_type, file in zip(doc_types, doc_files):
                if doc_type and file:
                    is_valid, message = validate_uploaded_file(file, ALLOWED_DOCUMENT_TYPES)
                    if not is_valid:
                        messages.error(request, f"Document validation failed: {message}")
                        return render(request, 'student_admission.html', {
                            **context,
                            'form_data': request.POST
                        })
                    document_data = base64.b64encode(file.read()).decode('utf-8')
                    documents.append({
                        'type': doc_type,
                        'name': file.name,
                        'data': document_data
                    })
            
            documents_json = json.dumps(documents) if documents else None
            
            # Call merged stored procedure
            with connection.cursor() as cursor:
                # Define output parameters
                user_code = None
                error_message = None
                
                # Debug: Log the exact parameters being passed
                logger.info(f"Calling stored procedure with fees_json: {fees_json}")
                logger.info(f"School ID: {school_id}, User ID: {user_id}")
                logger.info(f"Location IDs - Country: {country_id}, State: {state_id}, District: {district_id}")
                logger.info(f"Student Password Hash: {'Provided' if student_password_hash else 'None'}")
                
                # Execute stored procedure with user_id as CreatedBy
                cursor.execute("""
                    DECLARE @UserCode NVARCHAR(20);
                    DECLARE @ErrorMessage NVARCHAR(4000);
                    
                    EXEC Proc_Student_Admission_With_Documents
                        @FullName = %s, @Gender = %s, @DateOfBirth = %s, @Age = %s,
                        @BloodGroup = %s, @Category = %s, @Religion = %s,
                        @Nationality = %s, @MotherTongue = %s, @PresentAddress = %s,
                        @PermanentAddress = %s, @District = %s, @State = %s, @Country = %s,
                        @ParentMobile = %s, @AlternateNumber = %s, @Email = %s,
                        @FatherName = %s, @FatherOccupation = %s, @FatherQualification = %s,
                        @FatherAadhaar = %s, @FatherMobile = %s, @MotherName = %s,
                        @MotherOccupation = %s, @MotherQualification = %s, @MotherAadhaar = %s,
                        @MotherMobile = %s, @GuardianName = %s, @GuardianRelation = %s,
                        @GuardianMobile = %s, @LastSchool = %s, @LastClass = %s,
                        @TCNumber = %s, @MediumOfInstruction = %s, @AcademicYearID = %s, @AdmissionClass = %s,
                        @Section = %s, @Stream = %s, @ModeOfAdmission = %s,
                        @AdmissionDate = %s, @FatherSign = %s, @MotherSign = %s,
                        @GuardianSign = %s, @StudentSign = %s, @DeclarationDate = %s,
                        @PrincipalApproval = %s, @CreatedBy = %s, @StudentPassword = %s,
                        @StudentAadhaar = %s, @FeesJson = %s, @DocumentsJson = %s, @SchoolID = %s,
                        @UserCode = @UserCode OUTPUT, @ErrorMessage = @ErrorMessage OUTPUT;
                    
                    SELECT @UserCode AS UserCode, @ErrorMessage AS ErrorMessage;
                """, [
                    full_name, gender, date_of_birth, age, blood_group, category, religion,
                    nationality, mother_tongue, present_address, permanent_address,
                    district_id, state_id, country_id,
                    parent_mobile, alternate_number, email,
                    father_name, father_occupation, father_qualification, father_aadhaar,
                    father_mobile, mother_name, mother_occupation, mother_qualification,
                    mother_aadhaar, mother_mobile, guardian_name, guardian_relation,
                    guardian_mobile, last_school, last_class, tc_number, medium_of_instruction,
                    academic_year_id, admission_class, section, stream, mode_of_admission, admission_date,
                    father_sign, mother_sign, guardian_sign, student_sign, declaration_date,
                    principal_approval, user_id, student_password_hash,
                    student_aadhaar, fees_json, documents_json, school_id
                ])
                
                # Get the output values
                result = cursor.fetchone()
                if result:
                    user_code = result[0]
                    error_message = result[1]
                    
                if error_message:
                    raise ValueError(error_message)
                else:
                    logger.info(f"Student admission procedure completed successfully for user code: {user_code}")
            
            # Fees saved by stored procedure, documents saved directly
            logger.info(f"Student and fees saved successfully via stored procedure for user code: {user_code}")
            
            # Get student ID for documents and session data
            student_id = None
            with connection.cursor() as cursor:
                cursor.execute("SELECT StudentID FROM Student WHERE StudentCode = %s", [user_code])
                student_result = cursor.fetchone()
                if student_result:
                    student_id = student_result[0]
                    logger.info(f"Student ID retrieved: {student_id}")
            
            # Insert student documents directly into StudentDocuments
            if student_id and documents:
                try:
                    with connection.cursor() as cursor:
                        for doc in documents:
                            # Decode base64 back to bytes for VARBINARY insert
                            doc_bytes = base64.b64decode(doc['data']) if doc.get('data') else None
                            cursor.execute(
                                """
                                INSERT INTO StudentDocuments
                                (StudentID, DocumentType, DocumentName, DocumentData, UploadDate, CreatedBy, CreatedAt, IsDeleted)
                                VALUES (%s, %s, %s, %s, CURRENT_TIMESTAMP, %s, CURRENT_TIMESTAMP, 0)
                                """,
                                [student_id, doc.get('type'), doc.get('name'), doc_bytes, user_id]
                            )
                    logger.info(f"Successfully inserted {len(documents)} documents for student {student_id}")
                except Exception as e:
                    logger.error(f"Failed to save student documents: {e}")
            
            # Store minimal admission data in session using safe_json_obj to ensure JSON serialization
            admission_data_raw = {
                'student_id': int(student_id) if student_id else 0,
                'student_code': str(user_code) if user_code else '',
                'student_name': str(full_name) if full_name else '',
                'school_id': int(school_id) if school_id else 0,
                'academic_year_id': int(academic_year_id) if academic_year_id else 0,
                'admission_class': str(admission_class) if admission_class else '',
                'section': str(section) if section else '',
                'email': str(email) if email else '',
                'student_password': str(student_password) if student_password else '',
                'admission_date': str(admission_date) if admission_date else '',
                'gender': str(gender) if gender else '',
                'date_of_birth': str(date_of_birth) if date_of_birth else '',
                'age': str(age) if age else '',
                'blood_group': str(blood_group) if blood_group else '',
                'category': str(category) if category else '',
                'religion': str(religion) if religion else '',
                'nationality': str(nationality) if nationality else '',
                'mother_tongue': str(mother_tongue) if mother_tongue else '',
                'student_aadhaar': str(student_aadhaar) if student_aadhaar else '',
                'present_address': str(present_address) if present_address else '',
                'permanent_address': str(permanent_address) if permanent_address else '',
                'parent_mobile': str(parent_mobile) if parent_mobile else '',
                'alternate_number': str(alternate_number) if alternate_number else '',
                'father_name': str(father_name) if father_name else '',
                'father_occupation': str(father_occupation) if father_occupation else '',
                'father_qualification': str(father_qualification) if father_qualification else '',
                'father_aadhaar': str(father_aadhaar) if father_aadhaar else '',
                'father_mobile': str(father_mobile) if father_mobile else '',
                'mother_name': str(mother_name) if mother_name else '',
                'mother_occupation': str(mother_occupation) if mother_occupation else '',
                'mother_qualification': str(mother_qualification) if mother_qualification else '',
                'mother_aadhaar': str(mother_aadhaar) if mother_aadhaar else '',
                'mother_mobile': str(mother_mobile) if mother_mobile else '',
                'guardian_name': str(guardian_name) if guardian_name else '',
                'guardian_relation': str(guardian_relation) if guardian_relation else '',
                'guardian_mobile': str(guardian_mobile) if guardian_mobile else '',
                'last_school': str(last_school) if last_school else '',
                'last_class': str(last_class) if last_class else '',
                'tc_number': str(tc_number) if tc_number else '',
                'medium_of_instruction': str(medium_of_instruction) if medium_of_instruction else '',
                'stream': str(stream) if stream else '',
                'mode_of_admission': str(mode_of_admission) if mode_of_admission else '',
                'country_id': int(country_id) if country_id else 0,
                'state_id': int(state_id) if state_id else 0,
                'district_id': int(district_id) if district_id else 0
            }
            request.session['admission_data'] = safe_json_obj(admission_data_raw)
            
            # Redirect to payment page
            return redirect('payment_page')
            
        except Exception as e:
            logger.error(f"Error admitting student: {str(e)}", exc_info=True)
            error_msg = str(e)
            
            # Handle specific error cases
            if "User authentication required" in error_msg:
                messages.error(request, "Session expired. Please login again.")
                return redirect('login')
            elif "Invalid user account" in error_msg:
                messages.error(request, "Your account is no longer valid. Please contact administrator.")
                return redirect('login')
            elif "already exists in the system" in error_msg:
                messages.error(request, error_msg)
            else:
                messages.error(request, f"Error admitting student: {error_msg}")
                
            # Return to form with existing data
            return render(request, 'student_admission.html', {
                **context,
                'form_data': request.POST
            })
    
    # For GET request, render the form with necessary context data
    return render(request, 'student_admission.html', context)


@custom_login_required
def get_monthly_fee_types(request):
    """
    AJAX endpoint to get monthly fee types based on class selection using stored procedure
    """
    if request.method != 'GET':
        return JsonResponse({'error': 'Invalid request method'}, status=405)
    
    class_id = request.GET.get('class_id')
    school_id = request.session.get('SchoolID')
    
    if not class_id or not school_id:
        return JsonResponse({'error': 'Class ID and School ID are required'}, status=400)
    
    try:
        with connection.cursor() as cursor:
            cursor.execute("SELECT * FROM proc_monthly_fee_types_get(%s, %s)", [school_id, class_id])
            
            monthly_fees = []
            for row in cursor.fetchall():
                monthly_fees.append({
                    'FeeTypeId': row[0],
                    'SchoolId': row[1],
                    'FeeTypeName': row[2],
                    'DefaultAmount': float(row[3])
                })
            
            logger.info(f"Loaded {len(monthly_fees)} monthly fee types for School ID: {school_id}, Class ID: {class_id}")
            
            return JsonResponse({
                'success': True,
                'monthly_fees': monthly_fees,
                'school_id': school_id,
                'class_id': class_id
            })
            
    except Exception as e:
        logger.error(f"Error fetching monthly fee types for School ID {school_id}, Class ID {class_id}: {str(e)}")
        return JsonResponse({
            'error': 'Failed to load monthly fee types',
            'details': str(e),
            'school_id': school_id,
            'class_id': class_id
        }, status=500)

@custom_login_required
def payment_page(request):
    """
    Handle payment after successful admission
    """
    admission_data = request.session.get('admission_data')
    if not admission_data:
        messages.error(request, "No admission data found. Please start the admission process again.")
        return redirect('student_admission')
    
    # Read fees from Student_Fee_Assignment table
    total_fees = 0
    fee_breakdown = []
    
    try:
        from django.utils import timezone as dj_tz
        current_month = dj_tz.now().strftime('%Y%m')
        
        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT ft.FeeTypeName, sfa.FeeAmount, sfa.DiscountPercentage, sfa.FinalAmount
                FROM Student_Fee_Assignment sfa
                JOIN FeeType_Master ft ON sfa.FeeTypeId = ft.FeeTypeId
                WHERE sfa.StudentID = %s AND sfa.FeeMonth = %s AND ISNULL(sfa.IsDeleted,0)=0
                """,
                [admission_data.get('student_id'), current_month]
            )
            for row in cursor.fetchall():
                fee_data = {
                    'fee_name': str(row[0]) if row[0] else '',
                    'default_amount': float(row[1]) if row[1] is not None else 0.0,
                    'discount_percentage': float(row[2]) if row[2] is not None else 0.0,
                    'amount': float(row[3]) if row[3] is not None else 0.0
                }
                fee_breakdown.append(fee_data)
                total_fees += fee_data['amount']
        
        logger.info(f"Loaded {len(fee_breakdown)} fees from database for payment page")
    except Exception as e:
        logger.error(f"Error loading fees from database: {str(e)}")
        total_fees = 0
    
    # Get user context
    context = get_context(request)
    context.update({
        'admission_data': admission_data,
        'total_fees': total_fees,
        'fee_breakdown': fee_breakdown,
        'dark_mode': request.session.get('dark_mode', False)
    })
    
    if request.method == 'POST':
        # Process payment
        payment_mode = request.POST.get('payment_mode')
        paid_amount = request.POST.get('paid_amount')
        transaction_ref = request.POST.get('transaction_ref', '')
        user_id = request.session.get('UserId')
        
        try:
            # Read fee breakdown from Student_Fee_Assignment table
            from django.utils import timezone as dj_tz
            current_month = dj_tz.now().strftime('%Y%m')
            fee_breakdown = []
            fee_breakdown_serializable = []
            total_amount = 0
            
            # Get fees from Student_Fee_Assignment (already saved during admission)
            with connection.cursor() as cursor:
                cursor.execute(
                    """
                    SELECT sfa.FeeTypeId, ft.FeeTypeName, sfa.FeeAmount, sfa.DiscountPercentage, sfa.FinalAmount
                    FROM Student_Fee_Assignment sfa
                    JOIN FeeType_Master ft ON sfa.FeeTypeId = ft.FeeTypeId
                    WHERE sfa.StudentID = %s AND sfa.FeeMonth = %s AND ISNULL(sfa.IsDeleted,0)=0
                    """,
                    [admission_data['student_id'], current_month]
                )
                for row in cursor.fetchall():
                    # Convert ALL database types to JSON-serializable types
                    fee_data = {
                        'fee_type_id': int(row[0]) if row[0] is not None else 0,
                        'fee_name': str(row[1]) if row[1] else '',
                        'default_amount': float(row[2]) if row[2] is not None else 0.0,
                        'discount_percentage': float(row[3]) if row[3] is not None else 0.0,
                        'amount': float(row[4]) if row[4] is not None else 0.0
                    }
                    fee_breakdown.append(fee_data)
                    fee_breakdown_serializable.append(fee_data.copy())
                    total_amount += fee_data['amount']
            
            logger.info(f"Loaded {len(fee_breakdown)} fees from Student_Fee_Assignment for payment")
            
            # Generate unique receipt number for admission payment
            # Format: ADM-SCHOOLID-STUDENTCODE-YYYYMMDD-SEQUENCE
            receipt_number = None
            with connection.cursor() as cursor:
                # Get next sequence number for today's admission payments
                today_date = dj_tz.now().strftime('%Y%m%d')
                cursor.execute(
                    """
                    SELECT COUNT(*) + 1 
                    FROM Payment 
                    WHERE SchoolID = %s 
                      AND PaymentFor = 'Admission' 
                      AND CONVERT(DATE, PaymentDate) = CONVERT(DATE, CURRENT_TIMESTAMP)
                    """,
                    [admission_data['school_id']]
                )
                sequence = cursor.fetchone()[0]
                receipt_number = f"ADM-{admission_data['school_id']}-{admission_data['student_code']}-{today_date}-{sequence:03d}"
                logger.info(f"Generated receipt number: {receipt_number}")
                # Insert payment record with unique receipt number and complete fee breakdown
                cursor.execute(
                    """
                    INSERT INTO Payment (
                        SchoolID, PaymentFor, EntityID, EntityType, ReceiptNumber,
                        TotalAmount, PaidAmount, PaymentMode, TransactionRef, PaymentStatus,
                        PaymentDate, PaymentMonth, FeeBreakdown, Remarks, CreatedBy, CreatedAt, IsDeleted
                    ) VALUES (
                        %s, %s, %s, %s, %s,
                        %s, %s, %s, %s, %s,
                        CURRENT_TIMESTAMP, FORMAT(CURRENT_TIMESTAMP,'yyyyMM'), %s, %s, %s, CURRENT_TIMESTAMP, 0
                    )
                    """,
                    [
                        admission_data['school_id'], 'Admission', admission_data['student_id'], 'Student', receipt_number,
                        float(total_amount), float(paid_amount), payment_mode, transaction_ref, 'Paid',
                        json.dumps(fee_breakdown_serializable, cls=DjangoJSONEncoder), 'Admission payment', user_id
                    ]
                )
            
            # Ensure Section exists for the student's class; create if missing
            admission_class = admission_data.get('admission_class')
            section_name = admission_data.get('section')
            if admission_class and section_name:
                try:
                    with connection.cursor() as cursor:
                        cursor.execute(
                            "SELECT \"SectionID\" FROM \"SectionMaster\" WHERE \"IsDeleted\" = FALSE AND \"ClassID\" = %s AND \"SectionName\" = %s",
                            [admission_class, section_name]
                        )
                        found = cursor.fetchone()
                        if not found:
                            cursor.execute(
                                """
                                INSERT INTO "SectionMaster" ("ClassID", "SectionName", "IsActive", "CreatedBy", "CreatedAt", "IsDeleted")
                                VALUES (%s, %s, TRUE, %s, CURRENT_TIMESTAMP, FALSE)
                                """,
                                [admission_class, section_name, user_id]
                            )
                except Exception as e:
                    logger.error(f"Section ensure/create failed: {e}")

            # Build comprehensive receipt data for success screen and PDFs
            from django.utils import timezone as dj_tz
            
            payment_receipt_raw = {
                'receipt_number': str(receipt_number) if receipt_number else '',
                'student_name': str(admission_data.get('student_name', '')),
                'student_code': str(admission_data.get('student_code', '')),
                'payment_date': str(dj_tz.now().strftime('%Y-%m-%d %H:%M')),
                'payment_mode': str(payment_mode) if payment_mode else '',
                'amount_paid': float(paid_amount) if paid_amount else 0.0,
                'total_amount': float(total_amount),
                'transaction_ref': str(transaction_ref) if transaction_ref else '',
                'fee_breakdown': fee_breakdown_serializable,
                # Additional student details for comprehensive receipt
                'gender': str(admission_data.get('gender', '')),
                'date_of_birth': str(admission_data.get('date_of_birth', '')),
                'age': str(admission_data.get('age', '')),
                'blood_group': admission_data.get('blood_group'),
                'category': admission_data.get('category'),
                'religion': admission_data.get('religion'),
                'nationality': admission_data.get('nationality'),
                'mother_tongue': admission_data.get('mother_tongue'),
                'student_aadhaar': admission_data.get('student_aadhaar'),
                'present_address': admission_data.get('present_address'),
                'permanent_address': admission_data.get('permanent_address'),
                'parent_mobile': admission_data.get('parent_mobile'),
                'alternate_number': admission_data.get('alternate_number'),
                'father_name': admission_data.get('father_name'),
                'father_occupation': admission_data.get('father_occupation'),
                'father_mobile': admission_data.get('father_mobile'),
                'mother_name': admission_data.get('mother_name'),
                'mother_occupation': admission_data.get('mother_occupation'),
                'mother_mobile': admission_data.get('mother_mobile'),
                'guardian_name': admission_data.get('guardian_name'),
                'guardian_relation': admission_data.get('guardian_relation'),
                'guardian_mobile': admission_data.get('guardian_mobile'),
                'last_school': admission_data.get('last_school'),
                'last_class': admission_data.get('last_class'),
                'admission_class': admission_data.get('admission_class'),
                'section': admission_data.get('section'),
                'stream': admission_data.get('stream'),
                'mode_of_admission': admission_data.get('mode_of_admission'),
                'admission_date': str(admission_data.get('admission_date', '')),
            }
            # Ensure all data is JSON-serializable
            payment_receipt = safe_json_obj(payment_receipt_raw)

            # Fetch school name/logo for receipt header
            try:
                with connection.cursor() as cursor:
                    cursor.execute("SELECT SchoolName, SchoolLogo FROM SchoolMaster WHERE SchoolID = %s", [admission_data.get('school_id')])
                    row = cursor.fetchone()
                    if row:
                        school_name, school_logo_blob = row
                        payment_receipt['school_name'] = school_name
                        if school_logo_blob:
                            payment_receipt['school_logo'] = _bytes_to_data_uri(school_logo_blob)
            except Exception:
                pass

            # Get accurate data from database using procedures
            ack_data = {}
            receipt_data = {}
            
            try:
                # Get acknowledgment data from database
                with connection.cursor() as cursor:
                    cursor.execute("SELECT * FROM proc_student_acknowledgment_get(%s)", [admission_data['student_code']])
                    columns = [col[0] for col in cursor.description]
                    row = cursor.fetchone()
                    if row:
                        ack_data = dict(zip(columns, row))
                        # Convert binary logo to base64 if exists
                        if ack_data.get('school_logo'):
                            ack_data['school_logo'] = _bytes_to_data_uri(ack_data['school_logo'])
                    
                    # Get fee structure from database
                    cursor.execute("SELECT * FROM proc_student_fee_structure_get(NULL, %s)", [admission_data['student_code']])
                    fee_columns = [col[0] for col in cursor.description]
                    fee_rows = cursor.fetchall()
                    fee_breakdown = [dict(zip(fee_columns, row)) for row in fee_rows]
                    
                    # Build receipt data with accurate fee breakdown
                    receipt_data = {
                        **payment_receipt,
                        'fee_breakdown': fee_breakdown,
                        **ack_data  # Include all student data
                    }
                    
            except Exception as e:
                logger.error(f"Error fetching accurate data from database: {e}")
                # Fallback to session data
                ack_data = payment_receipt
                receipt_data = payment_receipt
            
            # Store completion data with accurate database info
            request.session['admission_completion'] = safe_json_obj({
                'acknowledgment': ack_data,
                'payment_receipt': receipt_data,
            })

            # Send emails asynchronously using database queue system
            if admission_data.get('email'):
                logger.info("Queuing admission emails")
                
                # Queue emails using database email tracking system
                try:
                    # Queue acknowledgment email
                    EmailTrackingManager.create_email_task(
                        email_code='ADMISSION_ACKNOWLEDGMENT',
                        to_email=admission_data.get('email'),
                        placeholders={'student_code': admission_data.get('student_code')},
                        school_id=admission_data.get('school_id'),
                        priority=5,
                        student_code=admission_data.get('student_code'),
                        has_attachments=True
                    )
                    
                    # Queue payment receipt email
                    EmailTrackingManager.create_email_task(
                        email_code='PAYMENT_RECEIPT',
                        to_email=admission_data.get('email'),
                        placeholders={
                            'student_code': admission_data.get('student_code'),
                            'receipt_number': receipt_number,
                            'amount_paid': paid_amount,
                            'payment_mode': payment_mode,
                            'transaction_ref': transaction_ref
                        },
                        school_id=admission_data.get('school_id'),
                        priority=5,
                        student_code=admission_data.get('student_code'),
                        has_attachments=True
                    )
                    
                    logger.info("Admission emails queued successfully")
                    messages.success(request, f"Payment of ₹{paid_amount} processed successfully! Confirmation emails are being sent to {admission_data.get('email')}")
                except Exception as email_error:
                    logger.error(f"Failed to queue admission emails: {str(email_error)}")
                    messages.success(request, f"Payment of ₹{paid_amount} processed successfully!")
            else:
                messages.success(request, f"Payment of ₹{paid_amount} processed successfully!")
                logger.info("No email address provided, skipping email sending")
            
            return redirect('admission_complete')
            
        except Exception as e:
            logger.error(f"Error processing payment: {str(e)}")
            messages.error(request, f"Error processing payment: {str(e)}")
    
    return render(request, 'payment.html', context)

@custom_login_required
def admission_complete(request):
    """Show acknowledgment and payment receipt with options to print and download PDFs, and email them."""
    data = request.session.get('admission_completion')
    if not data:
        messages.error(request, 'No completion data found.')
        return redirect('dashboard')

    student_code = data.get('acknowledgment', {}).get('student_code') or data.get('payment_receipt', {}).get('student_code')
    ack = {}
    
    if student_code:
        try:
            with connection.cursor() as cursor:
                cursor.execute("SELECT * FROM proc_student_acknowledgment_get(%s)", [student_code])
                row = cursor.fetchone()
                if row:
                    columns = [col[0] for col in cursor.description]
                    ack = dict(zip(columns, row))
                    if ack.get('StudentCode'):
                        ack['student_code'] = ack['StudentCode']
                    if ack.get('school_logo'):
                        import base64
                        ack['school_logo'] = f"data:image/png;base64,{base64.b64encode(ack['school_logo']).decode('utf-8')}"
                ack['instructions'] = []
                if cursor.nextset():
                    ack['instructions'] = [{'title': r[0], 'text': r[1]} for r in cursor.fetchall()]
                ack['documents'] = []
                if cursor.nextset():
                    ack['documents'] = [{'type': r[0], 'name': r[1]} for r in cursor.fetchall()]
                ack['fees'] = []
                if cursor.nextset():
                    ack['fees'] = [{'FeeTypeName': r[0], 'FeeAmount': float(r[1] or 0), 'DiscountPercentage': float(r[2] or 0), 'FinalAmount': float(r[3] or 0)} for r in cursor.fetchall()]
                if cursor.nextset():
                    total_row = cursor.fetchone()
                    ack['total_amount'] = float(total_row[0] or 0) if total_row else 0
        except Exception as e:
            logger.error(f"Error fetching acknowledgment data: {str(e)}")

    user_context = get_context(request)
    
    context = {
        'acknowledgment': ack,
        'payment_receipt': data.get('payment_receipt'),
        'student_code': student_code,
        **user_context,
        'dark_mode': request.session.get('dark_mode', False),
    }

    # Emails are already sent asynchronously via database queue in payment_page
    # No need to send emails again here to avoid duplicates
    logger.info("Admission complete page rendered - emails already queued in payment_page")

    return render(request, 'admission_complete.html', context)

@custom_login_required
@xframe_options_exempt
def print_acknowledgment(request):
    """Render acknowledgment as a printable page (used by payment success screen).
    Preview mode (preview=1) does not require login.
    """
    # Check if this is a preview request (no login required for preview)
    is_preview = request.GET.get('preview') == '1'
    
    if is_preview:
        # Generate sample data for preview
        from datetime import datetime as _dt
        ack = {
            'student_name': 'Rahul Kumar',
            'student_code': 'STU2024001',
            'student_id': 1,
            'gender': 'Male',
            'date_of_birth': '2010-05-15',
            'age': 14,
            'blood_group': 'O+',
            'category': 'General',
            'religion': 'Hindu',
            'nationality': 'Indian',
            'admission_class': 'Class 9',
            'section': 'A',
            'stream': 'Science',
            'admission_date': _dt.now().strftime('%Y-%m-%d'),
            'academic_year': '2024-2025',
            'mode_of_admission': 'Regular',
            'email': 'rahul.kumar@example.com',
            'parent_mobile': '9876543210',
            'present_address': '123, MG Road, Bangalore',
            'district_name': 'Bangalore Urban',
            'state_name': 'Karnataka',
            'country_name': 'India',
            'father_name': 'Suresh Kumar',
            'father_mobile': '9876543210',
            'mother_name': 'Priya Kumar',
            'mother_mobile': '9876543211',
            'school_name': 'Sample School',
            'school_phone': '080-12345678',
            'school_email': 'info@sampleschool.com',
            'current_date': _dt.now().strftime('%Y-%m-%d'),
            'instructions': [
                {'title': 'Uniform', 'text': 'Students must wear proper school uniform daily'},
                {'title': 'Attendance', 'text': 'Minimum 75% attendance is mandatory'},
                {'title': 'Fee Payment', 'text': 'Fees must be paid before 10th of every month'}
            ],
            'documents': [
                {'type': 'Birth Certificate', 'name': 'birth_cert.pdf'},
                {'type': 'Aadhaar Card', 'name': 'aadhaar.pdf'},
                {'type': 'Previous School TC', 'name': 'transfer_cert.pdf'}
            ],
            'fees': [
                {'name': 'Admission Fee', 'amount': 5000, 'discount': 0, 'final': 5000},
                {'name': 'Tuition Fee', 'amount': 10000, 'discount': 10, 'final': 9000},
                {'name': 'Lab Fee', 'amount': 2000, 'discount': 0, 'final': 2000}
            ]
        }
    else:
        data = request.session.get('admission_completion') or {}
        ack = data.get('acknowledgment', {}).copy()
    
    student_code = request.GET.get('student_code')
    
    if student_code:
        try:
            with connection.cursor() as cursor:
                cursor.execute("SELECT * FROM proc_student_acknowledgment_get(%s)", [student_code])
                
                # Result set 1: Student data
                row = cursor.fetchone()
                if row:
                    columns = [col[0] for col in cursor.description]
                    ack = dict(zip(columns, row))
                    # Map StudentCode to student_code for template
                    if ack.get('StudentCode'):
                        ack['student_code'] = ack['StudentCode']
                    if ack.get('school_logo'):
                        ack['school_logo'] = _bytes_to_data_uri(ack['school_logo'])
                else:
                    ack = {}
                
                # Result set 2: Instructions
                ack['instructions'] = []
                if cursor.nextset():
                    ack['instructions'] = [{'title': r[0], 'text': r[1]} for r in cursor.fetchall()]
                
                # Result set 3: Documents
                ack['documents'] = []
                if cursor.nextset():
                    ack['documents'] = [{'type': r[0], 'name': r[1]} for r in cursor.fetchall()]
                
                # Result set 4: Fees
                ack['fees'] = []
                if cursor.nextset():
                    ack['fees'] = [{'FeeTypeName': r[0], 'FeeAmount': float(r[1] or 0), 'DiscountPercentage': float(r[2] or 0), 'FinalAmount': float(r[3] or 0)} for r in cursor.fetchall()]
                
                if cursor.nextset():
                    total_row = cursor.fetchone()
                    ack['total_amount'] = float(total_row[0] or 0) if total_row else 0
                    
        except Exception as e:
            logger.error(f"Error: {e}", exc_info=True)
            ack = {'instructions': [], 'documents': [], 'fees': []}
    
    # Ensure required fields exist
    if not ack.get('school_name'):
        try:
            with connection.cursor() as cursor:
                cursor.execute("SELECT SchoolName, SchoolLogo FROM SchoolMaster WHERE SchoolID = %s", [request.session.get('SchoolID')])
                row = cursor.fetchone()
                if row:
                    ack['school_name'] = row[0]
                    if row[1]:
                        ack['school_logo'] = _bytes_to_data_uri(row[1])
        except Exception:
            pass
    
    if not ack.get('current_date'):
        from datetime import datetime as _dt
        ack['current_date'] = _dt.now().strftime('%Y-%m-%d')
    
    # Ensure lists exist
    if 'instructions' not in ack:
        ack['instructions'] = []
    if 'documents' not in ack:
        ack['documents'] = []
    if 'fees' not in ack:
        ack['fees'] = []
    
    # Get school's selected template or preview template
    template_file = 'core/document_templates/admission_acknowledgment/admission_acknowledgment.html'
    
    # Check if this is a preview request with template parameter
    preview_template = request.GET.get('template')
    if preview_template:
        template_file = preview_template
    elif not is_preview:
        # Use school's selected template from database (only for non-preview)
        school_id = request.session.get('SchoolID') or ack.get('SchoolID')
        if school_id:
            try:
                with connection.cursor() as cursor:
                    cursor.execute("""
                        SELECT "TemplateFile" FROM "TemplateSettings" 
                        WHERE "SchoolID" = %s AND "TemplateType" = 'AdmissionAcknowledgment' 
                        AND "IsActive" = TRUE AND "IsDeleted" = FALSE
                    """, [school_id])
                    row = cursor.fetchone()
                    if row and row[0]:
                        template_file = row[0]
            except Exception:
                pass
    
    return render(request, template_file, { 'acknowledgment': ack })

@custom_login_required
def print_receipt(request):
    data = request.session.get('admission_completion') or {}
    student_code = data.get('payment_receipt', {}).get('student_code') or data.get('acknowledgment', {}).get('student_code')
    rcpt = {}
    
    if student_code:
        try:
            with connection.cursor() as cursor:
                cursor.execute("SELECT * FROM proc_student_acknowledgment_get(%s)", [student_code])
                row = cursor.fetchone()
                if row:
                    columns = [col[0] for col in cursor.description]
                    rcpt = dict(zip(columns, row))
                    if rcpt.get('StudentCode'):
                        rcpt['student_code'] = rcpt['StudentCode']
                    if rcpt.get('school_logo'):
                        import base64
                        rcpt['school_logo'] = f"data:image/png;base64,{base64.b64encode(rcpt['school_logo']).decode('utf-8')}"
                if cursor.nextset():
                    cursor.fetchall()
                if cursor.nextset():
                    cursor.fetchall()
                fee_breakdown = []
                if cursor.nextset():
                    fee_breakdown = [{'FeeTypeName': r[0], 'FeeAmount': float(r[1] or 0), 'DiscountPercentage': float(r[2] or 0), 'FinalAmount': float(r[3] or 0)} for r in cursor.fetchall()]
                rcpt['fee_breakdown'] = fee_breakdown
                if cursor.nextset():
                    total_row = cursor.fetchone()
                    rcpt['total_amount'] = float(total_row[0] or 0) if total_row else 0
        except Exception as e:
            logger.error(f"Error fetching receipt data: {str(e)}")
    
    template_file = 'core/document_templates/payment_receipt/payment_success.html'
    school_id = request.session.get('SchoolID')
    
    if school_id:
        try:
            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT "TemplateFile" FROM "TemplateSettings" 
                    WHERE "SchoolID" = %s AND "TemplateType" = 'PaymentReceipt' 
                    AND "IsActive" = TRUE AND "IsDeleted" = FALSE
                """, [school_id])
                row = cursor.fetchone()
                if row and row[0]:
                    template_file = row[0]
        except Exception as e:
            logger.error(f"Error fetching receipt template preference: {str(e)}")
    
    return render(request, template_file, { 'payment_receipt': rcpt })

@require_POST
@custom_login_required
def clear_receipt_session(request):
    """Clear temporary admission receipt data from session."""
    try:
        if 'admission_completion' in request.session:
            del request.session['admission_completion']
        if 'admission_data' in request.session:
            del request.session['admission_data']
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@custom_login_required
def view_applications(request):
    # Session context
    user_id = request.session.get('UserId')
    profile_id = request.session.get('ProfileID')
    school_id = request.session.get('SchoolID')

    # Filters and pagination (aligned with user_list pattern)
    page_number = safe_int(request.GET.get('page', 1))
    page_size = safe_int(request.GET.get('per_page', 10))

    search = request.GET.get('search', '')
    class_id = request.GET.get('class_id') or None
    section_id = request.GET.get('section_id') or None
    gender = request.GET.get('gender') or None
    category = request.GET.get('category') or None
    status = request.GET.get('status') or None
    from_date = request.GET.get('from_date') or None
    to_date = request.GET.get('to_date') or None

    from_date_obj = None
    to_date_obj = None
    if from_date:
        try:
            from_date_obj = safe_strptime(from_date, '%Y-%m-%d')
        except ValueError:
            from_date_obj = None
    if to_date:
        try:
            to_date_obj = safe_strptime(to_date, '%Y-%m-%d')
        except ValueError:
            to_date_obj = None

    applications = []
    total_count = 0

    try:
        with connection.cursor() as cursor:
            cursor.execute(
                """
                EXEC dbo.usp_GetStudentApplications
                    @ProfileID=%s,
                    @UserID=%s,
                    @SchoolID=%s,
                    @Search=%s,
                    @ClassID=%s,
                    @SectionID=%s,
                    @Gender=%s,
                    @Category=%s,
                    @Status=%s,
                    @FromAdmissionDate=%s,
                    @ToAdmissionDate=%s,
                    @PageNumber=%s,
                    @PageSize=%s
                """,
                [
                    profile_id,
                    user_id,
                    school_id,
                    search or None,
                    class_id,
                    section_id,
                    gender,
                    category,
                    status,
                    from_date_obj,
                    to_date_obj,
                    page_number,
                    page_size,
                ],
            )
            columns = [col[0] for col in cursor.description]
            rows = cursor.fetchall()
            for row in rows:
                applications.append(dict(zip(columns, row)))
            if applications:
                total_count = applications[0].get('TotalCount', 0) or 0
    except Exception as e:
        logger.error(f"Error fetching applications: {e}")
        messages.error(request, "Error loading applications. Please try again.")

    # Compute simple pagination helpers
    start_index = (page_number - 1) * page_size + 1 if total_count else 0
    end_index = min(page_number * page_size, total_count) if total_count else 0
    has_next = end_index < total_count

    # Get user context for header
    context = get_context(request)
    context.update({
        'applications': applications,
        'page': page_number,
        'per_page': page_size,
        'total_count': total_count,
        'start_index': start_index,
        'end_index': end_index,
        'has_next': has_next,
        'search': search,
        'class_id': class_id or '',
        'section_id': section_id or '',
        'gender': gender or '',
        'category': category or '',
        'status': status or '',
        'from_date': from_date or '',
        'to_date': to_date or '',
    })
    return render(request, 'view_applications.html', context)


@custom_login_required
@custom_login_required
def view_application_detail(request, student_code):
    """View detailed information for a specific application using unified procedure"""
    school_id = request.session.get('SchoolID')
    
    application = None
    fee_structure = []
    documents = []
    
    try:
        with connection.cursor() as cursor:
            # Call unified stored procedure
            cursor.execute("EXEC Proc_application_details_get @StudentCode = %s, @SchoolID = %s", 
                         [student_code, school_id])
            
            # Result Set 1: Student Application Details
            columns = [col[0] for col in cursor.description]
            row = cursor.fetchone()
            if row:
                # Check for error status
                if len(row) >= 2 and row[0] == 'ERROR':
                    messages.error(request, row[1])
                    return redirect('admission_applicants')
                application = dict(zip(columns, row))
            
            # Result Set 2: Fee Structure
            if cursor.nextset():
                for fee_row in cursor.fetchall():
                    fee_structure.append({
                        'FeeTypeName': fee_row[0],
                        'FeeAmount': float(fee_row[1]) if fee_row[1] else 0.0,
                        'DiscountPercentage': float(fee_row[2]) if fee_row[2] else 0.0,
                        'FinalAmount': float(fee_row[3]) if fee_row[3] else 0.0
                    })
            
            # Result Set 3: Documents
            if cursor.nextset():
                doc_rows = cursor.fetchall()
                logger.info(f"Found {len(doc_rows)} documents for student {student_code}")
                
                for doc_row in doc_rows:
                    doc_data = {
                        'DocumentID': doc_row[0],
                        'DocumentType': doc_row[1],
                        'DocumentName': doc_row[2],
                        'UploadDate': doc_row[4],
                        'DocumentData': None,
                        'MimeType': None
                    }
                    
                    # Convert varbinary to base64 for viewing
                    if doc_row[3] and isinstance(doc_row[3], bytes):  # DocumentData
                        try:
                            import base64
                            doc_data['DocumentData'] = base64.b64encode(doc_row[3]).decode('utf-8')
                            # Determine MIME type based on file extension
                            file_ext = doc_row[2].lower().split('.')[-1] if '.' in doc_row[2] else ''
                            if file_ext in ['jpg', 'jpeg']:
                                doc_data['MimeType'] = 'image/jpeg'
                            elif file_ext == 'png':
                                doc_data['MimeType'] = 'image/png'
                            elif file_ext == 'pdf':
                                doc_data['MimeType'] = 'application/pdf'
                            else:
                                doc_data['MimeType'] = 'application/octet-stream'
                            logger.info(f"Document {doc_row[0]} ({doc_row[2]}): Successfully encoded, MIME: {doc_data['MimeType']}")
                        except Exception as encode_error:
                            logger.error(f"Error encoding document {doc_row[0]}: {encode_error}")
                    else:
                        logger.warning(f"Document {doc_row[0]} ({doc_row[2]}): No data or not bytes type")
                    
                    documents.append(doc_data)
                    print(f"DEBUG: Document {doc_row[0]} - Has DocumentData: {bool(doc_data['DocumentData'])}, Type: {type(doc_data['DocumentData'])}, Length: {len(doc_data['DocumentData']) if doc_data['DocumentData'] else 0}")
            else:
                logger.warning(f"No documents result set returned for student {student_code}")
                    
    except Exception as e:
        logger.error(f"Error fetching application detail: {e}")
        messages.error(request, "Error loading application details.")
        return redirect('admission_applicants')
    
    if not application:
        messages.error(request, "Application not found.")
        return redirect('admission_applicants')
    
    logger.info(f"Application detail for {student_code}: {len(documents)} documents loaded")
    
    context = get_context(request)
    context['application'] = application
    context['fee_structure'] = fee_structure
    context['documents'] = documents
    return render(request, 'application_detail.html', context)

@custom_login_required
def load_more_applications(request):
    if request.method != 'GET':
        return JsonResponse({'error': 'Method not allowed'}, status=405)

    user_id = request.session.get('UserId')
    profile_id = request.session.get('ProfileID')
    school_id = request.session.get('SchoolID')

    page_number = safe_int(request.GET.get('page', 1))
    page_size = safe_int(request.GET.get('per_page', 10))

    search = request.GET.get('search', '')
    class_id = request.GET.get('class_id') or None
    section_id = request.GET.get('section_id') or None
    gender = request.GET.get('gender') or None
    category = request.GET.get('category') or None
    status = request.GET.get('status') or None
    from_date = request.GET.get('from_date') or None
    to_date = request.GET.get('to_date') or None

    from_date_obj = None
    to_date_obj = None
    if from_date:
        try:
            from_date_obj = safe_strptime(from_date, '%Y-%m-%d')
        except ValueError:
            from_date_obj = None
    if to_date:
        try:
            to_date_obj = safe_strptime(to_date, '%Y-%m-%d')
        except ValueError:
            to_date_obj = None

    applications = []
    total_count = 0

    try:
        with connection.cursor() as cursor:
            cursor.execute(
                """
                EXEC dbo.usp_GetStudentApplications
                    @ProfileID=%s,
                    @UserID=%s,
                    @SchoolID=%s,
                    @Search=%s,
                    @ClassID=%s,
                    @SectionID=%s,
                    @Gender=%s,
                    @Category=%s,
                    @Status=%s,
                    @FromAdmissionDate=%s,
                    @ToAdmissionDate=%s,
                    @PageNumber=%s,
                    @PageSize=%s
                """,
                [
                    profile_id,
                    user_id,
                    school_id,
                    search or None,
                    class_id,
                    section_id,
                    gender,
                    category,
                    status,
                    from_date_obj,
                    to_date_obj,
                    page_number,
                    page_size,
                ],
            )
            columns = [col[0] for col in cursor.description]
            rows = cursor.fetchall()
            for row in rows:
                applications.append(dict(zip(columns, row)))
            if applications:
                total_count = applications[0].get('TotalCount', 0) or 0
    except Exception as e:
        logger.error(f"Error fetching applications: {e}")
        return JsonResponse({'error': 'Failed to load applications'}, status=500)

    start_index = (page_number - 1) * page_size + 1 if total_count else 0
    end_index = min(page_number * page_size, total_count) if total_count else 0
    has_next = end_index < total_count

    return JsonResponse({
        'applications': applications,
        'total_count': total_count,
        'start_index': start_index,
        'end_index': end_index,
        'has_next': has_next,
    })


# Face Template Management Functions
def authenticate_with_face_template(identifier, face_data):
    """
    Authenticate user using face template matching
    """
    try:
        import json
        
        # Parse face data (should be JSON string of face descriptor)
        try:
            current_descriptor = json.loads(face_data)
            if not isinstance(current_descriptor, list) or len(current_descriptor) != 128:
                return {'success': False, 'error': 'Invalid face descriptor format'}
        except (json.JSONDecodeError, TypeError):
            return {'success': False, 'error': 'Invalid face data format'}
        
        # Find user and their face templates
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT 
                    u."UserID",
                    u."UserName",
                    u."ProfileID",
                    ft."FaceDescriptor",
                    ft."TemplateVersion"
                FROM "UserMaster" u
                LEFT JOIN "FaceTemplates" ft ON u."UserID" = ft."UserID" AND ft."IsActive" = TRUE
                WHERE (u."UserName" = %s OR u."UserCode" = %s OR u."Email" = %s)
                  AND u."IsActive" = TRUE
                  AND u."IsDeleted" IS NOT TRUE
            """, [identifier, identifier, identifier])
            
            user_data = cursor.fetchone()
            
            if not user_data:
                return {'success': False, 'error': 'User not found'}
            
            user_id, user_name, profile_id, stored_descriptor, template_version = user_data
            
            if not stored_descriptor:
                return {'success': False, 'error': 'No face template registered for this user'}
            
            # Parse stored descriptor
            try:
                stored_descriptor_list = json.loads(stored_descriptor)
                if not isinstance(stored_descriptor_list, list) or len(stored_descriptor_list) != 128:
                    return {'success': False, 'error': 'Invalid stored face template'}
            except (json.JSONDecodeError, TypeError):
                return {'success': False, 'error': 'Corrupted face template data'}
            
            # Calculate similarity using Euclidean distance
            current_array = np.array(current_descriptor, dtype=np.float32)
            stored_array = np.array(stored_descriptor_list, dtype=np.float32)
            
            # Calculate Euclidean distance
            distance = np.linalg.norm(current_array - stored_array)
            
            # Convert distance to similarity score (0-1)
            # Typical face recognition threshold is around 0.6
            similarity = max(0, 1 - distance)
            
            # Check if similarity meets threshold (80% match)
            if similarity >= 0.8:
                logger.info(f"Face authentication successful for user {identifier}, similarity: {similarity:.3f}")
                return {
                    'success': True, 
                    'user_id': user_id,
                    'user_name': user_name,
                    'profile_id': profile_id,
                    'similarity': similarity
                }
            else:
                logger.warning(f"Face authentication failed for user {identifier}, similarity: {similarity:.3f}")
                return {'success': False, 'error': f'Face match too low: {similarity:.1%}'}
                
    except Exception as e:
        logger.error(f"Face authentication error for {identifier}: {str(e)}", exc_info=True)
        return {'success': False, 'error': 'Authentication system error'}


def store_face_template(user_id, face_descriptor, created_by=None):
    """
    Store or update face template for a user
    """
    try:
        import json
        
        # Validate face descriptor
        if not isinstance(face_descriptor, (list, tuple)) or len(face_descriptor) != 128:
            return {'success': False, 'error': 'Invalid face descriptor format'}
        
        # Convert to JSON string
        descriptor_json = json.dumps(face_descriptor)
        
        with connection.cursor() as cursor:
            # Check if user already has an active face template
            cursor.execute("""
                SELECT FaceTemplateID FROM FaceTemplates 
                WHERE UserID = %s AND IsActive = TRUE
            """, [user_id])
            
            existing_template = cursor.fetchone()
            
            if existing_template:
                # Update existing template
                cursor.execute("""
                    UPDATE FaceTemplates 
                    SET FaceDescriptor = %s, 
                        UpdatedAt = CURRENT_TIMESTAMP,
                        UpdatedBy = %s
                    WHERE UserID = %s AND IsActive = TRUE
                """, [descriptor_json, created_by, user_id])
            else:
                # Create new template
                cursor.execute("""
                    INSERT INTO FaceTemplates (UserID, FaceDescriptor, CreatedBy, UpdatedBy)
                    VALUES (%s, %s, %s, %s)
                """, [user_id, descriptor_json, created_by, created_by])
            
            logger.info(f"Face template stored/updated for user ID {user_id}")
            return {'success': True, 'message': 'Face template saved successfully'}
            
    except Exception as e:
        logger.error(f"Error storing face template for user {user_id}: {str(e)}", exc_info=True)
        return {'success': False, 'error': 'Failed to save face template'}


def get_face_template(user_id):
    """
    Retrieve face template for a user
    """
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT FaceDescriptor, TemplateVersion, CreatedAt, UpdatedAt
                FROM FaceTemplates 
                WHERE UserID = %s AND IsActive = TRUE
            """, [user_id])
            
            template_data = cursor.fetchone()
            
            if template_data:
                return {
                    'success': True,
                    'descriptor': template_data[0],
                    'version': template_data[1],
                    'created_at': template_data[2],
                    'updated_at': template_data[3]
                }
            else:
                return {'success': False, 'error': 'No face template found'}
                
    except Exception as e:
        logger.error(f"Error retrieving face template for user {user_id}: {str(e)}", exc_info=True)
        return {'success': False, 'error': 'Failed to retrieve face template'}


@custom_login_required
def register_face_template(request):
    """
    API endpoint to register/update face template for current user
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        import json
        
        user_id = request.session.get('UserId')
        if not user_id:
            return JsonResponse({'error': 'User not authenticated'}, status=401)
        
        # Get face descriptor from request
        face_descriptor = request.POST.get('face_descriptor')
        if not face_descriptor:
            return JsonResponse({'error': 'Face descriptor required'}, status=400)
        
        # Parse and validate descriptor
        try:
            descriptor_list = json.loads(face_descriptor)
            if not isinstance(descriptor_list, list) or len(descriptor_list) != 128:
                return JsonResponse({'error': 'Invalid face descriptor format'}, status=400)
        except (json.JSONDecodeError, TypeError):
            return JsonResponse({'error': 'Invalid face descriptor data'}, status=400)
        
        # Store the template
        result = store_face_template(user_id, descriptor_list, user_id)
        
        if result['success']:
            return JsonResponse({'success': True, 'message': result['message']})
        else:
            return JsonResponse({'success': False, 'error': result['error']}, status=500)
            
    except Exception as e:
        logger.error(f"Error in register_face_template: {str(e)}", exc_info=True)
        return JsonResponse({'success': False, 'error': 'Internal server error'}, status=500)


def register_face_template_by_identifier(request):
    """
    API endpoint to register face template by identifier (for login process)
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)
    
    try:
        import json
        
        identifier = request.POST.get('identifier', '').strip()
        face_descriptor = request.POST.get('face_descriptor')
        
        if not identifier:
            return JsonResponse({'success': False, 'error': 'Identifier required'}, status=400)
        
        if not face_descriptor:
            return JsonResponse({'success': False, 'error': 'Face descriptor required'}, status=400)
        
        # Find user by identifier
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT "UserID" FROM "UserMaster" 
                WHERE ("UserName" = %s OR "UserCode" = %s OR "Email" = %s)
                  AND "IsActive" = TRUE
                  AND COALESCE("IsDeleted", FALSE) = FALSE
            """, [identifier, identifier, identifier])
            
            user_data = cursor.fetchone()
            
            if not user_data:
                return JsonResponse({'success': False, 'error': 'User not found'}, status=404)
            
            user_id = user_data[0]
        
        # Parse and validate descriptor
        try:
            descriptor_list = json.loads(face_descriptor)
            if not isinstance(descriptor_list, list) or len(descriptor_list) != 128:
                return JsonResponse({'success': False, 'error': 'Invalid face descriptor format'}, status=400)
        except (json.JSONDecodeError, TypeError):
            return JsonResponse({'success': False, 'error': 'Invalid face descriptor data'}, status=400)
        
        # Store the template
        result = store_face_template(user_id, descriptor_list, user_id)
        
        if result['success']:
            return JsonResponse({'success': True, 'message': result['message']})
        else:
            return JsonResponse({'success': False, 'error': result['error']}, status=500)
            
    except Exception as e:
        logger.error(f"Error in register_face_template_by_identifier: {str(e)}", exc_info=True)
        return JsonResponse({'success': False, 'error': 'Internal server error'}, status=500)


def get_face_template_by_identifier(request):
    """
    API endpoint to get face template by identifier (for login process)
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)
    
    try:
        identifier = request.POST.get('identifier', '').strip()
        
        if not identifier:
            return JsonResponse({'success': False, 'error': 'Identifier required'}, status=400)
        
        # Find user and their face template
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT 
                    u."UserID",
                    ft."FaceDescriptor",
                    ft."TemplateVersion",
                    ft."CreatedAt",
                    ft."UpdatedAt"
                FROM "UserMaster" u
                LEFT JOIN "FaceTemplates" ft ON u."UserID" = ft."UserID" AND ft."IsActive" = TRUE
                WHERE (u."UserName" = %s OR u."UserCode" = %s OR u."Email" = %s)
                  AND u."IsActive" = TRUE
                  AND u."IsDeleted" IS NOT TRUE
            """, [identifier, identifier, identifier])
            
            user_data = cursor.fetchone()
            
            if not user_data:
                return JsonResponse({'success': False, 'error': 'User not found'}, status=404)
            
            user_id, face_descriptor, template_version, created_at, updated_at = user_data
            
            if face_descriptor:
                return JsonResponse({
                    'success': True,
                    'descriptor': json.loads(face_descriptor),
                    'version': template_version,
                    'created_at': created_at.isoformat() if created_at else None,
                    'updated_at': updated_at.isoformat() if updated_at else None
                })
            else:
                return JsonResponse({'success': False, 'error': 'No face template found'})
                
    except Exception as e:
        logger.error(f"Error in get_face_template_by_identifier: {str(e)}", exc_info=True)
        return JsonResponse({'success': False, 'error': 'Internal server error'}, status=500)


def get_user_photo(request):
    """
    API endpoint to get user photo by identifier (for simple face comparison)
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)
    
    try:
        identifier = request.POST.get('identifier', '').strip()
        
        if not identifier:
            return JsonResponse({'success': False, 'error': 'Identifier required'}, status=400)
        
        # Find user and their photo
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT 
                    u."UserID",
                    u."UserName",
                    u."UserPhoto"
                FROM "UserMaster" u
                WHERE (u."UserName" = %s OR u."UserCode" = %s OR u."Email" = %s)
                  AND u."IsActive" = TRUE
                  AND u."IsDeleted" IS NOT TRUE
            """, [identifier, identifier, identifier])
            
            user_data = cursor.fetchone()
            
            if not user_data:
                return JsonResponse({'success': False, 'error': 'User not found'}, status=404)
            
            user_id, user_name, user_photo = user_data
            
            if user_photo:
                # Convert binary photo to base64 data URI
                import base64
                photo_data_uri = f"data:image/jpeg;base64,{base64.b64encode(user_photo).decode('utf-8')}"
                
                return JsonResponse({
                    'success': True,
                    'user_id': user_id,
                    'user_name': user_name,
                    'user_photo': photo_data_uri
                })
            else:
                return JsonResponse({'success': False, 'error': 'No profile photo found'})
                
    except Exception as e:
        logger.error(f"Error in get_user_photo: {str(e)}", exc_info=True)
        return JsonResponse({'success': False, 'error': 'Internal server error'}, status=500)

@custom_login_required
def email_queue_status(request):
    """Admin view to monitor email queue status"""
    try:
        # Get user information using the same method as other pages
        sess = _get_custom_session_info(request)
        if not sess:
            messages.error(request, "Session expired. Please login again.")
            return redirect('login')
        
        # Get header context using the same helper function as other pages
        header_context = get_context(request)
        
        # Extract user information for the page
        user_name = header_context.get('user_name', 'User')
        user_email = sess.get('user_email', '')
        user_photo_src = header_context.get('user_photo_src', '')
        school_logo_src = header_context.get('school_logo_src', '')
        school_name = header_context.get('SchoolName', '')
        profile_name = header_context.get('profile_name', '')
        
        # Check if user is super admin
        is_super_admin = sess.get('profile_id') == 1
        
        # Get queue status from database with error handling
        try:
            status = database_email_queue.get_queue_status()
        except Exception as db_error:
            logger.error(f"Database connection error: {str(db_error)}")
            status = {
                'total_tasks': 0,
                'pending': 0,
                'processing': 0,
                'completed': 0,
                'failed': 0,
                'permanently_failed': 0,
                'worker_running': False,
                'email_types': {},
                'error': 'Database connection failed'
            }
        
        # Get recent email tasks from database with error handling
        try:
            recent_tasks = database_email_queue.get_recent_emails(limit=20)
        except Exception as db_error:
            logger.error(f"Error getting recent emails: {str(db_error)}")
            recent_tasks = []
        
        # Get email statistics with error handling
        try:
            email_stats = EmailTrackingManager.get_email_statistics()
        except Exception as db_error:
            logger.error(f"Error getting email statistics: {str(db_error)}")
            email_stats = {}
        
        # Get email activity list parameters
        search_term = request.GET.get('search', '').strip()
        page_number = safe_int(request.GET.get('page', 1))
        page_size = safe_int(request.GET.get('page_size', 10))
        school_id = request.GET.get('school_id')
        status = request.GET.get('status')
        email_type = request.GET.get('email_type')
        is_active = request.GET.get('is_active')
        sort_column = request.GET.get('sort_column', 'CreatedAt')
        sort_direction = request.GET.get('sort_direction', 'DESC')
        
        # Convert is_active to boolean if provided
        if is_active is not None:
            is_active = is_active.lower() == 'true'
        
        # Convert school_id to int if provided
        if school_id:
            try:
                school_id = int(school_id)
            except ValueError:
                school_id = None
        
        # Get email activity data using raw SQL (equivalent to stored procedure)
        from django.db import connection
        
        email_activities = []
        total_records = 0
        total_pages = 0
        previous_page = None
        next_page = None
        available_schools = []
        available_statuses = []
        available_email_types = []
        
        try:
            with connection.cursor() as cursor:
                # Calculate offset for pagination
                offset = (page_number - 1) * page_size
                
                # Build the main query
                base_query = """
                    SELECT 
                        et."EmailTrackingID",
                        et."EmailCode",
                        et."StudentCode",
                        et."ToEmail",
                        et."Status",
                        et."Priority",
                        et."AttemptCount",
                        et."MaxAttempts",
                        et."CreatedAt",
                        et."CompletedAt",
                        et."IsActive",
                        et."LastError" AS "ErrorMessage",
                        sm."SchoolName",
                        sm."SchoolCode" AS "SchoolCode",
                        -- Calculate additional fields
                        CASE 
                            WHEN et."Status" = 'Sent' THEN 'Completed'
                            WHEN et."Status" = 'Failed' AND et."AttemptCount" >= et."MaxAttempts" THEN 'Permanently Failed'
                            ELSE et."Status"
                        END AS "DisplayStatus",
                        EXTRACT(EPOCH FROM (COALESCE(et."CompletedAt", CURRENT_TIMESTAMP) - et."CreatedAt")) / 60 AS "ProcessingTimeMinutes",
                        -- Email type display name
                        CASE 
                            WHEN et."EmailCode" = 'ADMISSION_ACKNOWLEDGMENT' THEN 'Admission'
                            WHEN et."EmailCode" = 'PAYMENT_RECEIPT' THEN 'Payment'
                            WHEN et."EmailCode" = 'EMPLOYEE_REGISTRATION_CONFIRMATION' THEN 'Employee Registration'
                            ELSE et."EmailCode"
                        END AS "EmailTypeDisplay"
                    FROM "EmailTracking" et
                    LEFT JOIN "UserMaster" s ON s."UserCode" = et."StudentCode"
                    LEFT JOIN "SchoolMaster" sm ON sm."SchoolID" = COALESCE(s."SchoolID", et."SchoolID")
                    WHERE 1=1
                """
                
                # Add role-based filtering
                if profile_name != 'Super Admin':
                    base_query += ' AND COALESCE(s."SchoolID", et."SchoolID") = %s'
                
                # Add search filters
                params = []
                if profile_name != 'Super Admin':
                    params.append(sess.get('school_id'))
                
                if search_term:
                    base_query += """
                        AND (et."ToEmail" ILIKE %s OR
                             et."StudentCode" ILIKE %s OR
                             et."EmailCode" ILIKE %s OR
                             et."Status" ILIKE %s OR
                             sm."SchoolName" ILIKE %s)
                    """
                    search_param = f'%{search_term}%'
                    params.extend([search_param] * 5)
                
                if school_id:
                    base_query += ' AND COALESCE(s."SchoolID", et."SchoolID") = %s'
                    params.append(school_id)
                
                if status:
                    base_query += ' AND et."Status" = %s'
                    params.append(status)
                
                if email_type:
                    base_query += ' AND et."EmailCode" = %s'
                    params.append(email_type)
                
                if is_active is not None:
                    base_query += ' AND et."IsActive" = %s'
                    params.append(is_active)
                
                # Get total count
                count_query = f"SELECT COUNT(*) FROM ({base_query}) AS count_query"
                cursor.execute(count_query, params)
                total_records = cursor.fetchone()[0]
                
                # Add sorting
                valid_sort_columns = ['EmailTrackingID', 'EmailCode', 'StudentCode', 'ToEmail', 'Status', 'Priority', 'AttemptCount', 'CreatedAt', 'CompletedAt', 'SchoolName']
                if sort_column not in valid_sort_columns:
                    sort_column = 'CreatedAt'
                
                if sort_direction not in ['ASC', 'DESC']:
                    sort_direction = 'DESC'
                
                # Add ORDER BY clause
                if sort_column == 'SchoolName':
                    base_query += f' ORDER BY sm."SchoolName" {sort_direction}'
                else:
                    base_query += f' ORDER BY et."{sort_column}" {sort_direction}'
                
                # Add pagination
                base_query += " OFFSET %s LIMIT %s"
                params.extend([offset, page_size])
                
                # Execute main query
                cursor.execute(base_query, params)
                columns = [col[0] for col in cursor.description]
                email_activities = [dict(zip(columns, row)) for row in cursor.fetchall()]
            
            # Calculate pagination info
            total_pages = (total_records + page_size - 1) // page_size
            previous_page = page_number - 1 if page_number > 1 else None
            next_page = page_number + 1 if page_number < total_pages else None
            
            # Get available schools for filter (if super admin)
            if profile_name == 'Super Admin':
                from .models import SchoolMaster
                available_schools = SchoolMaster.objects.filter(is_deleted=False).values('school_id', 'school_name')
            
            # Get available statuses and email types for filters
            with connection.cursor() as cursor:
                cursor.execute('SELECT DISTINCT "Status" FROM "EmailTracking" WHERE "IsActive" = TRUE')
                available_statuses = [row[0] for row in cursor.fetchall()]
                
                cursor.execute('SELECT DISTINCT "EmailCode" FROM "EmailTracking" WHERE "IsActive" = TRUE')
                available_email_types = [row[0] for row in cursor.fetchall()]
                
        except Exception as db_error:
            logger.error(f"Error getting email activity list: {str(db_error)}")
            email_activities = []
            total_records = 0
        
        # Get current time for display
        from django.utils import timezone
        current_time = timezone.now()
        
        # Create context using the same structure as other pages
        context = {
            'user': sess,  # This is what the base template expects
            'user_name': user_name,
            'user_email': user_email,
            'user_photo_src': user_photo_src,
            'school_logo_src': school_logo_src,
            'school_name': school_name,
            'profile_name': profile_name,
            'is_super_admin': is_super_admin,
            'queue_status': status,
            'recent_tasks': recent_tasks,
            'email_stats': email_stats,
            'current_time': current_time,
            'email_activities': email_activities,
            'search_term': search_term,
            'page_number': page_number,
            'page_size': page_size,
            'total_records': total_records,
            'total_pages': total_pages,
            'previous_page': previous_page,
            'next_page': next_page,
            'sort_column': sort_column,
            'sort_direction': sort_direction,
            'available_schools': available_schools,
            'available_statuses': available_statuses,
            'available_email_types': available_email_types,
            'selected_school_id': school_id,
            'selected_status': status,
            'selected_email_type': email_type,
            'selected_is_active': is_active,
            'page_title': 'Email Track Dashboard',
            'breadcrumb': [
                {'name': 'Settings', 'url': '#'},
                {'name': 'Email Track', 'url': '#'}
            ]
        }
        
        return render(request, 'admin/email_queue_status.html', context)
        
    except Exception as e:
        logger.error(f"Error in email_queue_status: {str(e)}", exc_info=True)
        messages.error(request, f"Error loading email track page: {str(e)}")
        # Get user info for error case using the same method
        try:
            sess = _get_custom_session_info(request)
            header_context = get_context(request)
            
            error_context = {
                'user': sess or {},
                'user_name': header_context.get('user_name', 'User'),
                'user_email': sess.get('user_email', '') if sess else '',
                'user_photo_src': header_context.get('user_photo_src', ''),
                'school_logo_src': header_context.get('school_logo_src', ''),
                'school_name': header_context.get('SchoolName', ''),
                'profile_name': header_context.get('profile_name', ''),
                'is_super_admin': sess.get('profile_id') == 1 if sess else False,
                'queue_status': {'error': 'Failed to load data'},
                'recent_tasks': [],
                'email_stats': {},
                'current_time': timezone.now(),
                'email_activities': [],
                'search_term': '',
                'page_number': 1,
                'page_size': 10,
                'total_records': 0,
                'total_pages': 0,
                'previous_page': None,
                'next_page': None,
                'sort_column': 'CreatedAt',
                'sort_direction': 'DESC',
                'available_schools': [],
                'available_statuses': [],
                'available_email_types': [],
                'selected_school_id': None,
                'selected_status': None,
                'selected_email_type': None,
                'selected_is_active': None,
                'page_title': 'Email Track Dashboard',
                'breadcrumb': [
                    {'name': 'Settings', 'url': '#'},
                    {'name': 'Email Track', 'url': '#'}
                ]
            }
        except Exception:
            # Fallback if even the error handling fails
            error_context = {
                'user': {},
                'user_name': 'User',
                'user_email': '',
                'user_photo_src': '',
                'school_logo_src': '',
                'school_name': '',
                'profile_name': '',
                'is_super_admin': False,
                'queue_status': {'error': 'Failed to load data'},
                'recent_tasks': [],
                'email_stats': {},
                'current_time': timezone.now(),
                'email_activities': [],
                'search_term': '',
                'page_number': 1,
                'page_size': 10,
                'total_records': 0,
                'total_pages': 0,
                'previous_page': None,
                'next_page': None,
                'sort_column': 'CreatedAt',
                'sort_direction': 'DESC',
                'available_schools': [],
                'available_statuses': [],
                'available_email_types': [],
                'selected_school_id': None,
                'selected_status': None,
                'selected_email_type': None,
                'selected_is_active': None,
                'page_title': 'Email Track Dashboard',
                'breadcrumb': [
                    {'name': 'Settings', 'url': '#'},
                    {'name': 'Email Track', 'url': '#'}
                ]
            }
        
        return render(request, 'admin/email_queue_status.html', error_context)

@custom_login_required
def email_queue_status_api(request):
    """API endpoint to get email queue status"""
    try:
        status = database_email_queue.get_queue_status()
        return JsonResponse({'success': True, 'status': status})
        
    except Exception as e:
        logger.error(f"Error in email_queue_status_api: {str(e)}", exc_info=True)
        return JsonResponse({'success': False, 'error': 'Internal server error'}, status=500)

@custom_login_required
def email_track_debug(request):
    """Debug view to check user information"""
    try:
        # Get user information using the same method as other pages
        sess = _get_custom_session_info(request)
        header_context = get_context(request)
        
        debug_info = {
            'session_info': sess,
            'header_context': header_context,
            'user_name': header_context.get('user_name'),
            'user_email': sess.get('user_email') if sess else '',
            'user_photo_src': header_context.get('user_photo_src'),
            'school_logo_src': header_context.get('school_logo_src'),
            'school_name': header_context.get('SchoolName'),
            'profile_name': header_context.get('profile_name'),
            'is_super_admin': sess.get('profile_id') == 1 if sess else False,
            'profile_id': sess.get('profile_id') if sess else None,
            'session_keys': list(request.session.keys()),
        }
        
        return JsonResponse({'success': True, 'debug': debug_info})
        
    except Exception as e:
        logger.error(f"Error in email_track_debug: {str(e)}", exc_info=True)
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@custom_login_required
def retry_failed_emails(request):
    """Retry failed emails manually"""
    try:
        if request.method == 'POST':
            max_retries = safe_int(request.POST.get('max_retries', 10))
            retry_count = database_email_queue.retry_failed_emails(max_retries=max_retries)
            
            messages.success(request, f"Retried {retry_count} failed emails")
            return redirect('email_queue_status')
        else:
            return JsonResponse({'error': 'POST method required'}, status=405)
            
    except Exception as e:
        logger.error(f"Error in retry_failed_emails: {str(e)}", exc_info=True)
        messages.error(request, f"Error retrying emails: {str(e)}")
        return redirect('email_queue_status')

@custom_login_required
def cleanup_old_emails(request):
    """Clean up old completed emails"""
    try:
        if request.method == 'POST':
            days = safe_int(request.POST.get('days', 30))
            cleanup_count = database_email_queue.cleanup_old_emails(days=days)
            
            messages.success(request, f"Cleaned up {cleanup_count} old emails")
            return redirect('email_queue_status')
        else:
            return JsonResponse({'error': 'POST method required'}, status=405)
            
    except Exception as e:
        logger.error(f"Error in cleanup_old_emails: {str(e)}", exc_info=True)
        messages.error(request, f"Error cleaning up emails: {str(e)}")
        return redirect('email_queue_status')


# =============================================
# Menu Data Management Views
# =============================================



# =============================================
# Student Management Views
# =============================================

@custom_login_required
def view_students(request):
    """
    View Students page - Display students in card format with filtering and pagination
    """
    # Get user context for header
    context = get_context(request)
    
    # Also get session info for user object (needed for header template)
    sess = _get_custom_session_info(request)
    if sess:
        context['user'] = sess
    
    # Get user information
    user_id = request.session.get('UserId')
    school_id = request.session.get('SchoolID')
    profile_id = request.session.get('ProfileID')
    
    if not user_id:
        messages.error(request, "Please login to access student data")
        return redirect('login')
    
    if not school_id:
        messages.error(request, "School ID is required to access student data")
        return redirect('login')
    
    if not profile_id:
        messages.error(request, "Profile ID is required to access student data")
        return redirect('login')
    
    # Get query parameters
    page = safe_int(request.GET.get('page', 1))
    per_page = safe_int(request.GET.get('per_page', 10))  # 10 cards per page for better performance
    search = request.GET.get('search', '').strip()
    class_id = request.GET.get('class_id', '')  # No default filter
    section_id = request.GET.get('section_id', '')  # No default filter
    gender = request.GET.get('gender', '')
    category = request.GET.get('category', '')
    from_date = request.GET.get('from_date', '')
    to_date = request.GET.get('to_date', '')
    
    # Convert date strings to datetime objects
    from_date_obj = None
    to_date_obj = None
    if from_date:
        try:
            from_date_obj = safe_strptime(from_date, '%Y-%m-%d').date()
        except ValueError:
            from_date = ''
    if to_date:
        try:
            to_date_obj = safe_strptime(to_date, '%Y-%m-%d').date()
        except ValueError:
            to_date = ''
    
    # Initialize variables
    students = []
    total_count = 0
    start_index = 0
    end_index = 0
    has_next = False
    has_prev = False
    
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT * FROM "Proc_Student_Cards_Full_Get"(
                    %s, %s, %s, %s, %s, %s
                )
            """, [
                school_id,
                int(class_id) if class_id else None,
                int(section_id) if section_id else None,
                search if search else None,
                page,
                per_page
            ])
            
            columns = [col[0] for col in cursor.description]
            raw_students = cursor.fetchall()
            
            for row in raw_students:
                student = dict(zip(columns, row))
                
                # Process Photo (VARBINARY) to Base64
                if student.get('Photo') and isinstance(student['Photo'], bytes):
                    try:
                        student['PhotoBase64'] = base64.b64encode(student['Photo']).decode('utf-8')
                    except Exception as e:
                        logger.error(f"Error encoding photo for student {student.get('StudentCode')}: {e}")
                        student['PhotoBase64'] = None
                else:
                    student['PhotoBase64'] = None
                
                # Process SchoolLogo (VARBINARY) to Base64
                if student.get('SchoolLogo') and isinstance(student['SchoolLogo'], bytes):
                    try:
                        student['SchoolLogoBase64'] = base64.b64encode(student['SchoolLogo']).decode('utf-8')
                    except Exception as e:
                        logger.error(f"Error encoding school logo: {e}")
                        student['SchoolLogoBase64'] = None
                else:
                    student['SchoolLogoBase64'] = None
                
                # Format dates
                if student.get('AdmissionDate'):
                    student['AdmissionDateFormatted'] = student['AdmissionDate'].strftime('%Y-%m-%d')
                else:
                    student['AdmissionDateFormatted'] = 'N/A'
                    
                if student.get('DateOfBirth'):
                    student['DateOfBirthFormatted'] = student['DateOfBirth'].strftime('%Y-%m-%d')
                else:
                    student['DateOfBirthFormatted'] = 'N/A'
                
                students.append(student)
            
            # Get total count from first row (TotalCount column)
            if students:
                total_count = students[0].get('TotalCount', 0)
            
    except Exception as e:
        logger.error(f"Error fetching students: {str(e)}", exc_info=True)
        messages.error(request, "Error loading student data. Please try again.")
        students = []
        total_count = 0
    
    # Calculate pagination info
    start_index = (page - 1) * per_page + 1 if students else 0
    end_index = min(start_index + len(students) - 1, total_count) if students else 0
    has_next = end_index < total_count
    has_prev = page > 1
    
    # Get selected student card template from Template Management
    card_template = 'student_card_horizontal_1.html'
    try:
        with connection.cursor() as cursor:
            cursor.execute('SELECT "TemplateType", "TemplateFile" FROM "Proc_Template_Preference_Get"(%s)', [school_id])
            for row in cursor.fetchall():
                if row[0] == 'StudentCard':
                    card_template = row[1]
                    break
    except Exception as e:
        logger.error(f"Error fetching student card template: {e}")
    
    # Add pagination and filter context
    context.update({
        'students': students,
        'total_count': total_count,
        'page': page,
        'per_page': per_page,
        'start_index': start_index,
        'end_index': end_index,
        'has_next': has_next,
        'has_prev': has_prev,
        'search': search,
        'class_id': class_id,
        'section_id': section_id,
        'gender': gender,
        'category': category,
        'status': '',
        'from_date': from_date,
        'to_date': to_date,
        'dark_mode': request.session.get('dark_mode', False),
        'card_template': card_template
    })
    
    return render(request, 'view_students_cards.html', context)


def _is_mobile_request(request):
    """Helper function to detect mobile requests"""
    user_agent = request.META.get('HTTP_USER_AGENT', '')
    return 'Mobile' in user_agent or 'Android' in user_agent or 'iPhone' in user_agent


def debug_view(view_func):
    @wraps(view_func)
    def _wrapped_view(request, *args, **kwargs):
        try:
            return view_func(request, *args, **kwargs)
        except Exception as e:
            import traceback
            return HttpResponse(f"Debug Error: {str(e)} <br> <pre>{traceback.format_exc()}</pre>")
    return _wrapped_view

@custom_login_required
@debug_view
def view_teachers(request):
    """
    View Teachers/Staff page - Display teachers/staff in table format with filtering and pagination
    """
    # Get user context for header
    context = get_context(request)
    
    # Also get session info for user object (needed for header template)
    sess = _get_custom_session_info(request)
    if sess:
        context['user'] = sess
    
    # Get user information
    user_id = request.session.get('UserId')
    school_id = request.session.get('SchoolID')
    profile_id = request.session.get('ProfileID')
    
    # Robust Super Admin check: ProfileID 1
    is_super_admin = (str(profile_id) == '1')
    selected_encrypted_school_id = None
    schools = []
    
    if is_super_admin:
        # Fetch schools using Global API logic
        try:
            from .utils import get_school_dropdown
            raw_schools = get_school_dropdown()
            
            for s in raw_schools:
                encrypted_sid = encrypt_id(s['SchoolID'])
                schools.append({
                    'SchoolID': s['SchoolID'],
                    'SchoolName': s.get('DisplayName', s['SchoolName']), 
                    'SchoolCode': s.get('SchoolCode'),
                    'EncryptedSchoolID': encrypted_sid
                })
        except Exception as e:
            logger.error(f"Error fetching schools for view teachers: {str(e)}")

        # Check for school_id in GET request
        get_school_id = request.GET.get('school_id')
        if get_school_id:
            # Try as plain integer first (from filter form)
            try:
                school_id = int(get_school_id)
                # Find the encrypted ID for the selected school
                for s in schools:
                    if s['SchoolID'] == school_id:
                        selected_encrypted_school_id = s['EncryptedSchoolID']
                        break
                
                if selected_encrypted_school_id:
                    # Redirect to encrypted URL
                    params = request.GET.copy()
                    params['school_id'] = selected_encrypted_school_id
                    return redirect(f"{request.path}?{params.urlencode()}")
            except ValueError:
                # Try decrypting (from URL)
                decrypted_id = decrypt_id(get_school_id)
                if decrypted_id:
                    school_id = int(decrypted_id)
                    # Find the encrypted ID for the selected school for context/UI highlights
                    for s in schools:
                        if s['SchoolID'] == school_id:
                            selected_encrypted_school_id = s['EncryptedSchoolID']
                            break
                else:
                    school_id = None
        else:
            school_id = None
            
    # Fetch Countries for location filter
    countries_list = []
    try:
        with connection.cursor() as cursor:
            cursor.execute(
                'SELECT "Geog_Id", "Geog_Name" FROM "Geographical_Master" '
                'WHERE "Geog_Type" = \'Country\' AND "IsDeleted" = FALSE '
                'ORDER BY CASE WHEN "Geog_Name" = \'India\' THEN 0 ELSE 1 END, "Geog_Name"'
            )
            countries_list = [{'id': row[0], 'name': row[1]} for row in cursor.fetchall()]
    except Exception as e:
        logger.error(f"Error fetching countries for view teachers: {str(e)}")
    
    if not user_id:
        messages.error(request, "Please login to access teacher data")
        return redirect('login')
    
    if not school_id and not is_super_admin:
        messages.error(request, "School ID is required to access teacher data")
        return redirect('login')
    
    if not profile_id:
        messages.error(request, "Profile ID is required to access teacher data")
        return redirect('login')
    
    # Get query parameters
    page = safe_int(request.GET.get('page', 1))
    per_page = safe_int(request.GET.get('per_page', 25))
    
    # Optimize per_page for mobile devices
    is_mobile = _is_mobile_request(request)
    if is_mobile and per_page > 15:
        per_page = 15  # Reduce records per page on mobile for better performance
    search = request.GET.get('search', '').strip()
    employee_code = request.GET.get('employee_code', '').strip()
    profile_name_filter = request.GET.get('profile_name', '').strip()
    mobile_no = request.GET.get('mobile_no', '').strip()
    email = request.GET.get('email', '').strip()
    country = request.GET.get('country', '').strip()
    state = request.GET.get('state', '').strip()
    district = request.GET.get('district', '').strip()
    pincode = request.GET.get('pincode', '').strip()
    religion = request.GET.get('religion', '').strip()
    national_id = request.GET.get('national_id', '').strip()
    gender = request.GET.get('gender', '').strip()
    employee_name = request.GET.get('employee_name', '').strip()
    status = request.GET.get('status', '').strip()
    from_date = request.GET.get('from_date', '').strip()
    to_date = request.GET.get('to_date', '').strip()
    order_by = request.GET.get('order_by', 'EmployeeCode')
    order_direction = request.GET.get('order_direction', 'ASC')
    
    # Optimize search parameters - avoid duplicate searches
    if employee_code and employee_code == employee_name == profile_name_filter == mobile_no == email:
        # If all search fields have the same value, use only employee_code to avoid duplicate search
        employee_name = ''
        profile_name_filter = ''
        mobile_no = ''
        email = ''
    
    # Convert date strings to date objects
    from_date_obj = None
    to_date_obj = None
    if from_date:
        try:
            from_date_obj = safe_strptime(from_date, '%Y-%m-%d').date()
        except ValueError:
            from_date = ''
    if to_date:
        try:
            to_date_obj = safe_strptime(to_date, '%Y-%m-%d').date()
        except ValueError:
            to_date = ''
    
    # Initialize variables
    teachers = []
    total_count = 0
    active_count = 0
    start_index = 0
    end_index = 0
    has_next = False
    has_prev = False
    
    # Determine if any filters are applied
    has_filters = any([
        search, employee_code, profile_name_filter, mobile_no, email,
        country, state, district, pincode, religion,
        national_id, gender, employee_name, status,
        from_date, to_date
    ])

    # Super Admin Handling: Only show message if no school selected AND no filters applied
    show_school_filter_message = is_super_admin and not school_id and not has_filters

    if show_school_filter_message:
        context.update({
            'teachers': [],
            'total_count': 0,
            'active_count': 0,
            'page': page,
            'per_page': per_page,
            'start_index': 0,
            'end_index': 0,
            'has_next': False,
            'has_prev': False,
            'schools': schools,
            'is_super_admin': is_super_admin,
            'selected_school_id': None,
            'selected_encrypted_school_id': None,
            'search': search,
            'employee_code': employee_code,
            'filter_profile_name': profile_name_filter,
            'mobile_no': mobile_no,
            'email': email,
            'country': country,
            'state': state,
            'district': district,
            'pincode': pincode,
            'religion': religion,
            'national_id': national_id,
            'gender': gender,
            'employee_name': employee_name,
            'status': status,
            'from_date': from_date,
            'to_date': to_date,
            'order_by': order_by,
            'order_direction': order_direction,
            'dark_mode': request.session.get('dark_mode', False),
            'show_school_filter_message': True
        })
        return render(request, 'view_teachers.html', context)
    
    try:
        # Add performance logging
        import time
        start_time = time.time()
        
        with connection.cursor() as cursor:
            # Use the provided stored procedure Proc_Employee_List
            cursor.execute("""
                SELECT * FROM "Proc_Employee_List"(
                    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s
                )
            """, [
                school_id, user_id, profile_id,
                employee_code if employee_code else None,
                profile_name_filter if profile_name_filter else None,
                mobile_no if mobile_no else None,
                email if email else None,
                country if country else None,
                state if state else None,
                district if district else None,
                pincode if pincode else None,
                religion if religion else None,
                national_id if national_id else None,
                gender if gender else None,
                employee_name if employee_name else None,
                status if status else None,
                from_date_obj, to_date_obj,
                order_by, order_direction,
                page, per_page
            ])
            
            # Get the result set (employee data + counts)
            columns = [col[0] for col in cursor.description]
            raw_teachers = cursor.fetchall()
            
            # Log database query performance
            db_time = time.time() - start_time
            print(f"Database query time: {db_time:.3f}s - Mobile: {is_mobile} - Search: {employee_code}")
            
            # Get counts from the first row if available
            if raw_teachers:
                total_count = raw_teachers[0][-1]  # Last column is TotalCount
                # ActiveCount is not returned by new proc, default to 0 or remove usage
                active_count = 0 
            
            for row in raw_teachers:
                # Map row by column names, excluding TotalCount if not needed in dict
                teacher = dict(zip(columns[:-1], row[:-1])) # Exclude TotalCount col
                
                # Process photo data efficiently - skip on mobile for better performance
                if not is_mobile:
                    photo_data = None
                    photo_field_names = ['Photo', 'UserPhoto', 'PhotoData', 'Image', 'Avatar']
                    
                    for field_name in photo_field_names:
                        if teacher.get(field_name):
                            photo_data = teacher.get(field_name)
                            break
                    
                    if photo_data and isinstance(photo_data, bytes):
                        try:
                            import base64
                            teacher['PhotoBase64'] = base64.b64encode(photo_data).decode('utf-8')
                        except Exception as e:
                            logger.error(f"Error processing photo for teacher {teacher.get('EmployeeID')}: {e}")
                            teacher['PhotoBase64'] = None
                    else:
                        teacher['PhotoBase64'] = None
                else:
                    # Skip photo processing on mobile for better performance
                    teacher['PhotoBase64'] = None
                
                # Format dates
                if teacher.get('DateOfJoining'):
                    teacher['DateOfJoiningFormatted'] = teacher['DateOfJoining'].strftime('%Y-%m-%d')
                else:
                    teacher['DateOfJoiningFormatted'] = 'N/A'
                    
                if teacher.get('DateOfBirth'):
                    teacher['DateOfBirthFormatted'] = teacher['DateOfBirth'].strftime('%Y-%m-%d')
                else:
                    teacher['DateOfBirthFormatted'] = 'N/A'
                
                if teacher.get('CreatedAt'):
                    teacher['CreatedAtFormatted'] = teacher['CreatedAt'].strftime('%Y-%m-%d')
                else:
                    teacher['CreatedAtFormatted'] = 'N/A'
                
                # Set Status based on IsDeleted
                teacher['Status'] = 'Active'
                
                teachers.append(teacher)
            
            # Fetch and map Core Subjects for the teachers on this page
            if teachers:
                teacher_ids = [t.get('EmployeeID') for t in teachers if t.get('EmployeeID')]
                if teacher_ids:
                    try:
                        with connection.cursor() as subject_cursor:
                            placeholders = ', '.join(['%s'] * len(teacher_ids))
                            subject_cursor.execute(f"""
                                SELECT ecs."EmployeeID", sm."SubjectName"
                                FROM "EmployeeCoreSubjects" ecs
                                JOIN "SubjectMaster" sm ON ecs."SubjectID" = sm."SubjectID"
                                WHERE ecs."EmployeeID" IN ({placeholders})
                            """, teacher_ids)
                            
                            # Group subjects by EmployeeID
                            teacher_subjects = {}
                            for eid, sname in subject_cursor.fetchall():
                                if eid not in teacher_subjects:
                                    teacher_subjects[eid] = []
                                teacher_subjects[eid].append(sname)
                            
                            # Add comma-separated string to teacher objects
                            for t in teachers:
                                eid = t.get('EmployeeID')
                                if eid in teacher_subjects:
                                    t['CoreSubjects'] = ', '.join(teacher_subjects[eid])
                                else:
                                    t['CoreSubjects'] = None
                    except Exception as subject_e:
                        logger.error(f"Error fetching core subjects for listing: {str(subject_e)}")
            
    except Exception as e:
        logger.error(f"Error fetching teachers: {str(e)}", exc_info=True)
        messages.error(request, "Error loading teacher data. Please try again.")
        teachers = []
        total_count = 0
        active_count = 0
    
    # Calculate pagination info
    total_pages = (total_count + per_page - 1) // per_page if total_count > 0 else 1
    start_index = (page - 1) * per_page + 1 if teachers else 0
    end_index = min(start_index + len(teachers) - 1, total_count) if teachers else 0
    has_next = end_index < total_count
    has_prev = page > 1
    
    # Generate page range for the UI (e.g., current page +/- 2)
    pagination_range = []
    if total_pages <= 7:
        pagination_range = range(1, total_pages + 1)
    else:
        if page <= 4:
            pagination_range = range(1, 6)
        elif page >= total_pages - 3:
            pagination_range = range(total_pages - 4, total_pages + 1)
        else:
            pagination_range = range(page - 2, page + 3)

    # Add pagination and filter context
    context.update({
        'teachers': teachers,
        'total_count': total_count,
        'active_count': active_count,
        'page': page,
        'per_page': per_page,
        'total_pages': total_pages,
        'pagination_range': pagination_range,
        'start_index': start_index,
        'end_index': end_index,
        'has_next': has_next,
        'has_prev': has_prev,
        'schools': schools,
        'is_super_admin': is_super_admin,
        'selected_school_id': school_id,
        'selected_encrypted_school_id': selected_encrypted_school_id,
        'search': search,
        'employee_code': employee_code,
        'filter_profile_name': profile_name_filter,
        'mobile_no': mobile_no,
        'email': email,
        'country': country,
        'state': state,
        'district': district,
        'pincode': pincode,
        'religion': religion,
        'national_id': national_id,
        'gender': gender,
        'employee_name': employee_name,
        'status': status,
        'from_date': from_date,
        'to_date': to_date,
        'order_by': order_by,
        'order_direction': order_direction,
        'countries_list': countries_list,
        'dark_mode': request.session.get('dark_mode', False),
        'profile_name': context.get('profile_name')
    })
    
    # Log total processing time
    total_time = time.time() - start_time
    print(f"Total view processing time: {total_time:.3f}s - Mobile: {is_mobile} - Teachers: {len(teachers)}")
    
    return render(request, 'view_teachers.html', context)


@custom_login_required
def export_teachers(request):
    """Export teachers list to Excel/CSV"""
    import csv
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    
    # Check permissions
    # Get user context
    user_id = request.session.get('UserId')
    school_id = request.session.get('SchoolID')
    profile_id = request.session.get('ProfileID')
    if not user_id:
        return redirect('login')

    # Super Admin Handling
    is_super_admin = False
    if profile_id == 1:
        is_super_admin = True
        # Allow Super Admin to filter by school
        filter_school_id = request.GET.get('school_id')
        if filter_school_id:
            try:
                school_id = int(filter_school_id)
            except ValueError:
                pass # Use existing school_id or None

    # Get filter parameters
    employee_code = request.GET.get('employee_code', '').strip()
    profile_name = request.GET.get('profile_name', '').strip()
    mobile_no = request.GET.get('mobile_no', '').strip()
    email = request.GET.get('email', '').strip()
    country = request.GET.get('country', '').strip()
    state = request.GET.get('state', '').strip()
    district = request.GET.get('district', '').strip()
    pincode = request.GET.get('pincode', '').strip()
    religion = request.GET.get('religion', '').strip()
    national_id = request.GET.get('national_id', '').strip()
    gender = request.GET.get('gender', '').strip()
    employee_name = request.GET.get('employee_name', '').strip()
    status = request.GET.get('status', '').strip()
    
    # Sort parameters
    order_by = request.GET.get('sort', 'EmployeeName')
    order_direction = request.GET.get('order', 'asc')

    # Date filters
    from_date = request.GET.get('from_date', '')
    to_date = request.GET.get('to_date', '')
    
    # Optimize search parameters
    if employee_code and employee_code == employee_name == profile_name == mobile_no == email:
        employee_name = ''
        profile_name = ''
        mobile_no = ''
        email = ''

    # Date parsing
    from_date_obj = None
    to_date_obj = None
    if from_date:
        try:
            from_date_obj = safe_strptime(from_date, '%Y-%m-%d').date()
        except ValueError:
            pass
    if to_date:
        try:
            to_date_obj = safe_strptime(to_date, '%Y-%m-%d').date()
        except ValueError:
            pass

    # Fetch Data
    employees = []
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT * FROM "Proc_Employee_List"(
                    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s
                )
            """, [
                school_id, user_id, profile_id,
                employee_code if employee_code else None,
                profile_name if profile_name else None,
                mobile_no if mobile_no else None,
                email if email else None,
                country if country else None,
                state if state else None,
                district if district else None,
                pincode if pincode else None,
                religion if religion else None,
                national_id if national_id else None,
                gender if gender else None,
                employee_name if employee_name else None,
                status if status else None,
                from_date_obj, to_date_obj,
                order_by, order_direction,
                1, 100000 # Large page size for export
            ])
            
            columns = [col[0] for col in cursor.description]
            rows = cursor.fetchall()
            for row in rows:
                employees.append(dict(zip(columns, row)))
                
    except Exception as e:
        logger.error(f"Error fetching employees for export: {e}")
        return HttpResponse("Error generating export", status=500)

    # Format Data for Export
    export_data = []
    for emp in employees:
        row = {
            'Employee Code': emp.get('EmployeeCode'),
            'Name': emp.get('EmployeeName'),
            'Designation': emp.get('ProfileName'),
            'Mobile': emp.get('MobileNo'),
            'Email': emp.get('Email'),
            'Gender': emp.get('Gender'),
            'DOB': emp.get('DateOfBirth'),
            'Join Date': emp.get('DateOfJoining'),
            'Status': 'Active' if emp.get('IsActive') else 'Inactive',
            'School': emp.get('SchoolName'),
            'Religion': emp.get('Religion'),
            'National ID': emp.get('NationalIdNumber'),
            'Address': emp.get('CurrentAddress'),
            'City': emp.get('CurrentCity'),
            'District': emp.get('CurrentDistrict'),
            'State': emp.get('CurrentState'),
            'Country': emp.get('CurrentCountry'),
            'Pincode': emp.get('CurrentPincode'),
        }
        export_data.append(row)

    format_type = request.GET.get('format', 'excel')
    filename = f"Staff_List_{datetime.now().strftime('%Y%m%d_%H%M%S')}"

    if format_type == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
        
        if export_data:
            writer = csv.DictWriter(response, fieldnames=export_data[0].keys())
            writer.writeheader()
            writer.writerows(export_data)
        return response
        
    else: # Excel
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Staff List"
        
        if export_data:
            # Headers
            headers = list(export_data[0].keys())
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
                
            # Data
            for row_num, row_data in enumerate(export_data, 2):
                for col_num, (key, value) in enumerate(row_data.items(), 1):
                    ws.cell(row=row_num, column=col_num, value=str(value) if value is not None else "")
            
            # Auto-width
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(response)
        return response


@custom_login_required
def student_profile(request):
    """
    Student Profile page - Individual student details
    """
    # Get user context for header
    context = get_context(request)
    
    # Get student ID from URL parameters
    student_id = request.GET.get('id')
    
    if not student_id:
        messages.error(request, "Student ID is required")
        return redirect('view_students')
    
    # TODO: Implement student profile view
    context.update({
        'student_id': student_id,
        'dark_mode': request.session.get('dark_mode', False)
    })
    
    return render(request, 'student_profile.html', context)



# Student Promotion views moved to student_promote_views.py


# Error testing views (remove in production)
def test_404(request):
    """Test 404 error page"""
    from django.http import Http404
    raise Http404("This is a test 404 error")

def test_500(request):
    """Test 500 error page"""
    raise ValueError("This is a test 500 error")

def test_403(request):
    """Test 403 error page"""
    from django.core.exceptions import PermissionDenied
    raise PermissionDenied("This is a test 403 error")

def test_400(request):
    """Test 400 error page"""
    from django.core.exceptions import BadRequest
    raise BadRequest("This is a test 400 error")

def custom_404_view(request):
    """Custom 404 view that works in debug mode"""
    from core.error_handlers import safe_render_error
    return safe_render_error(request, 'errors/404.html', 404)


@custom_login_required
def add_employee_view(request):
    """Display the Add Employee form"""
    if request.method == 'GET':
        # Load dark mode preference
        dark_mode = request.session.get('dark_mode', False)
        
        # Fetch context data for header (same as student admission page)
        context = get_context(request)
        
        # Get user info
        school_id = request.session.get('SchoolID')
        profile_id = request.session.get('ProfileID')
        
        schools = []
        selected_school_id = school_id
        is_super_admin = (str(profile_id) == '1')
        
        selected_encrypted_school_id = None

        # Super Admin Logic: Fetch schools and allow selection
        if is_super_admin:
            # Check for school_id in GET request
            get_school_id = request.GET.get('school_id')
            if get_school_id:
                # Try decrypting first
                decrypted_id = decrypt_id(get_school_id)
                if decrypted_id:
                    selected_school_id = int(decrypted_id)
                else:
                    # Validation for integer ID (fallback or legacy)
                    try:
                        selected_school_id = int(get_school_id)
                    except ValueError:
                        selected_school_id = None
            else:
                selected_school_id = None
            
            # Fetch all active schools using utility
            try:
                from .utils import get_school_dropdown
                raw_schools = get_school_dropdown()
                
                # Transform to match template expectation (Name -> DisplayName)
                schools = []
                for s in raw_schools:
                    encrypted_sid = encrypt_id(s['SchoolID'])
                    schools.append({
                        'SchoolID': s['SchoolID'],
                        'SchoolName': s.get('DisplayName', s['SchoolName']), 
                        'SchoolCode': s.get('SchoolCode'),
                        'EncryptedSchoolID': encrypted_sid
                    })
                    
                    if selected_school_id == s['SchoolID']:
                        selected_encrypted_school_id = encrypted_sid

            except Exception as e:
                logger.error(f"Error fetching schools for super admin: {str(e)}")
        
        # Fetch data based on selected school
        salary_components = []
        profile_roles = []
        subjects = []
        
        # Fetch profile roles and global subjects (independent of school)
        try:
            with connection.cursor() as cursor:
                # Fetch profile roles
                cursor.execute('SELECT * FROM "Proc_Add_Staff_Profile_Role_Get"()')
                profile_columns = [col[0] for col in cursor.description]
                profile_roles = [dict(zip(profile_columns, row)) for row in cursor.fetchall()]

                # Fetch global subject specializations
                cursor.execute("""
                    SELECT "SpecializationID" as "SubjectID", "SpecializationName" as "SubjectName" 
                    FROM "SubjectSpecializationMaster" 
                    WHERE "IsDeleted" = false
                    ORDER BY "SpecializationName"
                """)
                columns = [col[0] for col in cursor.description]
                subjects = [dict(zip(columns, row)) for row in cursor.fetchall()]
        except Exception as e:
            logger.error(f"Error fetching profile roles or subjects: {str(e)}")

        if selected_school_id:
            try:
                with connection.cursor() as cursor:
                    # Fetch salary components
                    cursor.execute("""
                        SELECT "ComponentID", "ComponentName", "ComponentType"
                        FROM "SalaryComponentMaster" 
                        WHERE "SchoolID" = %s AND "IsDeleted" = false
                        ORDER BY "ComponentType", "ComponentName"
                    """, [selected_school_id])
                    
                    columns = [col[0] for col in cursor.description]
                    salary_components = [dict(zip(columns, row)) for row in cursor.fetchall()]

            except Exception as e:
                logger.error(f"Error fetching employee form data: {str(e)}")
        
        context.update({
            'dark_mode': dark_mode,
            'salary_components': salary_components,
            'profile_roles': profile_roles,
            'subjects': subjects,
            'schools': schools,
            'selected_school_id': selected_school_id,
            'selected_encrypted_school_id': selected_encrypted_school_id,
            'is_super_admin': is_super_admin,
        })
        
        return render(request, 'add_employee.html', context)
    return HttpResponseForbidden()


@custom_login_required
def assign_classes_view(request):
    """Display the Assign Classes page with teacher list"""
    if request.method == 'GET':
        # Load dark mode preference
        dark_mode = request.session.get('dark_mode', False)
        
        # Get user context for header
        context = get_context(request)
        
        # Get user information
        user_id = request.session.get('UserId')
        school_id = request.session.get('SchoolID')
        profile_id = request.session.get('ProfileID')
        
        if not user_id or not school_id:
            messages.error(request, "Please login to access this page")
            return redirect('login')
        
        teachers = []
        academic_years = []
        classes = []
        all_sections = {}  # Dictionary to store sections by class ID
        all_subjects = {}  # Dictionary to store subjects by class ID
        
        try:
            with connection.cursor() as cursor:
                # Get all dropdown data in one call
                cursor.execute('SELECT * FROM "Proc_AssignClasses_PageLoad"(%s)', [school_id])
                data_row = cursor.fetchone()
                
                if data_row:
                    # Parse JSON results
                    # Ensure we handle both string and list/dict returns depending on psycopg2 adaptation
                    teachers = data_row[0] if isinstance(data_row[0], list) else json.loads(data_row[0])
                    academic_years = data_row[1] if isinstance(data_row[1], list) else json.loads(data_row[1])
                    classes = data_row[2] if isinstance(data_row[2], list) else json.loads(data_row[2])
                    sections_list = data_row[3] if isinstance(data_row[3], list) else json.loads(data_row[3])
                    subjects_list = data_row[4] if isinstance(data_row[4], list) else json.loads(data_row[4])
                    
                    # Group sections by class ID
                    all_sections = {}
                    for sec in sections_list:
                        class_id = str(sec.get('class_id'))
                        if class_id not in all_sections:
                            all_sections[class_id] = []
                        all_sections[class_id].append({
                            "id": sec.get('id'),
                            "name": sec.get('name')
                        })
                        
                    # Group subjects by class ID
                    all_subjects = {}
                    for sub in subjects_list:
                        class_id = str(sub.get('ClassId'))
                        if class_id not in all_subjects:
                            all_subjects[class_id] = []
                        all_subjects[class_id].append({
                            "SubjectID": sub.get('SubjectID'),
                            "SubjectCode": sub.get('SubjectCode'),
                            "SubjectName": sub.get('SubjectName')
                        })
                
        except Exception as e:
            logger.error(f"Error fetching data for assign classes: {str(e)}")
            messages.error(request, "Error loading data")
        
        context.update({
            'dark_mode': dark_mode,
            'teachers': teachers,
            'academic_years': academic_years,
            'classes': classes,
            'all_sections': all_sections,
            'all_subjects': all_subjects,
        })
        
        return render(request, 'assign_classes.html', context)
    return HttpResponseForbidden()


@custom_login_required
def assign_classes_submit(request):
    """Handle teacher class assignment submission using stored procedure"""
    if request.method != 'POST':
        return JsonResponse({'status': 'FAILED', 'message': 'Invalid request method'})
    
    # Add debug logging
    logger.info(f"Assign classes submit called with method: {request.method}")
    logger.info(f"Request data: {dict(request.POST)}")
    
    try:
        # Get form data
        school_id = request.session.get('SchoolID')
        user_id = request.session.get('UserId')
        
        if not school_id or not user_id:
            return JsonResponse({'status': 'FAILED', 'message': 'Session expired. Please login again.'})
        
        # Required fields
        teacher_id = request.POST.get('teacher_id')
        class_id = request.POST.get('class_id')
        section_id = request.POST.get('section_id')
        subject_id = request.POST.get('subject_id')
        academic_year = request.POST.get('academic_year')
        is_class_teacher = request.POST.get('is_class_teacher', 'false').lower() == 'true'
        
        # Validate required fields
        if not teacher_id:
            return JsonResponse({'status': 'FAILED', 'message': 'Teacher is required'})
        if not class_id:
            return JsonResponse({'status': 'FAILED', 'message': 'Class is required'})
        if not subject_id:
            return JsonResponse({'status': 'FAILED', 'message': 'Subject is required'})
        if not academic_year:
            return JsonResponse({'status': 'FAILED', 'message': 'Academic Year is required'})
        
        # Check for class teacher conflict if marking as class teacher
        if is_class_teacher:
            with connection.cursor() as cursor:
                # Use PostgreSQL check function
                cursor.execute("""
                    SELECT * FROM "Proc_Check_Class_Teacher_Conflict"(%s, %s, %s, %s)
                """, [teacher_id, academic_year, school_id, class_id])
                
                conflict_row = cursor.fetchone()
                
                if conflict_row and conflict_row[0] > 0:
                    conflict_class_name = conflict_row[1]
                    return JsonResponse({
                        'Status': 'ERROR',
                        'Message': f'This teacher is already assigned as class teacher for {conflict_class_name} in academic year {academic_year}. A teacher can only be class teacher for one class per academic year.'
                    })
        
        # Convert to integers
        try:
            teacher_id = int(teacher_id)
            class_id = int(class_id)
            subject_id = int(subject_id)
            section_id = int(section_id) if section_id else None
        except (ValueError, TypeError):
            return JsonResponse({'status': 'FAILED', 'message': 'Invalid data format'})
        
        # Optional fields
        start_date = request.POST.get('start_date') or None
        end_date = request.POST.get('end_date') or None
        assignment_order = request.POST.get('assignment_order')
        remarks = request.POST.get('remarks', '').strip() or None
        
        try:
            assignment_order = int(assignment_order) if assignment_order else None
        except (ValueError, TypeError):
            assignment_order = None
        
        # Call stored procedure
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT * FROM "Proc_AssignTeacherToClass_set"(
                    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s
                )
            """, [
                None, # AssignmentID is NULL for insert
                school_id, teacher_id, class_id, section_id, subject_id, academic_year,
                start_date, end_date, is_class_teacher, assignment_order, remarks, 'INSERT', user_id
            ])
            
            result = cursor.fetchone()
            if result:
                # Result is (Status, Message)
                return JsonResponse({
                    'Status': result[0],
                    'Message': result[1]
                })
            else:
                logger.error("No result returned from stored procedure")
                return JsonResponse({'status': 'FAILED', 'message': 'No response from server'})
                
    except Exception as e:
        logger.error(f"Error in assign_classes_submit: {str(e)}")
        return JsonResponse({'status': 'FAILED', 'message': f'Error: {str(e)}'})


@custom_login_required
def assign_classes_test(request):
    """Test endpoint for debugging"""
    school_id = request.session.get('SchoolID')
    user_id = request.session.get('UserID')
    
    # Test database connections
    test_results = {
        'status': 'SUCCESS',
        'message': 'Test endpoint is working correctly',
        'session_data': {
            'school_id': school_id,
            'user_id': user_id
        },
        'api_tests': {}
    }
    
    try:
        # Test Academic Years
        with connection.cursor() as cursor:
            if school_id:
                cursor.execute("""
                    SELECT AcademicYearID, AcademicYear 
                    FROM AcademicYear 
                    WHERE SchoolID = %s
                """, [school_id])
                academic_years = cursor.fetchall()
                test_results['api_tests']['academic_years'] = {
                    'count': len(academic_years),
                    'data': [{'id': r[0], 'name': r[1]} for r in academic_years]
                }
            else:
                # Test with SchoolID=3 as fallback
                cursor.execute("""
                    SELECT AcademicYearID, AcademicYear 
                    FROM AcademicYear 
                    WHERE SchoolID = 3
                """)
                academic_years = cursor.fetchall()
                test_results['api_tests']['academic_years'] = {
                    'count': len(academic_years),
                    'data': [{'id': r[0], 'name': r[1]} for r in academic_years],
                    'note': 'Used SchoolID=3 as fallback'
                }
        
        # Test Classes
        with connection.cursor() as cursor:
            if school_id:
                cursor.execute("""
                    SELECT ClassID, ClassName 
                    FROM ClassMaster 
                    WHERE SchoolID = %s 
                    ORDER BY 1
                """, [school_id])
                classes = cursor.fetchall()
                test_results['api_tests']['classes'] = {
                    'count': len(classes),
                    'data': [{'id': r[0], 'name': r[1]} for r in classes]
                }
            else:
                # Test with SchoolID=3 as fallback
                cursor.execute("""
                    SELECT ClassID, ClassName 
                    FROM ClassMaster 
                    WHERE SchoolID = 3 
                    ORDER BY 1
                """)
                classes = cursor.fetchall()
                test_results['api_tests']['classes'] = {
                    'count': len(classes),
                    'data': [{'id': r[0], 'name': r[1]} for r in classes],
                    'note': 'Used SchoolID=3 as fallback'
                }
                
    except Exception as e:
        test_results['api_tests']['error'] = str(e)
    
    return JsonResponse(test_results)


@custom_login_required
def get_subjects_by_class(request):
    """Get subjects by class ID"""
    if request.method != 'GET':
        return JsonResponse({'status': 'FAILED', 'message': 'Invalid request method'})
    
    try:
        school_id = request.session.get('SchoolID')
        class_id = request.GET.get('class_id')
        exam_id = request.GET.get('exam_id')
        
        if not school_id:
            return JsonResponse({'status': 'FAILED', 'message': 'School ID not found in session'})
        
        if not class_id:
            return JsonResponse({'status': 'FAILED', 'message': 'Class ID is required'})
        
        try:
            class_id = int(class_id)
        except (ValueError, TypeError):
            return JsonResponse({'status': 'FAILED', 'message': 'Invalid Class ID format'})
        
        subjects = []
        
        with connection.cursor() as cursor:
            if exam_id:
                cursor.execute("""
                    SELECT "SubjectID", "SubjectCode", "SubjectName" 
                    FROM "SubjectMaster" 
                    WHERE "SchoolID" = %s AND "ClassId" = %s AND "IsDeleted" = FALSE
                    AND "SubjectID" NOT IN (
                        SELECT "SubjectID" FROM "ExamTimeTable" 
                        WHERE "ExamID" = %s AND "ClassID" = %s AND "SchoolID" = %s AND "IsActive" = FALSE
                    )
                    ORDER BY "SubjectName"
                """, [school_id, class_id, exam_id, class_id, school_id])
            else:
                cursor.execute("""
                    SELECT "SubjectID", "SubjectCode", "SubjectName" 
                    FROM "SubjectMaster" 
                    WHERE "SchoolID" = %s AND "ClassId" = %s AND "IsDeleted" = FALSE
                    ORDER BY "SubjectName"
                """, [school_id, class_id])
            
            columns = [col[0] for col in cursor.description]
            subjects = [dict(zip(columns, row)) for row in cursor.fetchall()]
        
        return JsonResponse({
            'status': 'SUCCESS',
            'subjects': subjects
        })
        
    except Exception as e:
        logger.error(f"Error fetching subjects by class: {str(e)}")
        return JsonResponse({'status': 'FAILED', 'message': f'Error: {str(e)}'})


@custom_login_required
def check_class_teacher_conflict(request):
    """Check if teacher is already class teacher for another class in same academic year"""
    if request.method != 'GET':
        return JsonResponse({'status': 'FAILED', 'message': 'Invalid request method'})
    
    try:
        school_id = request.session.get('SchoolID')
        if not school_id:
            return JsonResponse({'status': 'FAILED', 'message': 'School ID not found in session'})
        
        teacher_id = request.GET.get('teacher_id')
        academic_year = request.GET.get('academic_year')
        class_id = request.GET.get('class_id')
        
        if not all([teacher_id, academic_year, class_id]):
            return JsonResponse({'status': 'FAILED', 'message': 'Teacher ID, Academic Year, and Class ID are required'})
        
        try:
            teacher_id = int(teacher_id)
            class_id = int(class_id)
        except (ValueError, TypeError):
            return JsonResponse({'status': 'FAILED', 'message': 'Invalid ID format'})
        
        with connection.cursor() as cursor:
            # Check if teacher is already class teacher for another class in the same academic year
            cursor.execute("""
                SELECT COUNT(*) 
                FROM TeacherClassAssignment 
                WHERE TeacherID = %s 
                AND AcademicYear = %s 
                AND SchoolID = %s 
                AND IsClassTeacher = 1 
                AND ClassID != %s
            """, [teacher_id, academic_year, school_id, class_id])
            
            conflict_count = cursor.fetchone()[0]
            
            if conflict_count > 0:
                # Get details of the conflicting assignment
                cursor.execute("""
                    SELECT tc.ClassName, tcs.StartDate, tcs.EndDate
                    FROM TeacherClassAssignment tcs
                    JOIN ClassMaster tc ON tcs.ClassID = tc.ClassID
                    WHERE tcs.TeacherID = %s 
                    AND tcs.AcademicYear = %s 
                    AND tcs.SchoolID = %s 
                    AND tcs.IsClassTeacher = 1 
                    AND tcs.ClassID != %s
                """, [teacher_id, academic_year, school_id, class_id])
                
                conflict_details = cursor.fetchone()
                if conflict_details:
                    return JsonResponse({
                        'status': 'CONFLICT',
                        'message': f'Teacher is already assigned as class teacher for {conflict_details[0]} in academic year {academic_year}',
                        'conflict_class': conflict_details[0],
                        'start_date': conflict_details[1],
                        'end_date': conflict_details[2]
                    })
            
            return JsonResponse({
                'status': 'SUCCESS',
                'message': 'No conflict found'
            })
            
    except Exception as e:
        logger.error(f"Error checking class teacher conflict: {str(e)}")
        return JsonResponse({'status': 'FAILED', 'message': f'Error: {str(e)}'})


@custom_login_required
def check_national_id(request):
    """Check if National ID already exists for the school"""
    if request.method != 'POST':
        return JsonResponse({'status': 'FAILED', 'message': 'Invalid request method'})
    
    try:
        national_id = request.POST.get('national_id', '').strip()
        school_id = request.session.get('SchoolID')
        profile_id = request.session.get('ProfileID')
        
        # Super Admin (ProfileID=1) or if no session SchoolID, get from POST data
        if str(profile_id) == '1' or not school_id:
            form_school_id = request.POST.get('school_id', '').strip()
            if form_school_id:
                # Try decrypting first
                decrypted = decrypt_id(form_school_id)
                if decrypted:
                    school_id = int(decrypted)
                else:
                    try:
                        school_id = int(form_school_id)
                    except (ValueError, TypeError):
                        pass
        
        if not national_id:
            return JsonResponse({'status': 'FAILED', 'message': 'National ID is required'})
        
        if not school_id:
            return JsonResponse({'status': 'FAILED', 'message': 'Please select a school first'})
        
        # Check if National ID exists in EmployeeMaster
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT * FROM "Proc_Check_NationalID"(%s, %s)
            """, [national_id, school_id])
            
            result = cursor.fetchone()
            count = result[0] if result else 0
            
            if count > 0:
                return JsonResponse({
                    'status': 'SUCCESS',
                    'available': False,
                    'message': 'National ID already exists'
                })
            else:
                return JsonResponse({
                    'status': 'SUCCESS',
                    'available': True,
                    'message': 'National ID is available'
                })
                
    except Exception as e:
        logger.error(f"Error in check_national_id: {str(e)}", exc_info=True)
        return JsonResponse({'status': 'FAILED', 'message': f'Error checking National ID: {str(e)}'})


@custom_login_required
def add_employee_submit(request):
    """Handle employee form submission using stored procedure"""
    if request.method != 'POST':
        return JsonResponse({'status': 'FAILED', 'message': 'Invalid request method'})
    
    try:
        # Get form data
        school_id = request.session.get('SchoolID')
        created_by = request.session.get('UserId')
        
        # Super Admin Override: Use school_id from form if available
        profile_id_session = request.session.get('ProfileID')
        if str(profile_id_session) == '1':
            form_school_id = request.POST.get('school_id')
            if form_school_id:
                # Try decrypting first
                decrypted = decrypt_id(form_school_id)
                if decrypted:
                    school_id = int(decrypted)
                else:
                    try:
                        school_id = int(form_school_id)
                    except ValueError:
                        pass # Keep session school_id if invalid
        
        if not school_id or not created_by:
            return JsonResponse({'status': 'FAILED', 'message': 'Session expired. Please login again.'})
        
        # Required fields
        employee_name = request.POST.get('employeeName', '').strip()
        profile_id = request.POST.get('profileId')
        date_of_joining = request.POST.get('dateOfJoining')
        password = request.POST.get('password', '').strip()
        employment_type = request.POST.get('employmentType', '').strip()
        
        # Validate required fields
        if not employee_name:
            return JsonResponse({'status': 'FAILED', 'message': 'Employee name is required'})
        if not profile_id:
            return JsonResponse({'status': 'FAILED', 'message': 'Profile/Role is required'})
        if not date_of_joining:
            return JsonResponse({'status': 'FAILED', 'message': 'Date of joining is required'})
        if not password:
            return JsonResponse({'status': 'FAILED', 'message': 'Password is required'})
        if not employment_type:
            return JsonResponse({'status': 'FAILED', 'message': 'Employment Type is required'})
        
        # Convert profile_id to integer
        try:
            profile_id = int(profile_id)
        except (ValueError, TypeError):
            return JsonResponse({'status': 'FAILED', 'message': 'Invalid Profile/Role selected'})
        
        # Optional fields
        mobile_no = request.POST.get('mobileNo', '').strip() or None
        email = request.POST.get('email', '').strip() or None
        date_of_birth = request.POST.get('dateOfBirth') or None
        father_or_husband_name = request.POST.get('fatherOrHusbandName', '').strip() or None
        national_id = request.POST.get('nationalId', '').strip() or None
        gender = request.POST.get('gender') or None
        religion = request.POST.get('religion', '').strip() or None
        education = request.POST.get('education', '').strip() or None
        blood_group = request.POST.get('bloodGroup') or None
        country = request.POST.get('country_name', '').strip() or None
        state = request.POST.get('state_name', '').strip() or None
        district = request.POST.get('district_name', '').strip() or None
        pincode = request.POST.get('pincode', '').strip() or None
        home_address = request.POST.get('homeAddress', '').strip() or None
        experience = request.POST.get('experience', '').strip() or None
        core_subjects = request.POST.getlist('coreSubjects')
        
        # Process salary components
        salary_components = []
        salary_keys = [key for key in request.POST.keys() if key.startswith('salaryComponents[')]
        
        for key in salary_keys:
            if '][amount]' in key:
                # Extract component ID from the key (e.g., "salaryComponents[31][amount]" -> "31")
                component_id = key.split('[')[1].split(']')[0]
                amount = request.POST.get(key)
                
                if amount and float(amount) > 0:
                    salary_components.append({
                        'ComponentID': int(component_id),
                        'Amount': float(amount)
                    })
        
        # Process document components
        document_components = []
        
        # Process document types and files
        doc_types = request.POST.getlist('docType[]')
        doc_files = request.FILES.getlist('docFile[]')
        
        for i, doc_type in enumerate(doc_types):
            if i < len(doc_files) and doc_files[i]:
                file_obj = doc_files[i]
                file_name = file_obj.name
                file_extension = file_name.split('.')[-1].lower() if '.' in file_name else ''
                
                # Validate file (Security Check)
                is_valid, error_msg = validate_uploaded_file(file_obj)
                if not is_valid:
                    return JsonResponse({
                        'status': 'FAILED', 
                        'message': f'Document "{file_name}" validation failed: {error_msg}'
                    })
                
                # Read file content and convert to base64
                file_content = file_obj.read()
                
                document_components.append({
                    'DocumentType': doc_type,
                    'FilesName': file_name,
                    'FileExtension': file_extension,
                    'FileContent': base64.b64encode(file_content).decode('utf-8')
                })
        
        # Convert to JSON strings
        salary_json = json.dumps(salary_components) if salary_components else None
        document_json = json.dumps(document_components) if document_components else None
        
        # Hash the password using Django's make_password (same as create_user and student_admission)
        hashed_password = make_password(password)
        
        # Call PostgreSQL function
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT * FROM "Proc_Executive_set"(
                    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s
                )
            """, [
                school_id, employee_name, mobile_no, email, hashed_password, date_of_birth,
                profile_id, date_of_joining, father_or_husband_name, national_id,
                gender, religion, education, blood_group, country, state, district,
                pincode, home_address, experience, employment_type, created_by,
                salary_json, document_json
            ])
            
            result = cursor.fetchone()
            if result:
                # Result is (ResultJson, UserCode)
                result_json_str = result[0]
                user_code = result[1]
                
                # Parse JSON result
                result_json = json.loads(result_json_str) if result_json_str else {}
                
                # Save Core Subjects mapping if provided
                if result_json.get('Status') == 'SUCCESS' and core_subjects:
                    try:
                        employee_id = result_json.get('EmployeeID')
                        if employee_id:
                            with connection.cursor() as sub_cursor:
                                for subject_id in core_subjects:
                                    sub_cursor.execute("""
                                        INSERT INTO "EmployeeSpecialization" ("EmployeeID", "SpecializationID", "SchoolID")
                                        VALUES (%s, %s, %s)
                                    """, [employee_id, subject_id, school_id])
                    except Exception as sub_e:
                        logger.error(f"Error saving employee specializations: {str(sub_e)}")
                
                # Send email if employee registration was successful and email is provided
                if result_json.get('Status') == 'SUCCESS' and email:
                    try:
                        # Get employee details for email
                        employee_id = result_json.get('EmployeeID')
                        employee_code = user_code  # Use @UserCode from procedure output instead of JSON
                        
                        # Fallback to generated code if @UserCode is not provided
                        if employee_id and not employee_code:
                            employee_code = f"EMP{employee_id:06d}"
                        
                        if employee_id and employee_code:
                            # Store employee data in session for acknowledgment
                            request.session['employee_ack_data'] = safe_json_obj({
                                'employee_id': employee_id,
                                'employee_code': employee_code,
                                'employee_name': employee_name,
                                'email': email,
                                'mobile_no': mobile_no,
                                'position': result_json.get('Position', 'Staff'),
                                'date_of_joining': date_of_joining,
                                'school_id': school_id,
                                'profile_id': profile_id
                            })
                            
                            # Send employee registration email asynchronously
                            send_employee_registration_email_async({
                                'employee_id': employee_id,
                                'employee_code': employee_code,
                                'employee_name': employee_name,
                                'email': email,
                                'mobile_no': mobile_no,
                                'position': result_json.get('Position', 'Staff'),
                                'date_of_joining': date_of_joining,
                                'school_id': school_id,
                                'profile_id': profile_id,  # Add profile_id for dynamic rules
                                'username': result_json.get('Username', email),  # Use email as username if not provided
                                'password': password,  # Use actual password from form submission
                                'salary_components': salary_components,
                                'document_components': document_components
                            })
                            logger.info(f"Employee registration email queued for {email} (Employee ID: {employee_id}, Code: {employee_code})")
                        else:
                            logger.warning(f"Email not sent: EmployeeID={employee_id}, EmployeeCode={employee_code}")
                    except Exception as email_error:
                        logger.error(f"Failed to queue employee registration email: {str(email_error)}")
                        # Don't fail the main operation if email fails
                
                return JsonResponse(result_json)
            else:
                return JsonResponse({'status': 'FAILED', 'message': 'No result from stored procedure'})
                
    except Exception as e:
        logger.error(f"Error in add_employee_submit: {str(e)}", exc_info=True)
        return JsonResponse({'status': 'FAILED', 'message': f'Error adding employee: {str(e)}'})


@custom_login_required
def view_assign_class(request):
    """
    View Assign Class page - Display teacher class assignments in table format with filtering and pagination
    Uses the Proc_TeacherClassAssignment_Report stored procedure
    """
    # Get user context for header
    context = get_context(request)
    
    # Also get session info for user object (needed for header template)
    sess = _get_custom_session_info(request)
    if sess:
        context['user'] = sess
    
    # Get user information
    user_id = request.session.get('UserId')
    school_id = request.session.get('SchoolID')
    profile_id = request.session.get('ProfileID')
    
    if not user_id:
        messages.error(request, "Please login to access assignment data")
        return redirect('login')
    
    if not school_id:
        messages.error(request, "School ID is required to access assignment data")
        return redirect('login')
    
    if not profile_id:
        messages.error(request, "Profile ID is required to access assignment data")
        return redirect('login')
    
    # Get query parameters
    page = safe_int(request.GET.get('page', 1))
    per_page = safe_int(request.GET.get('per_page', 25))
    
    # Optimize per_page for mobile devices
    is_mobile = _is_mobile_request(request)
    if is_mobile and per_page > 15:
        per_page = 15  # Reduce records per page on mobile for better performance
    
    # Filter parameters
    academic_year = request.GET.get('academic_year', '').strip()
    teacher_name = request.GET.get('teacher_name', '').strip()
    class_name = request.GET.get('class_name', '').strip()
    subject_name = request.GET.get('subject_name', '').strip()
    search = request.GET.get('search', '').strip()
    order_by = request.GET.get('order_by', 'TeacherName')
    order_direction = request.GET.get('order_direction', 'ASC')
    
    # Initialize variables
    assignments = []
    total_count = 0
    active_count = 0
    start_index = 0
    end_index = 0
    has_next = False
    has_prev = False
    
    try:
        # Add performance logging
        import time
        start_time = time.time()
        
        with connection.cursor() as cursor:
            # Use PostgreSQL function
            cursor.execute("""
                SELECT * FROM "Proc_TeacherClassAssignment_Report"(
                    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s
                )
            """, [
                school_id,
                academic_year if academic_year else None,
                teacher_name if teacher_name else None,
                class_name if class_name else None,
                subject_name if subject_name else None,
                search if search else None,
                page,
                per_page,
                order_by,
                order_direction
            ])
            
            # Get result set (assignments + counts)
            columns = [col[0] for col in cursor.description]
            raw_rows = cursor.fetchall()
            
            # Log database query performance
            db_time = time.time() - start_time
            print(f"Database query time: {db_time:.3f}s - Mobile: {is_mobile} - Search: {search}")
            
            assignments = []
            if raw_rows:
                # Get counts from the first row (ActiveCount is last, TotalCount is second to last)
                # Columns: AssignmentID, TeacherName, ClassName, SectionName, SubjectName, AcademicYear, StartDate, EndDate, IsClassTeacher, TotalCount, ActiveCount
                
                total_count = raw_rows[0][-2]
                active_count = raw_rows[0][-1]
                
                for row in raw_rows:
                    # Exclude the last two count columns from the assignment dict
                    assignment = dict(zip(columns[:-2], row[:-2]))
                    
                    # Format dates
                    if assignment.get('StartDate'):
                        assignment['StartDateFormatted'] = assignment['StartDate'].strftime('%Y-%m-%d')
                    else:
                        assignment['StartDateFormatted'] = 'N/A'
                        
                    if assignment.get('EndDate'):
                        assignment['EndDateFormatted'] = assignment['EndDate'].strftime('%Y-%m-%d')
                    else:
                        assignment['EndDateFormatted'] = 'N/A'
                    
                    # Format boolean fields
                    assignment['IsClassTeacherText'] = 'Yes' if assignment.get('IsClassTeacher') else 'No'
                    
                    assignments.append(assignment)
            
    except Exception as e:
        logger.error(f"Error fetching teacher class assignments: {str(e)}", exc_info=True)
        messages.error(request, "Error loading assignment data. Please try again.")
        assignments = []
        total_count = 0
        active_count = 0
    
    # Calculate pagination info
    start_index = (page - 1) * per_page + 1 if assignments else 0
    end_index = min(start_index + len(assignments) - 1, total_count) if assignments else 0
    has_next = end_index < total_count
    has_prev = page > 1
    
    # Add pagination and filter context
    context.update({
        'assignments': assignments,
        'total_count': total_count,
        'active_count': active_count,
        'page': page,
        'per_page': per_page,
        'start_index': start_index,
        'end_index': end_index,
        'has_next': has_next,
        'has_prev': has_prev,
        'search': search,
        'academic_year': academic_year,
        'teacher_name': teacher_name,
        'class_name': class_name,
        'subject_name': subject_name,
        'order_by': order_by,
        'order_direction': order_direction,
        'dark_mode': request.session.get('dark_mode', False)
    })
    
    # Log total processing time
    total_time = time.time() - start_time
    print(f"Total view processing time: {total_time:.3f}s - Mobile: {is_mobile} - Assignments: {len(assignments)}")
    
    return render(request, 'view_assign_class.html', context)


# =============================================
# Fee Collection Management Views
# =============================================

# =============================================
# New Fee Collection Management Views
# =============================================

@custom_login_required
def student_attendance(request):
    """
    Student Attendance page - Load classes and sections like assign_classes_view
    """
    # Get user context for header
    context = get_context(request)
    
    # Get user information
    user_id = request.session.get('UserId')
    school_id = request.session.get('SchoolID')
    profile_id = request.session.get('ProfileID')
    
    # Super Admin School Support
    is_super_admin = profile_id == 1
    if is_super_admin:
        school_id_param = request.GET.get('school_id')
        if school_id_param:
            try:
                school_id = decrypt_id(school_id_param)
            except:
                school_id = school_id_param # Fallback for raw ID if any
    
    if not user_id:
        messages.error(request, "Please login to access attendance data")
        return redirect('login')
    
    if not school_id and not is_super_admin:
        messages.error(request, "School ID is required to access attendance data")
        return redirect('login')
    
    # Get query parameters
    selected_date = request.GET.get('date', '')
    class_id = request.GET.get('class_id', '')
    section_id = request.GET.get('section_id', '')
    
    # Default to today's date if not provided
    if not selected_date:
        selected_date = timezone.now().strftime('%Y-%m-%d')
    
    # Initialize data structures
    classes = []
    all_sections = {}  # Dictionary to store sections by class ID
    students = []
    attendance_data = {}
    
    try:
        with connection.cursor() as cursor:
            # Get classes using the same simple approach as assign_classes_view
            cursor.execute("""
                SELECT "ClassID", "ClassName" 
                FROM "ClassMaster" 
                WHERE "SchoolID" = %s 
                ORDER BY 1
            """, [school_id])
            
            classes = [{"id": r[0], "name": r[1]} for r in cursor.fetchall()]
            
            # Get all sections for all classes using the same approach as assign_classes_view
            cursor.execute("""
                SELECT "SectionID", "SectionName", "ClassID" 
                FROM "SectionMaster" 
                WHERE "ClassID" IN (SELECT "ClassID" FROM "ClassMaster" WHERE "SchoolID" = %s)
            """, [school_id])
            
            # Group sections by class ID (same logic as assign_classes_view)
            for row in cursor.fetchall():
                class_id_str = str(row[2])  # Convert to string for JavaScript compatibility
                if class_id_str not in all_sections:
                    all_sections[class_id_str] = []
                all_sections[class_id_str].append({
                    "id": row[0],
                    "name": row[1]
                })
    
    except Exception as e:
        logger.error(f"Error loading attendance data: {str(e)}", exc_info=True)
        messages.error(request, "Error loading attendance data. Please try again.")
            
    # Get sections for selected class if any
    sections = all_sections.get(class_id, []) if class_id else []

    # Get names for selected class and section for the summary
    selected_class_name = next((c['name'] for c in classes if str(c['id']) == class_id), "")
    selected_section_name = ""
    if class_id and section_id:
        class_secs = all_sections.get(class_id, [])
        selected_section_name = next((s['name'] for s in class_secs if str(s['id']) == section_id), "")
    
    # Get school name for super admin
    selected_school_name = ""
    if is_super_admin and school_id:
        try:
            with connection.cursor() as cursor:
                cursor.execute('SELECT "SchoolName" FROM "School" WHERE "SchoolID" = %s', [school_id])
                row = cursor.fetchone()
                if row:
                    selected_school_name = row[0]
        except:
            pass

    context.update({
        'classes': classes,
        'sections': sections,
        'all_sections': json.dumps(all_sections, cls=DjangoJSONEncoder),
        'students': students,
        'attendance_data': attendance_data,
        'selected_date': selected_date,
        'selected_class_id': class_id,
        'selected_section_id': section_id,
        'selected_class_name': selected_class_name,
        'selected_section_name': selected_section_name,
        'selected_school_id': encrypt_id(school_id) if is_super_admin and school_id else "",
        'selected_school_name': selected_school_name,
    })
    
    return render(request, 'core/student_attendance.html', context)



@custom_login_required
@csrf_exempt
def load_students_ajax(request):
    """
    Load students for attendance using Proc_StudentList_Get
    """
    try:
        # Get parameters
        class_id = request.POST.get('class_id')
        section_id = request.POST.get('section_id')
        date = request.POST.get('date')
        
        if not class_id:
            return JsonResponse({'status': 'ERROR', 'message': 'Class ID is required'})
        
        if not date:
            return JsonResponse({'status': 'ERROR', 'message': 'Date is required'})
        
        # Get user context
        user_id = request.session.get('UserId')
        school_id = request.session.get('SchoolID')
        profile_id = request.session.get('ProfileID')

        # Super Admin School Support
        if profile_id == 1:
            school_id_param = request.POST.get('school_id') or request.GET.get('school_id')
            if school_id_param:
                try:
                    school_id = decrypt_id(school_id_param)
                except:
                    school_id = school_id_param

        if not school_id:
            return JsonResponse({'status': 'ERROR', 'message': 'School ID is required'})
        
        # Prepare parameters for the PostgreSQL procedure
        section_id_param = int(section_id) if section_id and section_id != 'None' and section_id != '' else None
        
        try:
            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT * FROM "Proc_StudentList_Get"(%s, %s, %s, %s::DATE)
                """, [school_id, int(class_id), section_id_param, date])
                
                columns = [col[0] for col in cursor.description]
                students = []
                
                for row in cursor.fetchall():
                    student_dict = dict(zip(columns, row))
                    students.append(student_dict)
            
            return JsonResponse({
                'status': 'SUCCESS',
                'students': students,
                'count': len(students)
            })
        except Exception as db_err:
            logger.error(f"Database error loading students: {str(db_err)}")
            return JsonResponse({
                'status': 'ERROR',
                'message': f'Database error: {str(db_err)}'
            })
                    
    except Exception as e:
        return JsonResponse({
            'status': 'ERROR',
            'message': f'Error loading students: {str(e)}'
        })


@custom_login_required
@csrf_exempt
@require_POST
def submit_attendance_ajax(request):
    """
    Submit attendance data using Proc_Student_Attendance_set
    """
    try:
        # Get parameters
        class_id = request.POST.get('class_id')
        section_id = request.POST.get('section_id')
        date = request.POST.get('date')
        attendance_data = request.POST.get('attendance_data')
        
        if not all([class_id, date, attendance_data]):
            return JsonResponse({'status': 'ERROR', 'message': 'Missing required parameters'})
        
        # Parse attendance data
        attendance_dict = json.loads(attendance_data)
        
        # Get user context
        user_id = request.session.get('UserId')
        school_id = request.session.get('SchoolID')
        profile_id = request.session.get('ProfileID')
        
        # Super Admin School Support
        if profile_id == 1:
            school_id_param = request.POST.get('school_id')
            if school_id_param:
                try:
                    school_id = decrypt_id(school_id_param)
                except:
                    school_id = school_id_param

        if not school_id:
            return JsonResponse({'status': 'ERROR', 'message': 'School ID is required'})
        
        # Prepare attendance data for the stored procedure
        attendance_records = []
        for student_id, status in attendance_dict.items():
            attendance_records.append({
                'SchoolID': school_id,
                'StudentID': int(student_id),
                'ClassID': int(class_id),
                'SectionID': int(section_id) if section_id and section_id != '' else 0,
                'AttendanceDate': date,
                'Status': status,
                'Remarks': '',
                'CreatedBy': user_id
            })
        
        attendance_json = json.dumps(attendance_records)
        
        try:
            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT "Status", "Message" FROM "Proc_Student_Attendance_set"(%s, %s, %s, %s, %s, %s)
                """, [
                    school_id, 
                    int(class_id), 
                    int(section_id) if section_id and section_id != '' else 0,
                    date,
                    attendance_json,
                    user_id
                ])
                
                result = cursor.fetchone()
                
                if result and result[0] == 'SUCCESS':
                    return JsonResponse({'status': 'SUCCESS', 'message': result[1]})
                else:
                    return JsonResponse({'status': 'ERROR', 'message': result[1] or 'Error saving attendance'})
        except Exception as db_err:
            logger.error(f"Database error saving attendance: {str(db_err)}")
            return JsonResponse({'status': 'ERROR', 'message': f'Database error: {str(db_err)}'})
            
    except Exception as e:
        return JsonResponse({'status': 'ERROR', 'message': f'Error saving attendance: {str(e)}'})


@custom_login_required
def view_attendance(request):
    """
    View Attendance page with graphs and analytics
    """
    context = get_context(request)
    
    user_id = request.session.get('UserId')
    school_id = request.session.get('SchoolID')
    profile_id = request.session.get('ProfileID')
    
    # Super Admin School Support
    is_super_admin = profile_id == 1
    if is_super_admin:
        school_id_param = request.GET.get('school_id')
        if school_id_param:
            try:
                school_id = decrypt_id(school_id_param)
            except:
                school_id = school_id_param

    if not user_id:
        return redirect('login')
        
    if not school_id and not is_super_admin:
        messages.error(request, "School ID is required")
        return redirect('login')
        
    selected_date = request.GET.get('date', timezone.now().strftime('%Y-%m-%d'))
    class_id = request.GET.get('class_id', '')
    section_id = request.GET.get('section_id', '')
    
    classes = []
    all_sections = {}
    attendance_data = {}
    attendance_stats = {}
    
    try:
        with connection.cursor() as cursor:
            # Get classes
            cursor.execute("""
                SELECT "ClassID", "ClassName" FROM "ClassMaster" WHERE "SchoolID" = %s ORDER BY 1
            """, [school_id])
            classes = [{"id": r[0], "name": r[1]} for r in cursor.fetchall()]
            
            # Get sections
            cursor.execute("""
                SELECT "SectionID", "SectionName", "ClassID" FROM "SectionMaster" 
                WHERE "ClassID" IN (SELECT "ClassID" FROM "ClassMaster" WHERE "SchoolID" = %s)
            """, [school_id])
            for r in cursor.fetchall():
                cid = str(r[2])
                if cid not in all_sections: all_sections[cid] = []
                all_sections[cid].append({"id": r[0], "name": r[1]})
                
            # Get data if filtered
            if class_id:
                cursor.execute("""
                    SELECT * FROM "Proc_StudentAttendance_Get"(%s, %s, %s::DATE)
                """, [school_id, int(class_id), selected_date])
                
                records = cursor.fetchall()
                for r in records:
                    attendance_data[r[0]] = {
                        'name': r[1], 'code': r[2], 'status': r[3],
                        'date': r[4], 'class': r[5], 'section': r[6] or 'N/A'
                    }
                
                total = len(attendance_data)
                present = sum(1 for d in attendance_data.values() if d['status'] == 'present')
                absent = sum(1 for d in attendance_data.values() if d['status'] == 'absent')
                late = sum(1 for d in attendance_data.values() if d['status'] == 'late')
                holiday = sum(1 for d in attendance_data.values() if d['status'] == 'holiday')
                
                attendance_stats = {
                    'total': total,
                    'present': present,
                    'absent': absent,
                    'late': late,
                    'holiday': holiday,
                    'present_percentage': round((present / total * 100) if total > 0 else 0, 1),
                    'absent_percentage': round((absent / total * 100) if total > 0 else 0, 1),
                    'late_percentage': round((late / total * 100) if total > 0 else 0, 1),
                    'holiday_percentage': round((holiday / total * 100) if total > 0 else 0, 1)
                }
    except Exception as e:
        logger.error(f"Error in view_attendance: {e}", exc_info=True)
        messages.error(request, "Error loading attendance report")

    sections = all_sections.get(class_id, []) if class_id else []
    
    context.update({
        'classes': classes,
        'sections': sections,
        'all_sections': json.dumps(all_sections),
        'attendance_data': attendance_data,
        'attendance_stats': attendance_stats,
        'selected_date': selected_date,
        'selected_class_id': class_id,
        'selected_section_id': section_id,
        'selected_class_name': next((c['name'] for c in classes if str(c['id']) == class_id), None),
        'selected_section_name': next((s['name'] for s in sections if str(s['id']) == section_id), None),
    })
    
    return render(request, 'core/view_attendance.html', context)


@custom_login_required
def submit_attendance_ajax(request):
    """
    Submit attendance data using Proc_Student_Attendance_set stored procedure
    """
    if request.method != 'POST':
        return JsonResponse({'status': 'ERROR', 'message': 'Invalid request method'})
    
    try:
        # Get parameters
        class_id = request.POST.get('class_id')
        section_id = request.POST.get('section_id')
        date = request.POST.get('date')
        attendance_data = request.POST.get('attendance_data')
        
        if not all([class_id, date, attendance_data]):
            return JsonResponse({'status': 'ERROR', 'message': 'Missing required parameters'})
        
        # Parse attendance data
        import json
        attendance_dict = json.loads(attendance_data)
        
        # Get user context
        user_id = request.session.get('UserId')
        school_id = request.session.get('SchoolID')
        profile_id = request.session.get('ProfileID')
        
        # Super Admin School Support
        if profile_id == 1:
            school_id_param = request.POST.get('school_id')
            if school_id_param:
                try:
                    school_id = decrypt_id(school_id_param)
                except:
                    school_id = school_id_param

        if not user_id:
            return JsonResponse({'status': 'ERROR', 'message': 'User ID is required'})
        
        if not school_id:
            return JsonResponse({'status': 'ERROR', 'message': 'School ID is required'})
        
        # Permission check removed - using same simple approach as assign_classes_view
        
        # Get student data to extract section IDs
        student_sections = {}
        if not section_id:  # Only if no specific section is selected
            try:
                with connection.cursor() as cursor:
                    cursor.execute("""
                        SELECT "StudentID", 
                        CAST(NULLIF(REGEXP_REPLACE("Section"::TEXT, '[^0-9]', '', 'g'), '') AS INTEGER) AS "SectionID"
                        FROM "Student"
                        WHERE "SchoolID" = %s 
                        AND CAST(NULLIF(REGEXP_REPLACE("AdmissionClass"::TEXT, '[^0-9]', '', 'g'), '') AS INTEGER) = %s 
                        AND "IsDeleted" = FALSE
                    """, [school_id, int(class_id)])
                    
                    for row in cursor.fetchall():
                        student_sections[row[0]] = row[1]
            except Exception as e:
                logger.error(f"Error fetching student sections for bulk attendance: {str(e)}")

        # Prepare attendance data for the stored procedure
        attendance_records = []
        for student_id, status in attendance_dict.items():
            # Use the section_id from the form if provided, otherwise use the student's actual section
            actual_section_id = int(section_id) if section_id else student_sections.get(int(student_id))
            
            
            attendance_records.append({
                'SchoolID': school_id,
                'StudentID': int(student_id),
                'ClassID': int(class_id),
                'SectionID': actual_section_id,
                'AttendanceDate': date,
                'Status': status,
                'Remarks': '',
                'UserID': user_id
            })
        
        # Use the newly created PostgreSQL procedure for bulk attendance save
        with connection.cursor() as cursor:
            try:
                # Convert attendance_records to JSON for the procedure
                attendance_json = json.dumps([
                    {
                        'StudentID': r['StudentID'],
                        'Status': r['Status'],
                        'Remarks': r['Remarks'],
                        'SectionID': r['SectionID']
                    } for r in attendance_records
                ])

                # Call the PostgreSQL bulk marking function with explicit type casts
                cursor.execute("""
                    SELECT "Result", "Message" FROM "Proc_StudentAttendance_Mark_Bulk"(%s, %s, %s, %s::DATE, %s::JSONB, %s)
                """, [
                    school_id, 
                    int(class_id), 
                    int(section_id) if section_id else 0, # Default Section 0 if none
                    date,
                    attendance_json,
                    user_id
                ])
                
                result = cursor.fetchone()
                
                if result and result[0] == 'SUCCESS':
                    return JsonResponse({
                        'status': 'SUCCESS',
                        'message': result[1]
                    })
                else:
                    return JsonResponse({
                        'status': 'ERROR',
                        'message': result[1] if result else 'Unknown error occurred in stored procedure'
                    })
                
            except Exception as db_err:
                logger.error(f"Database error saving attendance: {str(db_err)}")
                return JsonResponse({
                    'status': 'ERROR',
                    'message': f'Database error: {str(db_err)}'
                })
        
    except json.JSONDecodeError as e:
        return JsonResponse({
            'status': 'ERROR', 
            'message': 'Invalid attendance data format'
        })
    except Exception as e:
        return JsonResponse({
            'status': 'ERROR', 
            'message': f'Error saving attendance: {str(e)}'
        })


@custom_login_required
def view_attendance(request):
    """
    View Attendance page with graphs and analytics
    """
    # Get user context for header
    context = get_context(request)
    
    # Get user information
    user_id = request.session.get('UserId')
    school_id = request.session.get('SchoolID')
    profile_id = request.session.get('ProfileID')
    
    # Super Admin School Support
    is_super_admin = profile_id == 1
    if is_super_admin:
        school_id_param = request.GET.get('school_id')
        if school_id_param:
            try:
                school_id = decrypt_id(school_id_param)
            except:
                school_id = school_id_param

    if not user_id:
        messages.error(request, "Please login to access attendance data")
        return redirect('login')
        
    if not school_id and not is_super_admin:
        messages.error(request, "School ID is required to access attendance data")
        return redirect('login')
        
    # Get query parameters
    selected_date = request.GET.get('date', '')
    class_id = request.GET.get('class_id', '')
    section_id = request.GET.get('section_id', '')
    
    # Sorting and Pagination parameters
    sort_by = request.GET.get('sort_by', 'name')
    sort_order = request.GET.get('sort_order', 'ASC')
    try:
        page = int(request.GET.get('page', 1))
    except (ValueError, TypeError):
        page = 1
    page_size = 20
    
    # Default to today's date if not provided
    if not selected_date:
        selected_date = timezone.now().strftime('%Y-%m-%d')
    
    # Initialize data structures
    classes = []
    all_sections = {}
    students_list = []  # Use a list for sorted data
    attendance_stats = {}
    pagination_data = {}
    
    try:
        with connection.cursor() as cursor:
            # Get classes using simple approach
            cursor.execute("""
                SELECT "ClassID", "ClassName" 
                FROM "ClassMaster" 
                WHERE "SchoolID" = %s 
                ORDER BY 1
            """, [school_id])
            
            classes = [{"id": r[0], "name": r[1]} for r in cursor.fetchall()]
            
            # Get all sections for all classes
            cursor.execute("""
                SELECT "SectionID", "SectionName", "ClassID" 
                FROM "SectionMaster" 
                WHERE "ClassID" IN (SELECT "ClassID" FROM "ClassMaster" WHERE "SchoolID" = %s)
            """, [school_id])
            
            sections_rows = cursor.fetchall()
            
            # Group sections by class ID
            for row in sections_rows:
                class_id_str = str(row[2])
                if class_id_str not in all_sections:
                    all_sections[class_id_str] = []
                all_sections[class_id_str].append({"id": row[0], "name": row[1]})
    
    except Exception as e:
        logger.error(f"Error loading metadata: {str(e)}")
        messages.error(request, "Error loading filter options.")
    
    # Get sections for selected class if any
    sections = all_sections.get(class_id, []) if class_id else []
    
    # Load attendance data if class is selected
    if class_id:
        try:
            with connection.cursor() as cursor:
                # Call the enhanced procedure with explicit type casting
                p_section_id = int(section_id) if section_id else None
                # school_id might be encrypted string or integer, cast it explicitly
                # p_section_id might be None, cast it to integer
                cursor.execute("""
                    SELECT Proc_Student_AttendanceReport_Get(%s::INTEGER, %s::INTEGER, %s::INTEGER, %s::DATE, %s::VARCHAR, %s::VARCHAR, %s::INTEGER, %s::INTEGER)
                """, [school_id, int(class_id), p_section_id, selected_date, sort_by, sort_order, page, page_size])
                
                result_json = cursor.fetchone()[0]
                
                if result_json:
                    # Explicitly parse if it is returned as a string
                    if isinstance(result_json, str):
                        try:
                            result_json = json.loads(result_json)
                        except Exception as parse_error:
                            logger.error(f"Error parsing JSON result from procedure: {parse_error}")
                            result_json = None
                            
                    if result_json and isinstance(result_json, dict):
                        attendance_stats = result_json.get('stats', {})
                        students_list = result_json.get('students', [])
                        pagination_data = result_json.get('pagination', {})
                    else:
                        logger.error(f"Unexpected result type from procedure: {type(result_json)}")

        except Exception as e:
            logger.error(f"Error calling Proc_Student_AttendanceReport_Get: {str(e)}")
            messages.error(request, "Error loading attendance report.")
    
    context.update({
        'classes': classes,
        'sections': sections,
        'all_sections': json.dumps(all_sections),
        'students_list': students_list,
        'attendance_stats': attendance_stats,
        'pagination': pagination_data,
        'start_index': (page - 1) * page_size,
        'selected_date': selected_date,
        'selected_class_id': class_id,
        'selected_section_id': section_id,
        'sort_by': sort_by,
        'sort_order': sort_order,
        'selected_class_name': next((c['name'] for c in classes if str(c['id']) == class_id), None),
        'selected_section_name': next((s['name'] for s in sections if str(s['id']) == section_id), None),
    })
    
    return render(request, 'core/view_attendance.html', context)


# =============================================
# Fee Report Management Views
# =============================================



# Profile Menu Mapping Views moved to core/menus_views.py













# Exam Management Views
@custom_login_required
@custom_login_required
def test_documents_view(request):
    """Test view to check document data in database"""
    from django.http import JsonResponse
    
    # If it's not an AJAX request, render the HTML template
    if request.headers.get('X-Requested-With') != 'XMLHttpRequest' and not request.path.endswith('.json'):
        context = get_context(request)
        return render(request, 'debug_documents.html', context)
    
    try:
        with connection.cursor() as cursor:
            # Check total documents
            cursor.execute("""
                SELECT COUNT(*) as total_docs,
                       COUNT(CASE WHEN DocumentData IS NOT NULL THEN 1 END) as docs_with_data,
                       COUNT(CASE WHEN DocumentData IS NULL THEN 1 END) as docs_without_data
                FROM StudentDocuments 
                WHERE ISNULL(IsDeleted, 0) = 0
            """)
            result = cursor.fetchone()
            
            # Get sample documents
            cursor.execute("""
                SELECT TOP 10 
                    sd.DocumentID, sd.StudentID, s.StudentCode, sd.DocumentType, sd.DocumentName, 
                    CASE WHEN sd.DocumentData IS NULL THEN 'NULL' ELSE 'HAS_DATA' END as DataStatus,
                    LEN(sd.DocumentData) as DataLength,
                    sd.UploadDate
                FROM StudentDocuments sd
                LEFT JOIN Student s ON sd.StudentID = s.StudentID
                WHERE ISNULL(sd.IsDeleted, 0) = 0
                ORDER BY sd.UploadDate DESC
            """)
            
            sample_docs = []
            for row in cursor.fetchall():
                sample_docs.append({
                    'DocumentID': row[0],
                    'StudentID': row[1], 
                    'StudentCode': row[2],
                    'DocumentType': row[3],
                    'DocumentName': row[4],
                    'DataStatus': row[5],
                    'DataLength': row[6],
                    'UploadDate': str(row[7]) if row[7] else None
                })
            
            # Test stored procedure with a sample student
            cursor.execute("""
                SELECT TOP 1 StudentCode, StudentID, SchoolID
                FROM Student 
                WHERE ISNULL(IsDeleted, 0) = 0
                ORDER BY CreatedAt DESC
            """)
            student_result = cursor.fetchone()
            
            procedure_result = None
            if student_result:
                student_code, student_id, school_id = student_result
                
                # Test the stored procedure
                cursor.execute("EXEC Proc_application_details_get @StudentCode = %s, @SchoolID = %s", 
                             [student_code, school_id])
                
                # Skip first two result sets
                cursor.fetchall()  # Application details
                if cursor.nextset():
                    cursor.fetchall()  # Fee structure
                    
                # Get documents result set
                if cursor.nextset():
                    documents = cursor.fetchall()
                    procedure_result = {
                        'student_code': student_code,
                        'student_id': student_id,
                        'school_id': school_id,
                        'documents_count': len(documents),
                        'documents': [{'type': doc[1], 'name': doc[2], 'has_data': bool(doc[3])} for doc in documents]
                    }
            
            return JsonResponse({
                'total_documents': result[0],
                'documents_with_data': result[1], 
                'documents_without_data': result[2],
                'sample_documents': sample_docs,
                'procedure_test': procedure_result
            })
            
    except Exception as e:
        return JsonResponse({'error': str(e)})

def test_send_admission_email(request, student_code):
    """Test endpoint to send admission emails for existing student"""
    try:
        school_id = request.session.get('SchoolID')
        
        # Get student email
        with connection.cursor() as cursor:
            cursor.execute("SELECT Email FROM Student WHERE StudentCode = %s", [student_code])
            row = cursor.fetchone()
            if not row or not row[0]:
                return HttpResponse(f"Student {student_code} not found or has no email", status=404)
            email = row[0]
        
        # Queue emails
        email_data = {
            'email': email,
            'student_code': student_code,
            'school_id': school_id,
            'payment_receipt': {'student_code': student_code}
        }
        
        send_admission_emails_async_database(safe_json_obj(email_data))
        
        return HttpResponse(f"✅ Emails queued for {student_code} ({email}). Check email queue status.")
    except Exception as e:
        logger.error(f"Test email failed: {e}")
        return HttpResponse(f"❌ Error: {str(e)}", status=500)




@custom_login_required
def template_settings(request):
    school_id = request.session.get('SchoolID')
    context = get_context(request)
    
    if request.method == 'POST':
        ack_template = request.POST.get('acknowledgment_template')
        receipt_template = request.POST.get('receipt_template')
        card_template = request.POST.get('student_card_template')
        user_id = request.session.get('UserId')
        
        try:
            with connection.cursor() as cursor:
                # Update or insert AdmissionAcknowledgment
                cursor.execute("""
                    UPDATE "TemplateSettings" SET "TemplateFile" = %s, "ModifiedBy" = %s, "ModifiedAt" = CURRENT_TIMESTAMP
                    WHERE "SchoolID" = %s AND "TemplateType" = 'AdmissionAcknowledgment'
                """, [ack_template, user_id, school_id])
                
                if cursor.rowcount == 0:
                    cursor.execute("""
                        INSERT INTO "TemplateSettings" ("SchoolID", "TemplateType", "TemplateFile", "IsActive", "CreatedBy", "CreatedAt", "IsDeleted")
                        VALUES (%s, 'AdmissionAcknowledgment', %s, TRUE, %s, CURRENT_TIMESTAMP, FALSE)
                    """, [school_id, ack_template, user_id])
                
                # Update or insert PaymentReceipt
                cursor.execute("""
                    UPDATE "TemplateSettings" SET "TemplateFile" = %s, "ModifiedBy" = %s, "ModifiedAt" = CURRENT_TIMESTAMP
                    WHERE "SchoolID" = %s AND "TemplateType" = 'PaymentReceipt'
                """, [receipt_template, user_id, school_id])
                
                if cursor.rowcount == 0:
                    cursor.execute("""
                        INSERT INTO "TemplateSettings" ("SchoolID", "TemplateType", "TemplateFile", "IsActive", "CreatedBy", "CreatedAt", "IsDeleted")
                        VALUES (%s, 'PaymentReceipt', %s, TRUE, %s, CURRENT_TIMESTAMP, FALSE)
                    """, [school_id, receipt_template, user_id])
                
                # Update or insert StudentCard
                cursor.execute("""
                    UPDATE "TemplateSettings" SET "TemplateFile" = %s, "ModifiedBy" = %s, "ModifiedAt" = CURRENT_TIMESTAMP
                    WHERE "SchoolID" = %s AND "TemplateType" = 'StudentCard'
                """, [card_template, user_id, school_id])
                
                if cursor.rowcount == 0:
                    cursor.execute("""
                        INSERT INTO "TemplateSettings" ("SchoolID", "TemplateType", "TemplateFile", "IsActive", "CreatedBy", "CreatedAt", "IsDeleted")
                        VALUES (%s, 'StudentCard', %s, TRUE, %s, CURRENT_TIMESTAMP, FALSE)
                    """, [school_id, card_template, user_id])
                
            messages.success(request, 'Template settings saved successfully!')
            return redirect('template_settings')
        except Exception as e:
            logger.error(f"Error saving template settings: {e}")
            messages.error(request, 'Error saving settings')
    
    # Load current settings
    current_settings = {}
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT "TemplateType", "TemplateFile" 
                FROM "TemplateSettings" 
                WHERE "SchoolID" = %s AND "IsActive" = TRUE AND "IsDeleted" = FALSE
            """, [school_id])
            for row in cursor.fetchall():
                current_settings[row[0]] = row[1]
    except Exception as e:
        logger.error(f"Error loading template settings: {e}")
    
    context.update({
        'ack_template': current_settings.get('AdmissionAcknowledgment', 'admission_acknowledgment.html'),
        'receipt_template': current_settings.get('PaymentReceipt', 'payment_success.html'),
        'card_template': current_settings.get('StudentCard', 'student_card_template1.html')
    })
    
    return render(request, 'template_settings.html', context)

# Import fee receipt view


@custom_login_required
def print_employee_acknowledgment(request):
    """Print employee acknowledgment/job letter"""
    employee_code = request.GET.get('employee_code')
    if not employee_code:
        messages.error(request, "Employee code is required")
        return redirect('view_teachers')
    
    # Redirect to the more robust print_job_letter view in template_views
    # Redirect to the more robust print_job_letter view in template_views
    return redirect(reverse('print_job_letter') + f'?employee_code={employee_code}')


@custom_login_required
def serve_user_photo(request, user_id):
    """Serve user photo as binary file for better caching/performance"""
    try:
        from django.db import connection
        from django.http import HttpResponse
        
        with connection.cursor() as cursor:
            cursor.execute("SELECT \"UserPhoto\" FROM \"UserMaster\" WHERE \"UserID\" = %s AND \"IsDeleted\" = FALSE", [user_id])
            row = cursor.fetchone()
            if row and row[0]:
                blob = row[0]
                if isinstance(blob, memoryview):
                    blob = blob.tobytes()
                
                # Simple MIME detection
                mime = "image/jpeg"
                if blob.startswith(b'\x89PNG\r\n\x1a\n'):
                    mime = "image/png"
                elif blob.startswith(b'GIF87a') or blob.startswith(b'GIF89a'):
                    mime = "image/gif"
                
                return HttpResponse(blob, content_type=mime)
    except Exception as e:
        logger.error(f"Error serving user photo: {e}")
    
    return HttpResponse(status=404)


@custom_login_required
def serve_school_logo(request, school_id):
    """Serve school logo as binary file for better caching/performance"""
    try:
        from django.db import connection
        from django.http import HttpResponse
        
        with connection.cursor() as cursor:
            cursor.execute("SELECT \"SchoolLogo\" FROM \"SchoolMaster\" WHERE \"SchoolID\" = %s AND \"IsDeleted\" = FALSE", [school_id])
            row = cursor.fetchone()
            if row and row[0]:
                blob = row[0]
                if isinstance(blob, memoryview):
                    blob = blob.tobytes()
                
                # Simple MIME detection
                mime = "image/jpeg"
                if blob.startswith(b'\x89PNG\r\n\x1a\n'):
                    mime = "image/png"
                
                return HttpResponse(blob, content_type=mime)
    except Exception as e:
        logger.error(f"Error serving school logo: {e}")
    
    return HttpResponse(status=404)



# -----------------
# Password Reset Views
# -----------------

def password_reset_request_view(request):
    """
    Handle initial request for password reset by sending an OTP.
    """
    if request.method == "POST":
        identifier = request.POST.get('identifier', '').strip()
        if not identifier:
            messages.error(request, "Email or Username is required.")
            return render(request, "core/password_reset.html")

        try:
            # Check if user exists
            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT u."UserID", u."Email", u."UserName"
                    FROM "UserMaster" u
                    WHERE (u."UserName" = %s OR u."UserCode" = %s OR u."Email" = %s)
                      AND u."IsActive" = TRUE
                      AND u."IsDeleted" IS NOT TRUE
                """, [identifier, identifier, identifier])
                row = cursor.fetchone()

            if not row:
                # Security best practice: don't reveal if user exists. 
                # But for academic/internal apps, sometimes a clear message is preferred.
                # Here we follow the user requirement to be secure.
                messages.success(request, "If an account matches those details, a reset OTP has been sent.")
                return redirect('login')

            user_id, email, user_name = row
            if not email:
                messages.error(request, "No email address associated with this account. Please contact Admin.")
                return render(request, "core/password_reset.html")

            # Reuse OTP logic
            generate_and_store_otp(identifier=identifier, purpose='password_reset', request=request)
            
            resp = redirect('password_reset_confirm')
            resp.set_cookie('reset_identifier', identifier, max_age=600, httponly=True, samesite='Lax')
            messages.success(request, f"A reset OTP has been sent to your registered email.")
            return resp

        except Exception as e:
            logger.error(f"Password reset request failed: {e}", exc_info=True)
            messages.error(request, "Service unavailable. Please try again later.")

    return render(request, "core/password_reset.html")


def password_reset_confirm_view(request):
    """
    Verify OTP and allow setting a new password.
    """
    identifier = request.COOKIES.get('reset_identifier')
    if not identifier:
        messages.error(request, "Reset session expired. Please start again.")
        return redirect('password_reset_request')

    if request.method == "POST":
        otp = request.POST.get('otp', '').strip()
        new_password = request.POST.get('new_password', '').strip()
        confirm_password = request.POST.get('confirm_password', '').strip()

        if not all([otp, new_password, confirm_password]):
            messages.error(request, "All fields are required.")
            return render(request, "core/password_reset_confirm.html")

        if new_password != confirm_password:
            messages.error(request, "Passwords do not match.")
            return render(request, "core/password_reset_confirm.html")
        
        if len(new_password) < 8 or not any(c.isdigit() for c in new_password) or not any(c.isalpha() for c in new_password):
            messages.error(request, "Password must be at least 8 characters long and contain both letters and numbers.")
            return render(request, "core/password_reset_confirm.html")

        try:
            is_valid, _ = verify_otp(identifier, otp, purpose='password_reset')
            if is_valid:
                hashed_pw = make_password(new_password)
                with connection.cursor() as cursor:
                    # Update password and reset strike counters
                    cursor.execute("""
                        UPDATE "UserMaster"
                        SET "PasswordHash" = %s,
                            "FailedLoginAttempts" = 0,
                            "BlockedUntil" = NULL
                        WHERE ("UserName" = %s OR "UserCode" = %s OR "Email" = %s)
                    """, [hashed_pw, identifier, identifier, identifier])

                # Send password changed notification
                try:
                    with connection.cursor() as cursor:
                        cursor.execute('''
                            SELECT 
                                u."Email", u."UserName", p."ProfileName", 
                                s."SchoolName", u."ProfileID", u."SchoolID"
                            FROM "UserMaster" u
                            LEFT JOIN "ProfileMaster" p ON u."ProfileID" = p."ProfileID"
                            LEFT JOIN "SchoolMaster" s ON u."SchoolID" = s."SchoolID"
                            WHERE u."UserName" = %s OR u."UserCode" = %s OR u."Email" = %s
                        ''', [identifier, identifier, identifier])
                        res = cursor.fetchone()
                        
                        if res and res[0]:
                            target_email, u_name, profile, s_name, p_id, s_id = res
                            
                            # Determine Branding
                            branding_name = get_branding_title(p_id, s_name)
                            
                            placeholders = {
                                'user_name': u_name,
                                'login_id': identifier,
                                'profile': profile,
                                'school_name': branding_name,
                                'school_logo': None,
                                'ip_address': request.META.get('REMOTE_ADDR', 'Unknown'),
                                'browser': request.META.get('HTTP_USER_AGENT', 'Unknown'),
                                'timestamp': timezone.now().strftime('%Y-%m-%d %H:%M:%S')
                            }
                            send_email_by_code(
                                code='PASSWORD_CHANGED_NOTIFICATION',
                                to_emails=[target_email],
                                placeholders=placeholders,
                                school_id=s_id
                            )
                except Exception as ex:
                    logger.error(f"Failed to send password changed notification: {ex}")
                
                messages.success(request, "Password reset successful. You can now login with your new password.")
                resp = redirect('login')
                resp.delete_cookie('reset_identifier')
                return resp
            else:
                messages.error(request, "Invalid or expired OTP.")
        except Exception as e:
            logger.error(f"Password reset confirmation failed: {e}", exc_info=True)
            messages.error(request, "Service unavailable.")

    return render(request, "core/password_reset_confirm.html")


# definitive reload trigger 02/20/2026 10:05:00

# core/data_import/templates_generator.py
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from django.conf import settings


def generate_excel_template(import_type):
    """Generate Excel template with headers and sample data"""
    
    templates = {
        'Students': {
            'columns': [
                'StudentCode', 'FullName*', 'Gender*', 'DateOfBirth*', 'Age', 'BloodGroup',
                'Category', 'Religion', 'Nationality', 'StudentAadhaar', 'MotherTongue',
                'PresentAddress', 'PermanentAddress', 'District', 'State', 'Country',
                'ParentMobile*', 'AlternateNumber', 'Email', 'FatherName*', 'FatherOccupation',
                'FatherQualification', 'FatherAadhaar', 'FatherMobile', 'MotherName',
                'MotherOccupation', 'MotherQualification', 'MotherAadhaar', 'MotherMobile',
                'GuardianName', 'GuardianRelation', 'GuardianMobile', 'LastSchool', 'LastClass',
                'TCNumber', 'MediumOfInstruction', 'AdmissionClass*', 'Section*', 'Stream',
                'ModeOfAdmission', 'AdmissionDate*'
            ],
            'sample_data': [
                ['STU001', 'John Doe', 'Male', '2010-05-15', '14', 'O+', 'General', 'Hindu',
                 'Indian', '123456789012', 'Hindi', '123 Main St', '123 Main St', 'Mumbai',
                 'Maharashtra', 'India', '9876543210', '9876543211', 'john@example.com',
                 'Mr. Doe', 'Engineer', 'B.Tech', '234567890123', '9876543212',
                 'Mrs. Doe', 'Teacher', 'M.A.', '345678901234', '9876543213',
                 'Mr. Guardian', 'Uncle', '9876543214', 'ABC School', 'Class 5',
                 'TC12345', 'English', '1', '1', 'Science', 'Regular', '2024-04-01']
            ],
            'instructions': [
                '* = Required field',
                'StudentCode: Leave blank for auto-generation',
                'Gender: Male/Female/Other',
                'DateOfBirth: YYYY-MM-DD format',
                'Age: Auto-calculated if blank',
                'StudentAadhaar, FatherAadhaar, MotherAadhaar: 12 digits (optional)',
                'Mobile numbers: 10 digits starting with 6-9',
                'AdmissionClass: ClassID from your school (required)',
                'Section: SectionID from your school (required)',
                'AdmissionDate: YYYY-MM-DD format',
                'All address fields are optional but recommended',
                'Parent/Guardian information helps in communication'
            ]
        },
        'Teachers': {
            'columns': [
                'EmployeeCode', 'EmployeeName*', 'Email*', 'Phone*', 'ProfileID',
                'DateOfJoining*', 'Qualification', 'Experience', 'Address'
            ],
            'sample_data': [
                ['EMP001', 'Jane Smith', 'jane@example.com', '9876543210', '3',
                 '2024-01-01', 'M.Sc', '5 years', '456 Teacher St']
            ],
            'instructions': [
                '* = Required field',
                'EmployeeCode: Leave blank for auto-generation',
                'ProfileID: 3 for Teacher (default)',
                'Phone: 10 digits starting with 6-9',
                'DateOfJoining: YYYY-MM-DD format'
            ]
        },
        'Salary': {
            'columns': [
                'EmployeeCode*', 'ComponentName*', 'Amount*'
            ],
            'sample_data': [
                ['EMP001', 'Basic Salary', '25000'],
                ['EMP001', 'HRA', '10000'],
                ['EMP001', 'PF Deduction', '2500']
            ],
            'instructions': [
                '* = Required field',
                'EmployeeCode: Must exist in system',
                'ComponentName: Must exist in SalaryComponentMaster',
                'Amount: Positive decimal number'
            ]
        },
        'Fee': {
            'columns': [
                'StudentCode*', 'FeeTypeName*', 'FeeMonth*', 'FeeAmount*',
                'DiscountPercentage', 'FinalAmount*'
            ],
            'sample_data': [
                ['STU001', 'Tuition Fee', '202404', '5000', '10', '4500'],
                ['STU001', 'Transport Fee', '202404', '1000', '0', '1000']
            ],
            'instructions': [
                '* = Required field',
                'StudentCode: Must exist in system',
                'FeeTypeName: Must exist in FeeType_Master',
                'FeeMonth: YYYYMM format (e.g., 202404 for April 2024)',
                'DiscountPercentage: 0-100',
                'FinalAmount: Auto-calculated if blank'
            ]
        },
        'Attendance': {
            'columns': [
                'StudentCode*', 'AttendanceDate*', 'Status*', 'Remarks'
            ],
            'sample_data': [
                ['STU001', '2024-04-01', 'Present', ''],
                ['STU002', '2024-04-01', 'Absent', 'Sick leave']
            ],
            'instructions': [
                '* = Required field',
                'StudentCode: Must exist in system',
                'AttendanceDate: YYYY-MM-DD format',
                'Status: Present/Absent/Leave/Holiday'
            ]
        },
        'Exam': {
            'columns': [
                'ExamName*', 'ExamType', 'ClassID*', 'StartDate*', 'EndDate*',
                'AcademicYearId*', 'IsActive'
            ],
            'sample_data': [
                ['Mid Term Exam', 'Mid Term', '1', '2024-07-01', '2024-07-10', '1', '1']
            ],
            'instructions': [
                '* = Required field',
                'ExamType: Unit Test/Mid Term/Final',
                'ClassID: Must exist for your school',
                'Dates: YYYY-MM-DD format',
                'AcademicYearId: Must exist for your school',
                'IsActive: 1 for active, 0 for inactive'
            ]
        },
        'ExamResult': {
            'columns': [
                'StudentCode*', 'ExamID*', 'SubjectName*', 'MarksObtained*',
                'MaxMarks*', 'Grade', 'Remarks'
            ],
            'sample_data': [
                ['STU001', '1', 'Mathematics', '85', '100', 'A', 'Excellent'],
                ['STU001', '1', 'Science', '78', '100', 'B+', 'Good']
            ],
            'instructions': [
                '* = Required field',
                'StudentCode: Must exist in system',
                'ExamID: Must exist for your school',
                'SubjectName: Must exist in SubjectMaster',
                'MarksObtained: Must be <= MaxMarks'
            ]
        },
        'ClassMaster': {
            'columns': [
                'ClassName*', 'ClassCode*', 'EducationLevel', 'Description'
            ],
            'sample_data': [
                ['Class 1', 'CLS1', 'Primary', 'First standard'],
                ['Class 2', 'CLS2', 'Primary', 'Second standard']
            ],
            'instructions': [
                '* = Required field',
                'ClassName: Unique within school',
                'ClassCode: Unique code for class'
            ]
        },
        'SectionMaster': {
            'columns': [
                'ClassName*', 'SectionName*', 'Capacity', 'RoomNumber'
            ],
            'sample_data': [
                ['Class 1', 'A', '40', 'R101'],
                ['Class 1', 'B', '40', 'R102']
            ],
            'instructions': [
                '* = Required field',
                'ClassName: Must exist in ClassMaster',
                'SectionName: Section identifier (A, B, C, etc.)',
                'Capacity: Maximum students (optional)'
            ]
        },
        'SubjectMaster': {
            'columns': [
                'SubjectName*', 'SubjectCode*', 'ClassName*', 'Description'
            ],
            'sample_data': [
                ['Mathematics', 'MATH', 'Class 1', 'Basic mathematics'],
                ['Science', 'SCI', 'Class 1', 'General science']
            ],
            'instructions': [
                '* = Required field',
                'SubjectName: Name of subject',
                'SubjectCode: Unique code',
                'ClassName: Must exist in ClassMaster'
            ]
        }
    }
    
    template_config = templates.get(import_type)
    if not template_config:
        return None
    
    # Create workbook
    wb = Workbook()
    
    # Data sheet
    ws_data = wb.active
    ws_data.title = "Data"
    
    # Header row styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Write headers
    for col_idx, col_name in enumerate(template_config['columns'], start=1):
        cell = ws_data.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Write sample data
    for row_idx, row_data in enumerate(template_config['sample_data'], start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws_data.cell(row=row_idx, column=col_idx, value=value)
    
    # Auto-adjust column widths
    for column in ws_data.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_data.column_dimensions[column_letter].width = adjusted_width
    
    # Instructions sheet
    ws_instructions = wb.create_sheet("Instructions")
    ws_instructions.column_dimensions['A'].width = 80
    
    instruction_font = Font(size=11)
    title_font = Font(size=14, bold=True)
    
    ws_instructions.cell(row=1, column=1, value=f"{import_type} Import Instructions").font = title_font
    ws_instructions.cell(row=2, column=1, value="")
    
    for idx, instruction in enumerate(template_config['instructions'], start=3):
        cell = ws_instructions.cell(row=idx, column=1, value=instruction)
        cell.font = instruction_font
        cell.alignment = Alignment(wrap_text=True)
    
    # Save file
    template_dir = os.path.join(settings.MEDIA_ROOT, 'templates')
    os.makedirs(template_dir, exist_ok=True)
    
    file_path = os.path.join(template_dir, f"{import_type}_Import_Template.xlsx")
    wb.save(file_path)
    
    return file_path

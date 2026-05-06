# core/school_views.py
# School management views - extracted for better code organization

import base64
import re
import time
import json
import csv
import logging
import io
from datetime import datetime
from urllib.parse import urlparse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from django.shortcuts import render, redirect
from django.contrib import messages
from django.http import JsonResponse, HttpResponseRedirect, HttpResponse
from django.db import connection
from django.urls import reverse
from django.core.paginator import Paginator

from .decorators import custom_login_required
from .utils import (
    get_context, safe_int, safe_strptime, validate_uploaded_file,
    ALLOWED_IMAGE_TYPES
)
from .url_encryption import decrypt_id
from .models import SchoolMaster

logger = logging.getLogger(__name__)


@custom_login_required
def get_school_list_api(request):
    """API to fetch list of schools for dropdowns"""
    try:
        schools = SchoolMaster.objects.filter(is_deleted=False).values(
            'school_id', 'school_code', 'school_name'
        ).order_by('school_name')
        
        school_list = []
        for school in schools:
            code = school['school_code'] if school['school_code'] else 'N/A'
            school_list.append({
                'id': school['school_id'],
                'code': code,
                'name': school['school_name'],
                'display_name': f"[{code}] {school['school_name']}"
            })
            
        return JsonResponse({'schools': school_list})
    except Exception as e:
        logger.error(f"Error fetching school list: {str(e)}")
        return JsonResponse({'error': 'Failed to fetch schools'}, status=500)


@custom_login_required
def schools_create(request):
    """View to create a new school"""
    
    # Clear any existing messages to prevent cross-talk from previous actions
    # This ensures the Create School page starts with a clean slate
    if request.method == 'GET':
        storage = messages.get_messages(request)
        for _ in storage:
            pass

    context = get_context(request)
    if context.get('profile_name') != 'Super Admin':
        messages.error(request, "Only Super Admins can create new schools.")
        return redirect('schools_list')

    if request.method == 'POST':
        # Extract form data
        school_name = request.POST.get('school_name')
        registration_number = request.POST.get('registration_number') or None
        address = request.POST.get('address') or None
        district = request.POST.get('district')
        state = request.POST.get('state')
        country = request.POST.get('country')
        pincode = request.POST.get('pincode') or None
        phone = request.POST.get('phone') or None
        email = request.POST.get('email') or None
        website = request.POST.get('website') or None
        created_by = request.session.get('UserId')
        
        logo_data = None
        if 'logo' in request.FILES:
            file = request.FILES['logo']
            from .utils import validate_uploaded_file
            is_valid, error_msg = validate_uploaded_file(file)
            if not is_valid:
                return JsonResponse({'status': 'ERROR', 'message': f'School logo validation failed: {error_msg}'}, status=400)
            is_valid, message = validate_uploaded_file(file, ALLOWED_IMAGE_TYPES)
            if not is_valid:
                messages.error(request, message)
                return render(request, 'create_school.html', get_context(request))
            logo_data = file.read()
        
        board_id = request.POST.get('board') or None
        medium_id = request.POST.get('medium') or None
        principal_name = request.POST.get('principal_name') or None
        principal_contact_mail = request.POST.get('principal_contact_mail') or None
        principal_contact_phone = request.POST.get('principal_contact_phone') or None
        director_name = request.POST.get('director_name') or None
        director_contact_phone = request.POST.get('director_contact_phone') or None
        director_contact_email = request.POST.get('director_contact_email') or None
        establish_date = request.POST.get('establish_date') or None

        # Validate created_by
        if not created_by:
            logger.warning("No UserId found in session for school creation")
            messages.error(request, "Session expired. Please login again.")
            return redirect('login')

        # Validate school name
        if not school_name or len(school_name) > 255:
            logger.warning(f"Invalid school name: {school_name}")
            messages.error(request, 'School name is required and must be 255 characters or less.')
            return render(request, 'create_school.html', get_context(request))

        # Convert district, state, country, board_id, medium_id to integers
        try:
            district = int(district) if district else None
            state = int(state) if state else None
            country = int(country) if country else None
            board_id = int(board_id) if board_id else None
            medium_id = int(medium_id) if medium_id else None
        except (ValueError, TypeError) as e:
            logger.warning(f"Invalid geographic/board/medium selection: {str(e)}")
            messages.error(request, 'Invalid geographic, board, or medium selection.')
            return render(request, 'create_school.html', get_context(request))

        # Validate phone numbers
        phone_fields = [
            ('phone', phone),
            ('principal_contact_phone', principal_contact_phone),
            ('director_contact_phone', director_contact_phone)
        ]
        for field_name, value in phone_fields:
            if value:
                # Remove dashes and check if it's exactly 10 digits
                clean_value = value.replace('-', '').replace(' ', '')
                if not (clean_value.isdigit() and len(clean_value) == 10):
                    logger.warning(f"Invalid {field_name}: {value}")
                    messages.error(request, f'{field_name.replace("_", " ").title()} must be exactly 10 digits.')
                    return render(request, 'create_school.html', get_context(request))
                # Store the clean value for database
                if field_name == 'phone':
                    phone = clean_value
                elif field_name == 'principal_contact_phone':
                    principal_contact_phone = clean_value
                elif field_name == 'director_contact_phone':
                    director_contact_phone = clean_value

        # Validate emails
        email_regex = r'^[^\s@]+@[^\s@]+\.[^\s@]+$'
        email_fields = [
            ('email', email),
            ('principal_contact_mail', principal_contact_mail),
            ('director_contact_email', director_contact_email)
        ]
        for field_name, value in email_fields:
            if value and not re.match(email_regex, value):
                logger.warning(f"Invalid {field_name}: {value}")
                messages.error(request, f'{field_name.replace("_", " ").title()} must be a valid email address.')
                return render(request, 'create_school.html', get_context(request))

        # Validate pincode
        if pincode and not (pincode.isdigit() and len(pincode) == 6):
            logger.warning(f"Invalid pincode: {pincode}")
            messages.error(request, 'Pincode must be exactly 6 digits.')
            return render(request, 'create_school.html', get_context(request))

        # Validate website
        if website:
            try:
                parsed_url = urlparse(website)
                if not all([parsed_url.scheme, parsed_url.netloc]):
                    raise ValueError("Invalid URL")
            except:
                logger.warning(f"Invalid website URL: {website}")
                messages.error(request, 'Please enter a valid URL (e.g., https://example.com).')
                return render(request, 'create_school.html', get_context(request))

        # Validate establish date
        if establish_date:
            try:
                input_date = safe_strptime(establish_date, '%Y-%m-%d').date()
                if input_date > datetime.now().date():
                    logger.warning(f"Future establish date: {establish_date}")
                    messages.error(request, 'Establish date cannot be in the future.')
                    return render(request, 'create_school.html', get_context(request))
            except ValueError as e:
                logger.warning(f"Invalid establish date format: {establish_date}")
                messages.error(request, 'Invalid establish date format.')
                return render(request, 'create_school.html', get_context(request))

        # Check for duplicate school
        with connection.cursor() as cursor:
            query = """
                SELECT COUNT(*) FROM "SchoolMaster" 
                WHERE ("SchoolName" = %s OR "RegistrationNumber" = %s) AND "IsDeleted" = FALSE
            """
            params = [school_name, registration_number or '']
            cursor.execute(query, params)
            count = cursor.fetchone()[0]
            if count > 0:
                logger.warning(f"Duplicate school detected: {school_name}, {registration_number}")
                messages.error(request, 'A school with this name or registration number already exists.')
                return render(request, 'create_school.html', get_context(request))

        # Call stored procedure
        try:
            with connection.cursor() as cursor:
                cursor.execute(
                    """
                    SELECT * FROM "Proc_CreateSchool_Set"(
                        %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                        %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s
                    )
                    """,
                    [
                        school_name,
                        registration_number,
                        address,
                        district,
                        state,
                        country,
                        pincode,
                        phone,
                        email,
                        website,
                        created_by,
                        logo_data,
                        board_id,
                        medium_id,
                        principal_name,
                        principal_contact_mail,
                        principal_contact_phone,
                        director_name,
                        director_contact_phone,
                        director_contact_email,
                        establish_date,
                    ]
                )
                result = cursor.fetchone()
                if result:
                    school_code, status = result
                else:
                    school_code, status = None, 'Error: No response from stored procedure.'
                    logger.error("No response from Proc_CreateSchool_Set")
        except Exception as e:
            logger.error(f"Error executing stored procedure: {str(e)}")
            messages.error(request, f'Error creating school: {str(e)}')
            return render(request, 'create_school.html', get_context(request))

        # Handle the result
        if school_code is not None:
            logger.info(f"School created successfully: {school_code}, {status}")
            messages.success(request, f"School '{school_name}' created successfully: {status}")
            return redirect('schools_create')
        else:
            logger.error(f"Failed to create school: {status}")
            messages.error(request, status)
            return render(request, 'create_school.html', get_context(request))

    # GET request
    context = get_context(request)
    return render(request, 'create_school.html', context)



@custom_login_required
def schools_list(request):
    # Get user context for header
    context = get_context(request)
    
    schools = []
    page_number = safe_int(request.GET.get('page', 1))
    page_size = safe_int(request.GET.get('per_page', 10))  # Support configurable page size

    # Search and filter parameters
    search_params = {
        'school_code': request.GET.get('school_code', ''),
        'school_name': request.GET.get('school_name', ''),
        'registration_number': request.GET.get('registration_number', ''),
        'created_at': request.GET.get('created_at', ''),
        'to_date': request.GET.get('to_date', ''),
        'phone': request.GET.get('phone', ''),
        'email': request.GET.get('email', ''),
        'principal_name': request.GET.get('principal_name', ''),
        'director_name': request.GET.get('director_name', ''),
        'board': request.GET.get('board', ''),
        'medium': request.GET.get('medium', ''),
        'status': request.GET.get('status', ''),
        'country': request.GET.get('country', ''),
        'state': request.GET.get('state', ''),
        'district': request.GET.get('district', ''),
        'pincode': request.GET.get('pincode', ''),
        'show_deleted': request.GET.get('show_deleted', '0'),
    }

    # Convert string IDs to integers where applicable
    try:
        board = int(search_params['board']) if search_params['board'] else None
        medium = int(search_params['medium']) if search_params['medium'] else None
        country = int(search_params['country']) if search_params['country'] else None
        state = int(search_params['state']) if search_params['state'] else None
        district = int(search_params['district']) if search_params['district'] else None
    except ValueError:
        board = medium = country = state = district = None

    # Convert date strings to datetime objects
    from_date = None
    to_date = None
    try:
        if search_params['created_at']:
            from_date = safe_strptime(search_params['created_at'], '%Y-%m-%d').date()
        if search_params['to_date']:
            to_date = safe_strptime(search_params['to_date'], '%Y-%m-%d').date()
    except ValueError:
        search_params['created_at'] = ''
        search_params['to_date'] = ''

    sort_column = request.GET.get('order_by', request.GET.get('sort_column', 'SchoolID'))
    sort_direction = request.GET.get('order_direction', request.GET.get('sort_direction', 'DESC'))

    # Get session data for filtering
    user_id = request.session.get('UserId')
    school_id = request.session.get('SchoolID')

    # If no userId, force login
    if not user_id:
        messages.error(request, "Session expired. Please login again.")
        return redirect('login')

    # Fetch profile_id if not in session
    profile_id = request.session.get('ProfileID')
    if not profile_id:
        try:
            with connection.cursor() as cursor:
                cursor.execute(
                    "SELECT \"ProfileID\" FROM \"UserMaster\" WHERE \"UserID\" = %s AND \"IsDeleted\" = FALSE",
                    [user_id]
                )
                result = cursor.fetchone()
                if result:
                    profile_id = result[0]
                    request.session['ProfileID'] = profile_id
                else:
                    messages.error(request, "Invalid user data.")
                    return redirect('login')
        except Exception as e:
            logger.error(f"Error fetching user profile: {e}")
            messages.error(request, "Error fetching user data.")
            return redirect('login')

    # Fetch user photo
    try:
        with connection.cursor() as cursor:
            cursor.execute(
                "SELECT \"UserPhoto\" FROM \"UserMaster\" WHERE \"UserID\" = %s AND \"IsDeleted\" = FALSE",
                [user_id]
            )
            photo_data = cursor.fetchone()
            if photo_data and photo_data[0]:
                user_photo_src = f"data:image/jpeg;base64,{base64.b64encode(photo_data[0]).decode('utf-8')}"
    except Exception as e:
        logger.error(f"Error fetching user photo: {e}")

    # Fetch school logo if exists
    if school_id:
        try:
            with connection.cursor() as cursor:
                cursor.execute(
                    "SELECT \"SchoolLogo\" FROM \"SchoolMaster\" WHERE \"SchoolID\" = %s AND \"IsDeleted\" = FALSE",
                    [school_id]
                )
                logo_data = cursor.fetchone()
                if logo_data and logo_data[0]:
                    school_logo_src = f"data:image/jpeg;base64,{base64.b64encode(logo_data[0]).decode('utf-8')}"
        except Exception as e:
            logger.error(f"Error fetching school logo: {e}")

    # Fetch board and medium options
    boards = []
    mediums = []
    try:
        with connection.cursor() as cursor:
            cursor.execute('SELECT "BoardID", "BoardName" FROM "Board_Master" WHERE "IsDeleted" = FALSE')
            boards = [{'BoardID': row[0], 'BoardName': row[1]} for row in cursor.fetchall()]
            cursor.execute('SELECT "MediumID", "MediumName" FROM "Medium_Master" WHERE "IsDeleted" = FALSE')
            mediums = [{'MediumID': row[0], 'MediumName': row[1]} for row in cursor.fetchall()]
    except Exception as e:
        logger.error(f"Error fetching boards or mediums: {e}")

    # Role-based filtering logic
    filter_school_id = None
    if context.get('profile_name') == 'School Admin':
        filter_school_id = school_id
    
    # Fetch schools
    total_schools = 0
    active_schools = 0
    deleted_schools = 0
    try:
        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT * FROM "Proc_SchoolList_get"(
                    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s
                )
                """,
                [
                    from_date,
                    to_date,
                    search_params['school_code'] or None,
                    search_params['school_name'] or None,
                    country,
                    state,
                    district,
                    search_params['phone'] or None,
                    search_params['email'] or None,
                    board,
                    medium,
                    search_params['pincode'] or None,
                    search_params['registration_number'] or None,
                    search_params['principal_name'] or None,
                    search_params['director_name'] or None,
                    search_params['status'] or None,
                    bool(int(search_params['show_deleted'])),
                    sort_column,
                    sort_direction,
                    page_number,
                    page_size,
                    user_id,
                    filter_school_id  # Added role-based filter
                ]
            )
            columns = [col[0] for col in cursor.description]
            rows = cursor.fetchall()
            schools = [dict(zip(columns, row)) for row in rows]
            if schools:
                total_schools = schools[0].get('TotalCount', 0)
                active_schools = schools[0].get('ActiveSchools', 0)
                deleted_schools = schools[0].get('DeletedSchools', 0)

        # Convert SchoolLogo to base64
        for school in schools:
            if school.get('SchoolLogo'):
                school['SchoolLogo'] = f"data:image/jpeg;base64,{base64.b64encode(school['SchoolLogo']).decode('utf-8')}"
    except Exception as e:
        logger.error(f"Error fetching schools: {e}")
        messages.error(request, "An error occurred while fetching schools.")
        schools = []

    # Paginate results
    paginator = Paginator(schools, page_size)
    page_obj = paginator.get_page(page_number)

    # Update context with schools list specific data
    import json
    context.update({
        'page_obj': page_obj,
        'total_schools': total_schools,
        'active_schools': active_schools,
        'deleted_schools': deleted_schools,
        'boards': boards,
        'mediums': mediums,
        'boards_json': json.dumps(boards),
        'mediums_json': json.dumps(mediums),
        'search_params': search_params,
        'sort_column': sort_column,
        'sort_direction': sort_direction,
        'page_size': page_size,
    })

    return render(request, 'schools_list.html', context)


@custom_login_required
def schools_export(request):
    """Export schools list to Excel or CSV"""
    context = get_context(request)
    user_id = context.get('user_id')
    school_id = request.session.get('SchoolID')
    profile_name = context.get('profile_name')
    
    # Extract parameters
    format_type = request.GET.get('format', 'excel').lower()
    school_code = request.GET.get('school_code', '')
    school_name = request.GET.get('school_name', '')
    registration_number = request.GET.get('registration_number', '')
    created_at = request.GET.get('created_at', '')
    to_date = request.GET.get('to_date', '')
    phone = request.GET.get('phone', '')
    email = request.GET.get('email', '')
    principal_name = request.GET.get('principal_name', '')
    director_name = request.GET.get('director_name', '')
    board = safe_int(request.GET.get('board'), default=None)
    medium = safe_int(request.GET.get('medium'), default=None)
    pincode = request.GET.get('pincode', '')
    status = request.GET.get('status', '')
    show_deleted = safe_int(request.GET.get('show_deleted', '0'), default=0)
    sort_column = request.GET.get('order_by', request.GET.get('sort_column', 'SchoolID'))
    sort_direction = request.GET.get('order_direction', request.GET.get('sort_direction', 'DESC'))

    # Convert geographic IDs
    country = safe_int(request.GET.get('country'), default=None)
    state = safe_int(request.GET.get('state'), default=None)
    district = safe_int(request.GET.get('district'), default=None)

    # Date parsing
    from_date = safe_strptime(created_at, '%Y-%m-%d') if created_at else None
    to_date_parsed = safe_strptime(to_date, '%Y-%m-%d') if to_date else None

    # Role-based filtering
    filter_school_id = None
    if profile_name == 'School Admin':
        filter_school_id = school_id

    # Fetch data (using a large page size to get all records)
    schools = []
    try:
        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT * FROM "Proc_SchoolList_get"(
                    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s
                )
                """,
                [
                    from_date, to_date_parsed, school_code or None, school_name or None,
                    country, state, district, phone or None, email or None,
                    board, medium, pincode or None, registration_number or None,
                    principal_name or None, director_name or None, status or None,
                    bool(show_deleted), sort_column, sort_direction,
                    1, 1000000, user_id, filter_school_id
                ]
            )
            columns = [col[0] for col in cursor.description]
            # Exclude metadata columns
            exclude_cols = {'TotalCount', 'ActiveSchools', 'DeletedSchools', 'SchoolLogo', 'SchoolID', 'BoardID', 'MediumID', 'CountryID', 'StateID', 'DistrictID', 'CreatedBy', 'UpdatedBy', 'UpdatedAt', 'IsDeleted'}
            export_columns = [col for col in columns if col not in exclude_cols]
            
            rows = cursor.fetchall()
            schools = [dict(zip(columns, row)) for row in rows]
    except Exception as e:
        logger.error(f"Export fetch error: {e}")
        return HttpResponse("Error fetching data for export", status=500)

    filename = f"Schools_List_{datetime.now().strftime('%Y%m%d_%H%M%S')}"

    if format_type == 'csv':
        delimiter_type = request.GET.get('delimiter', 'comma').lower()
        delimiters = {
            'comma': ',',
            'pipe': '|',
            'tab': '\t'
        }
        actual_delimiter = delimiters.get(delimiter_type, ',')
        
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
        
        writer = csv.DictWriter(response, fieldnames=export_columns, extrasaction='ignore', delimiter=actual_delimiter)
        writer.writeheader()
        for school in schools:
            # Format dates
            for key, value in school.items():
                if isinstance(value, datetime):
                    school[key] = value.strftime('%Y-%m-%d %H:%M:%S')
            writer.writerow(school)
        return response

    else:  # Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Schools List"
        
        header_fill = PatternFill(start_color="004aad", end_color="004aad", fill_type="solid")
        
        # Write headers
        for col_idx, col_name in enumerate(export_columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            
        # Write data
        for row_idx, school in enumerate(schools, start=2):
            for col_idx, col_name in enumerate(export_columns, start=1):
                value = school.get(col_name)
                if isinstance(value, datetime):
                    value = value.replace(tzinfo=None)
                ws.cell(row=row_idx, column=col_idx, value=value)
                
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except: pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
        wb.save(response)
        return response

    return render(request, 'schools_list.html', context)

# School Delete and Restore Views

@custom_login_required
def school_soft_delete(request, school_id):
    """Soft delete a school using the stored procedure"""
    if request.session.get('ProfileName') != 'Super Admin':
        return JsonResponse({'success': False, 'message': 'Only Super Admins can delete schools.'})
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method'})
    
    try:
        user_id = request.session.get('UserId')
        if not user_id:
            return JsonResponse({'success': False, 'message': 'Session expired. Please login again.'})
        
        with connection.cursor() as cursor:
            cursor.execute(
                'SELECT "Proc_SoftDeleteSchool"(%s, %s, %s)',
                [school_id, user_id, 'DELETE']
            )
            
        return JsonResponse({'success': True, 'message': 'School deleted successfully.'})
        
    except Exception as e:
        logger.error(f"Error deleting school {school_id}: {e}")
        return JsonResponse({'success': False, 'message': 'An error occurred while deleting the school.'})


@custom_login_required
def school_restore(request, school_id):
    """Restore a soft deleted school using the stored procedure"""
    if request.session.get('ProfileName') != 'Super Admin':
        return JsonResponse({'success': False, 'message': 'Only Super Admins can restore schools.'})
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method'})
    
    try:
        user_id = request.session.get('UserId')
        if not user_id:
            return JsonResponse({'success': False, 'message': 'Session expired. Please login again.'})
        
        with connection.cursor() as cursor:
            cursor.execute(
                'SELECT "Proc_SoftDeleteSchool"(%s, %s, %s)',
                [school_id, user_id, 'Restore']
            )
            
        return JsonResponse({'success': True, 'message': 'School restored successfully.'})
        
    except Exception as e:
        logger.error(f"Error restoring school {school_id}: {e}")
        return JsonResponse({'success': False, 'message': 'An error occurred while restoring the school.'})

# School Update Views

@custom_login_required
def school_update(request, encrypted_id):
    """View to display school update form with pre-filled data"""
    school_id = decrypt_id(encrypted_id)
    if not school_id:
        messages.error(request, "Invalid or expired link.")
        return redirect('schools_list')

    # Get user context for header
    context = get_context(request)
    
    # Role-based permission check
    profile_name = context.get('profile_name')
    user_school_id = request.session.get('SchoolID')
    
    if profile_name == 'School Admin':
        if str(school_id) != str(user_school_id):
            logger.warning(f"Unauthorized update attempt by School Admin for SchoolID {school_id}. User is assigned to SchoolID {user_school_id}")
            messages.error(request, "You do not have permission to update other schools.")
            return redirect('schools_list')
    
    # Fetch school details
    school_data = None
    try:
        with connection.cursor() as cursor:
            cursor.execute('SELECT * FROM "Proc_GetSchoolDetails_ByID"(%s)', [school_id])
            columns = [col[0] for col in cursor.description]
            rows = cursor.fetchall()
            if rows:
                school_data = dict(zip(columns, rows[0]))
                
                # Check if there was an error
                if school_data.get('Status') == 'Error':
                    messages.error(request, school_data.get('ErrorMessage', 'Error retrieving school details'))
                    return redirect('schools_list')
                
                # Map keys for template compatibility
                school_data['PrincipalPhone'] = school_data.get('PrincipalContactPhone')
                school_data['PrincipalEmail'] = school_data.get('PrincipalContactMail')
                school_data['DirectorPhone'] = school_data.get('DirectorContactPhone')
                school_data['DirectorEmail'] = school_data.get('DirectorContactEmail')

                # Handle School Logo
                if school_data.get('SchoolLogo'):
                    logo = school_data['SchoolLogo']
                    # Handle memoryview (from certain DB drivers)
                    if isinstance(logo, memoryview):
                        try:
                            logo = bytes(logo)
                        except Exception:
                            pass

                    if isinstance(logo, bytes):
                        # Check if already base64 (PNG or JPG signature)
                        if logo.startswith(b'iVBORw') or logo.startswith(b'/9j/'):
                            try:
                                logo = logo.decode('utf-8')
                            except:
                                pass # Keep as bytes if decode fails
                        else:
                            logo = base64.b64encode(logo).decode('utf-8')
                    
                    if not str(logo).startswith('data:image'):
                        school_data['SchoolLogo'] = f"data:image/png;base64,{logo}"
                    else:
                        school_data['SchoolLogo'] = logo
    except Exception as e:
        logger.error(f"Error fetching school details: {e}")
        messages.error(request, "An error occurred while fetching school details.")
        return redirect('schools_list')
    
    if not school_data:
        messages.error(request, "School not found.")
        return redirect('schools_list')
    
    # Dropdown data will be loaded via JavaScript API calls
    # This matches the approach used in the create school page
    
    # Update context with school data
    context.update({
        'school_data': school_data,
        'school_id': school_id,
        'encrypted_id': encrypted_id,
    })
    
    return render(request, 'update_school.html', context)

@custom_login_required
def school_update_submit(request, encrypted_id):
    """View to handle school update form submission"""
    school_id = decrypt_id(encrypted_id)
    if not school_id:
        messages.error(request, "Invalid or expired link.")
        return redirect('schools_list')

    if request.method != 'POST':
        return redirect('school_update', encrypted_id=encrypted_id)
    
    # Role-based permission check for submission
    profile_name = request.session.get('ProfileName')
    user_school_id = request.session.get('SchoolID')
    
    if profile_name == 'School Admin':
        if str(school_id) != str(user_school_id):
            logger.warning(f"Unauthorized update submission attempt by School Admin for SchoolID {school_id}")
            return HttpResponseForbidden("You do not have permission to update this school.")

    # Get form data
    school_name = request.POST.get('school_name', '').strip()
    registration_number = request.POST.get('registration_number', '').strip()
    address = request.POST.get('address', '').strip()
    district = request.POST.get('district', '').strip()
    state = request.POST.get('state', '').strip()
    country = request.POST.get('country', '').strip()
    pincode = request.POST.get('pincode', '').strip()
    phone = request.POST.get('phone', '').strip()
    email = request.POST.get('email', '').strip()
    website = request.POST.get('website', '').strip()
    board_id = request.POST.get('board', '').strip()
    medium_id = request.POST.get('medium', '').strip()
    principal_name = request.POST.get('principal_name', '').strip()
    principal_contact_mail = request.POST.get('principal_contact_mail', '').strip()
    principal_contact_phone = request.POST.get('principal_contact_phone', '').strip()
    director_name = request.POST.get('director_name', '').strip()
    director_contact_phone = request.POST.get('director_contact_phone', '').strip()
    director_contact_email = request.POST.get('director_contact_email', '').strip()
    establish_date = request.POST.get('establish_date', '').strip()
    
    # Debug logging
    logger.info(f"Form data received for school {school_id}:")
    logger.info(f"  School Name: {school_name}")
    logger.info(f"  Country: {country}")
    logger.info(f"  State: {state}")
    logger.info(f"  District: {district}")
    logger.info(f"  Board ID: {board_id}")
    logger.info(f"  Medium ID: {medium_id}")
    
    # Get user ID for UpdatedBy
    user_id = request.session.get('UserId')
    if not user_id:
        messages.error(request, 'User session expired. Please login again.')
        return redirect('login')
    
    # Validate required fields
    if not school_name:
        messages.error(request, 'School name is required.')
        return render(request, 'update_school.html', get_context(request))
    
    # Convert district, state, country, board_id, medium_id to integers
    try:
        district = int(district) if district else None
        state = int(state) if state else None
        country = int(country) if country else None
        board_id = int(board_id) if board_id else None
        medium_id = int(medium_id) if medium_id else None
    except (ValueError, TypeError) as e:
        logger.warning(f"Invalid geographic/board/medium selection: {str(e)}")
        messages.error(request, 'Invalid geographic, board, or medium selection.')
        return render(request, 'update_school.html', get_context(request))
    
    # Validate phone numbers
    phone_fields = [
        ('phone', phone),
        ('principal_contact_phone', principal_contact_phone),
        ('director_contact_phone', director_contact_phone)
    ]
    for field_name, value in phone_fields:
        if value:
            logger.info(f"Validating {field_name}: '{value}' (type: {type(value)})")
            
            # Convert to string and remove all non-digit characters except +
            clean_value = str(value).strip()
            
            # Remove all formatting characters
            import re
            clean_value = re.sub(r'[^\d+]', '', clean_value)
            
            # Remove country codes
            if clean_value.startswith('+91'):
                clean_value = clean_value[3:]
            elif clean_value.startswith('91') and len(clean_value) > 10:
                clean_value = clean_value[2:]
            elif clean_value.startswith('0') and len(clean_value) == 11:
                clean_value = clean_value[1:]
            
            logger.info(f"Cleaned {field_name}: '{clean_value}' (length: {len(clean_value)})")
            
            # Validate phone number
            is_valid = False
            
            if len(clean_value) == 10 and clean_value.isdigit():
                # Valid 10-digit number
                if clean_value[0] in '6789':
                    is_valid = True
                else:
                    logger.warning(f"Phone number doesn't start with 6-9: '{clean_value}'")
                    # Allow it but log a warning
                    is_valid = True
            elif len(clean_value) == 9 and clean_value.isdigit():
                # 9-digit number - likely incomplete, allow but warn
                logger.warning(f"Incomplete phone number (9 digits): '{clean_value}' - allowing but should be completed")
                is_valid = True
            elif len(clean_value) < 9:
                # Too short - invalid
                is_valid = False
            else:
                # Too long or invalid format
                is_valid = False
            
            if not is_valid:
                logger.warning(f"Invalid {field_name}: '{value}' -> cleaned: '{clean_value}' (length: {len(clean_value)})")
                messages.error(request, f'{field_name.replace("_", " ").title()} must be a valid Indian mobile number (10 digits starting with 6, 7, 8, or 9). Received: "{value}" (cleaned: "{clean_value}")')
                return render(request, 'update_school.html', get_context(request))
            
            logger.info(f"Valid {field_name}: '{clean_value}'")
            
            # Store the clean value for database
            if field_name == 'phone':
                phone = clean_value
            elif field_name == 'principal_contact_phone':
                principal_contact_phone = clean_value
            elif field_name == 'director_contact_phone':
                director_contact_phone = clean_value
    
    # Validate emails
    email_regex = r'^[^\s@]+@[^\s@]+\.[^\s@]+$'
    email_fields = [
        ('email', email),
        ('principal_contact_mail', principal_contact_mail),
        ('director_contact_email', director_contact_email)
    ]
    for field_name, value in email_fields:
        if value and not re.match(email_regex, value):
            logger.warning(f"Invalid {field_name}: {value}")
            messages.error(request, f'{field_name.replace("_", " ").title()} must be a valid email address.')
            return render(request, 'update_school.html', get_context(request))
    
    # Validate pincode
    if pincode and not (pincode.isdigit() and len(pincode) == 6):
        messages.error(request, 'Pincode must be exactly 6 digits.')
        return render(request, 'update_school.html', get_context(request))
    
    # Handle school logo upload
    school_logo = None
    if 'school_logo' in request.FILES:
        logo_file = request.FILES['school_logo']
        from .utils import validate_uploaded_file
        is_valid, error_msg = validate_uploaded_file(logo_file)
        if not is_valid:
            return JsonResponse({'status': 'ERROR', 'message': f'Logo validation failed: {error_msg}'}, status=400)
        is_valid, message = validate_uploaded_file(logo_file, ALLOWED_IMAGE_TYPES)
        if not is_valid:
            messages.error(request, message)
            return render(request, 'update_school.html', get_context(request))
        school_logo = logo_file.read()
    
    # Convert establish_date to date object
    establish_date_obj = None
    if establish_date:
        try:
            establish_date_obj = safe_strptime(establish_date, '%Y-%m-%d').date()
        except ValueError:
            messages.error(request, 'Invalid establish date format.')
            return render(request, 'update_school.html', get_context(request))
    
    # Call stored procedure to update school
    try:
        logger.info(f"Calling stored procedure Proc_UpdateSchool_Set for school {school_id}")
        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT * FROM "Proc_UpdateSchool_Set"(
                    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s
                )
                """,
                [
                    school_id,
                    school_name,
                    registration_number or None,
                    address or None,
                    district,
                    state,
                    country,
                    pincode or None,
                    phone or None,
                    email or None,
                    website or None,
                    user_id,
                    school_logo,
                    board_id,
                    medium_id,
                    principal_name or None,
                    principal_contact_mail or None,
                    principal_contact_phone or None,
                    director_name or None,
                    director_contact_phone or None,
                    director_contact_email or None,
                    establish_date_obj,
                ]
            )
            result = cursor.fetchone()
            status = result[0] if result else 'Error'
            error_message = result[1] if result else 'Unknown error occurred'
            
            logger.info(f"Stored procedure result: Status={status}, ErrorMessage={error_message}")
            
            if status == 'Success':
                logger.info(f"School {school_id} update successful, clearing messages and setting success message")
                # Clear any existing messages
                list(messages.get_messages(request))
                messages.success(request, 'School updated successfully!')
                logger.info(f"Success message set, redirecting to schools list")
                # Redirect with unique parameter to break back button history
                # Create redirect URL with parameters
                redirect_url = reverse('schools_list') + f'?updated={school_id}&t={int(time.time())}'
                response = HttpResponseRedirect(redirect_url)
                
                # Add cache control headers to prevent back button issues
                response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
                response['Pragma'] = 'no-cache'
                response['Expires'] = '0'
                return response
            else:
                messages.error(request, f'Error updating school: {error_message}')
                context = get_context(request)
                context['school_id'] = school_id
                # Fetch school data for error display
                try:
                    with connection.cursor() as cursor:
                        cursor.execute('SELECT * FROM "Proc_GetSchoolDetails_ByID"(%s)', [school_id])
                        columns = [col[0] for col in cursor.description]
                        rows = cursor.fetchall()
                        if rows:
                            school_data = dict(zip(columns, rows[0]))
                            if school_data.get('SchoolLogo'):
                                school_data['SchoolLogo'] = f"data:image/jpeg;base64,{base64.b64encode(school_data['SchoolLogo']).decode('utf-8')}"
                            context['school_data'] = school_data
                except Exception as e:
                    logger.error(f"Error fetching school data for error display: {e}")
                return render(request, 'update_school.html', context)
                
    except Exception as e:
        logger.error(f"Error updating school: {e}")
        # Only show error message if we haven't already shown a success message
        if not any(msg.tags == 'success' for msg in messages.get_messages(request)):
            messages.error(request, 'An error occurred while updating the school.')
            context = get_context(request)
            context['school_id'] = school_id
            # Fetch school data for error display
            try:
                with connection.cursor() as cursor:
                    cursor.execute('SELECT * FROM "Proc_GetSchoolDetails_ByID"(%s)', [school_id])
                    columns = [col[0] for col in cursor.description]
                    rows = cursor.fetchall()
                    if rows:
                        school_data = dict(zip(columns, rows[0]))
                        if school_data.get('SchoolLogo'):
                            school_data['SchoolLogo'] = f"data:image/jpeg;base64,{base64.b64encode(school_data['SchoolLogo']).decode('utf-8')}"
                        context['school_data'] = school_data
            except Exception as fetch_error:
                logger.error(f"Error fetching school data for error display: {fetch_error}")
            return render(request, 'update_school.html', context)

@custom_login_required
def load_more_schools(request):
    user_id = request.session.get('UserId')
    if not user_id:
        return JsonResponse({'success': False, 'message': 'Session expired'}, status=401)

    page_number = safe_int(request.GET.get('page', 1))
    page_size = safe_int(request.GET.get('per_page', 10))

    # Search and filter parameters
    search_params = {
        'school_code': request.GET.get('school_code', ''),
        'school_name': request.GET.get('school_name', ''),
        'registration_number': request.GET.get('registration_number', ''),
        'created_at': request.GET.get('created_at', ''),
        'to_date': request.GET.get('to_date', ''),
        'phone': request.GET.get('phone', ''),
        'email': request.GET.get('email', ''),
        'principal_name': request.GET.get('principal_name', ''),
        'director_name': request.GET.get('director_name', ''),
        'board': request.GET.get('board', ''),
        'medium': request.GET.get('medium', ''),
        'status': request.GET.get('status', ''),
        'country': request.GET.get('country', ''),
        'state': request.GET.get('state', ''),
        'district': request.GET.get('district', ''),
        'pincode': request.GET.get('pincode', ''),
    }

    # Convert string IDs to integers
    try:
        board = int(search_params['board']) if search_params['board'] else None
        medium = int(search_params['medium']) if search_params['medium'] else None
        country = int(search_params['country']) if search_params['country'] else None
        state = int(search_params['state']) if search_params['state'] else None
        district = int(search_params['district']) if search_params['district'] else None
    except ValueError:
        board = medium = country = state = district = None

    # Convert date strings
    from_date = None
    to_date = None
    try:
        if search_params['created_at']:
            from_date = safe_strptime(search_params['created_at'], '%Y-%m-%d').date()
        if search_params['to_date']:
            to_date = safe_strptime(search_params['to_date'], '%Y-%m-%d').date()
    except ValueError:
        search_params['created_at'] = ''
        search_params['to_date'] = ''

    sort_column = request.GET.get('sort_column', 'SchoolID')
    sort_direction = request.GET.get('sort_direction', 'DESC')

    schools = []
    total_schools = 0
    active_schools = 0
    try:
        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT * FROM "Proc_SchoolList_get"(
                    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                    %s, %s, %s, %s, %s, %s, FALSE, %s, %s, %s, %s, %s
                )
                """,
                [
                    from_date,
                    to_date,
                    search_params['school_code'] or None,
                    search_params['school_name'] or None,
                    country,
                    state,
                    district,
                    search_params['phone'] or None,
                    search_params['email'] or None,
                    board,
                    medium,
                    search_params['pincode'] or None,
                    search_params['registration_number'] or None,
                    search_params['principal_name'] or None,
                    search_params['director_name'] or None,
                    search_params['status'] or None,
                    sort_column,
                    sort_direction,
                    page_number,
                    page_size,
                    user_id
                ]
            )
            columns = [col[0] for col in cursor.description]
            rows = cursor.fetchall()
            schools = [dict(zip(columns, row)) for row in rows]
            if schools:
                total_schools = schools[0].get('TotalCount', 0)
                active_schools = schools[0].get('ActiveSchools', 0)
                deleted_schools = schools[0].get('DeletedSchools', 0)

        # Convert SchoolLogo to base64
        for school in schools:
            if school.get('SchoolLogo'):
                school['SchoolLogo'] = f"data:image/jpeg;base64,{base64.b64encode(school['SchoolLogo']).decode('utf-8')}"
    except Exception as e:
        logger.error(f"Error fetching schools: {e}")
        return JsonResponse({'success': False, 'message': 'Error fetching schools'}, status=500)

    # Prepare JSON response
    response_data = {
        'success': True,
        'schools': schools,
        'total_count': total_schools,
        'active_schools': active_schools,
        'start_index': (page_number - 1) * page_size + 1,
        'end_index': min(page_number * page_size, total_schools),
        'has_next': page_number * page_size < total_schools,
        'has_previous': page_number > 1
    }

    return JsonResponse(response_data)

@custom_login_required
def export_schools(request):
    user_id = request.session.get('UserId')
    if not user_id:
        messages.error(request, "Session expired. Please login again.")
        return redirect('login')

    # Search and filter parameters
    search_params = {
        'school_code': request.GET.get('school_code', ''),
        'school_name': request.GET.get('school_name', ''),
        'registration_number': request.GET.get('registration_number', ''),
        'created_at': request.GET.get('created_at', ''),
        'to_date': request.GET.get('to_date', ''),
        'phone': request.GET.get('phone', ''),
        'email': request.GET.get('email', ''),
        'principal_name': request.GET.get('principal_name', ''),
        'director_name': request.GET.get('director_name', ''),
        'board': request.GET.get('board', ''),
        'medium': request.GET.get('medium', ''),
        'status': request.GET.get('status', ''),
        'country': request.GET.get('country', ''),
        'state': request.GET.get('state', ''),
        'district': request.GET.get('district', ''),
        'pincode': request.GET.get('pincode', ''),
    }

    # Convert string IDs to integers
    try:
        board = int(search_params['board']) if search_params['board'] else None
        medium = int(search_params['medium']) if search_params['medium'] else None
        country = int(search_params['country']) if search_params['country'] else None
        state = int(search_params['state']) if search_params['state'] else None
        district = int(search_params['district']) if search_params['district'] else None
    except ValueError:
        board = medium = country = state = district = None

    # Convert date strings
    from_date = None
    to_date = None
    try:
        if search_params['created_at']:
            from_date = safe_strptime(search_params['created_at'], '%Y-%m-%d').date()
        if search_params['to_date']:
            to_date = safe_strptime(search_params['to_date'], '%Y-%m-%d').date()
    except ValueError:
        search_params['created_at'] = ''
        search_params['to_date'] = ''

    schools = []
    try:
        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT * FROM "Proc_SchoolList_get"(
                    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                    %s, %s, %s, %s, %s, %s, FALSE, %s, %s, %s, %s, %s
                )
                """,
                [
                    from_date,
                    to_date,
                    search_params['school_code'] or None,
                    search_params['school_name'] or None,
                    country,
                    state,
                    district,
                    search_params['phone'] or None,
                    search_params['email'] or None,
                    board,
                    medium,
                    search_params['pincode'] or None,
                    search_params['registration_number'] or None,
                    search_params['principal_name'] or None,
                    search_params['director_name'] or None,
                    search_params['status'] or None,
                    'SchoolID',
                    'ASC',
                    1,
                    10000,  # Fetch all records for export
                    user_id
                ]
            )
            columns = [col[0] for col in cursor.description]
            rows = cursor.fetchall()
            schools = [dict(zip(columns, row)) for row in rows]
    except Exception as e:
        logger.error(f"Error fetching schools for export: {e}")
        messages.error(request, "An error occurred while exporting schools.")
        return redirect('schools_list')

    # Create CSV response
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="schools_export.csv"'

    writer = csv.writer(response)
    # Write header
    headers = [
        'SchoolID', 'SchoolCode', 'SchoolName', 'RegistrationNumber', 'Address', 
        'DistrictName', 'StateName', 'CountryName', 'Pincode', 'Phone', 'Email', 
        'Website', 'BoardName', 'MediumName', 'PrincipalName', 'PrincipalContactMail', 
        'PrincipalContactPhone', 'DirectorName', 'DirectorContactPhone', 
        'DirectorContactEmail', 'EstablishDate', 'CreatedAt', 'STATUS'
    ]
    writer.writerow(headers)

    # Write data
    for school in schools:
        writer.writerow([
            school.get('SchoolID', ''),
            school.get('SchoolCode', ''),
            school.get('SchoolName', ''),
            school.get('RegistrationNumber', 'N/A'),
            school.get('Address', 'N/A'),
            school.get('DistrictName', 'N/A'),
            school.get('StateName', 'N/A'),
            school.get('CountryName', 'N/A'),
            school.get('Pincode', 'N/A'),
            school.get('Phone', 'N/A'),
            school.get('Email', 'N/A'),
            school.get('Website', 'N/A'),
            school.get('BoardName', 'N/A'),
            school.get('MediumName', 'N/A'),
            school.get('PrincipalName', 'N/A'),
            school.get('PrincipalContactMail', 'N/A'),
            school.get('PrincipalContactPhone', 'N/A'),
            school.get('DirectorName', 'N/A'),
            school.get('DirectorContactPhone', 'N/A'),
            school.get('DirectorContactEmail', 'N/A'),
            school.get('EstablishDate', 'N/A') if not school.get('EstablishDate') else school['EstablishDate'].strftime('%Y-%m-%d'),
            school.get('CreatedAt', 'N/A') if not school.get('CreatedAt') else school['CreatedAt'].strftime('%Y-%m-%d'),
            school.get('STATUS', 'N/A')
        ])

    return response

@custom_login_required
def school_soft_delete(request, school_id):
    """Soft delete a school"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method'})
    
    try:
        user_id = request.session.get('UserId')
        if not user_id:
            return JsonResponse({'success': False, 'message': 'Session expired'})
            
        with connection.cursor() as cursor:
            cursor.execute(
                'SELECT * FROM "Proc_School_DeleteRestore"(%s, %s, %s)',
                [school_id, 'DELETE', user_id]
            )
            result = cursor.fetchone()
            
            if result and result[0].get('success'):
                return JsonResponse({'success': True, 'message': result[0].get('message')})
            else:
                return JsonResponse({'success': False, 'message': result[0].get('message') if result else 'Operation failed'})
                
    except Exception as e:
        logger.error(f"Error deleting school {school_id}: {e}")
        return JsonResponse({'success': False, 'message': str(e)})

@custom_login_required
def school_restore(request, school_id):
    """Restore a school"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method'})
        
    try:
        user_id = request.session.get('UserId')
        if not user_id:
            return JsonResponse({'success': False, 'message': 'Session expired'})
            
        with connection.cursor() as cursor:
            cursor.execute(
                'SELECT * FROM "Proc_School_DeleteRestore"(%s, %s, %s)',
                [school_id, 'RESTORE', user_id]
            )
            result = cursor.fetchone()
            
            if result and result[0].get('success'):
                return JsonResponse({'success': True, 'message': result[0].get('message')})
            else:
                return JsonResponse({'success': False, 'message': result[0].get('message') if result else 'Operation failed'})
                
    except Exception as e:
        logger.error(f"Error restoring school {school_id}: {e}")
        return JsonResponse({'success': False, 'message': str(e)})

#--Dark Mode Toggle View

@custom_login_required
def school_stats(request, school_id):
    """
    AJAX view to fetch statistics for a specific school.
    Returns counts of Active and Inactive users grouped by ProfileName (e.g., Student, Employee).
    Only accessible by Super Admins.
    """
    if request.method != 'GET':
        return JsonResponse({'success': False, 'message': 'Invalid request method'})

    # Check for Super Admin access
    if request.session.get('ProfileID') != 1:
        return JsonResponse({'success': False, 'message': 'Unauthorized access'}, status=403)

    try:
        data = {
            'school_name': '',
            'stats': []
        }
        
        with connection.cursor() as cursor:
            # Fetch School Name
            cursor.execute('SELECT "SchoolName" FROM "SchoolMaster" WHERE "SchoolID" = %s', [school_id])
            result = cursor.fetchone()
            data['school_name'] = result[0] if result else "Unknown School"
            
            # Fetch Active Users grouped by Profile
            cursor.execute('''
                SELECT
                    pm."ProfileName",
                    COUNT(um."UserID") AS total_users
                FROM public."UserMaster" um
                JOIN public."ProfileMaster" pm
                    ON pm."ProfileID" = um."ProfileID"
                WHERE um."SchoolID" = %s
                AND um."IsDeleted" = FALSE
                GROUP BY pm."ProfileName"
                ORDER BY pm."ProfileName"
            ''', [school_id])
            active_counts = {row[0]: row[1] for row in cursor.fetchall()}
            
            # Fetch Inactive Users grouped by Profile
            cursor.execute('''
                SELECT
                    pm."ProfileName",
                    COUNT(um."UserID") AS total_users
                FROM public."UserMaster" um
                JOIN public."ProfileMaster" pm
                    ON pm."ProfileID" = um."ProfileID"
                WHERE um."SchoolID" = %s
                AND um."IsDeleted" = TRUE
                GROUP BY pm."ProfileName"
                ORDER BY pm."ProfileName"
            ''', [school_id])
            inactive_counts = {row[0]: row[1] for row in cursor.fetchall()}
            
            # Merge counts
            all_profiles = set(active_counts.keys()) | set(inactive_counts.keys())
            
            # If no data, ensure we at least show standard categories if needed, 
            # but dynamic is better. If empty, it's fine.
            
            for profile in sorted(all_profiles):
                data['stats'].append({
                    'profile_name': profile,
                    'active': active_counts.get(profile, 0),
                    'inactive': inactive_counts.get(profile, 0),
                    'total': active_counts.get(profile, 0) + inactive_counts.get(profile, 0)
                })

        return JsonResponse({'success': True, 'data': data})

    except Exception as e:
        logger.error(f"Error fetching school stats for ID {school_id}: {str(e)}")
        return JsonResponse({'success': False, 'message': 'Error fetching statistics'})
import csv
import logging
import base64
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
from django.shortcuts import render, redirect
from django.contrib import messages
from django.db import connection
from django.http import JsonResponse, HttpResponse
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.contrib.auth.hashers import make_password
from django.db.models import F, Q
from .url_encryption import encrypt_id, decrypt_id_int
from mail.utils import send_email_by_code

from .models import UserMaster, ProfileMaster, SchoolMaster
from .utils import (
    get_context, safe_int, safe_strptime, 
    validate_uploaded_file, ALLOWED_IMAGE_TYPES, 
    bytes_to_data_uri, ERP_DEFAULT_LOGO_STATIC
)
from .decorators import custom_login_required

logger = logging.getLogger(__name__)

@custom_login_required
def user_list(request):
    # Get user context for header
    
    # Get session data for filtering
    user_id = request.session.get('UserId')
    profile_id = request.session.get('ProfileID')
    school_id = request.session.get('SchoolID')

    # Get pagination and search parameters
    page_size = safe_int(request.GET.get('per_page', 10))
    search_term = request.GET.get('search_name', '')
    user_code = request.GET.get('user_code', '')
    email = request.GET.get('email', '')
    phone = request.GET.get('phone', '')
    profile = request.GET.get('profile', '')
    school = request.GET.get('school', '')
    status = request.GET.get('status', '')
    order_by = request.GET.get('order_by', 'UserCode')
    order_direction = request.GET.get('order_direction', 'ASC')
    
    # Get date filter parameters
    from_date = request.GET.get('from_date', '')
    to_date = request.GET.get('to_date', '')
    
    # Convert date strings to datetime objects if provided
    from_date_obj = None
    to_date_obj = None
    if from_date:
        try:
            from_date_obj = safe_strptime(from_date, '%Y-%m-%d')
        except ValueError:
            from_date = ''
    if to_date:
        try:
            to_date_obj = safe_strptime(to_date, '%Y-%m-%d')
        except ValueError:
            to_date = ''

    # Fetch profiles and schools for filter dropdowns using ORM
    logged_in_profile_id = request.session.get('ProfileID')
    logged_in_school_id = request.session.get('SchoolID')
    
    # Fetch Profile Name
    try:
        current_profile_name = ProfileMaster.objects.get(profile_id=logged_in_profile_id).profile_name
    except ProfileMaster.DoesNotExist:
        current_profile_name = "Unknown"

    # Fetch profiles and schools for filter dropdowns logic
    if current_profile_name == 'Super Admin': # Super Admin
        profiles = list(ProfileMaster.objects.all().order_by('profile_name').values(ProfileID=F('profile_id'), ProfileName=F('profile_name')))
        schools = list(SchoolMaster.objects.filter(is_deleted=False).order_by('school_name').values(SchoolID=F('school_id'), SchoolName=F('school_name')))
    else:
        # Non-Super Admins: Limit Profiles (Exclude Super Admin and Support Executive)
        profiles = list(ProfileMaster.objects.filter(~Q(profile_name__in=['Super Admin', 'Support Executive'])).order_by('profile_name').values(ProfileID=F('profile_id'), ProfileName=F('profile_name')))
        
        # Non-Super Admins: Limit Schools to their own school
        schools = list(SchoolMaster.objects.filter(school_id=logged_in_school_id, is_deleted=False).order_by('school_name').values(SchoolID=F('school_id'), SchoolName=F('school_name')))

    # Check if this is an AJAX request
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    page_number = request.GET.get('page', 1)

    # Fetch users using stored procedure
    users = []
    total_count = 0
    try:
        with connection.cursor() as cursor:
            # PostgreSQL function call uses SELECT * FROM "Proc_UserList_Get"
            cursor.execute("""
                SELECT * FROM "Proc_UserList_Get"(
                    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s
                )
            """, [
                1,  # PageNumber (Fetch all initially for Django Paginator)
                10000, # PageSize
                user_id,
                safe_int(profile) if profile else (profile_id if (current_profile_name not in ['Super Admin', 'School Admin']) else None),
                safe_int(school) if school else school_id,
                search_term,
                from_date_obj,
                to_date_obj,
                status,
                order_by,
                order_direction
            ])
            
            columns = [col[0] for col in cursor.description]
            raw_users = cursor.fetchall()
            
            for row in raw_users:
                user = dict(zip(columns, row))
                if user.get('UserPhoto'):
                    try:
                        user['UserPhotoBase64'] = base64.b64encode(user['UserPhoto']).decode('utf-8')
                    except Exception as e:
                        logger.error(f"Error encoding user photo: {e}")
                        user['UserPhotoBase64'] = None
                else:
                    user['UserPhotoBase64'] = None
                    
                if user.get('CreatedAt'):
                    user['CreatedAtFormatted'] = user['CreatedAt'].strftime('%Y-%m-%d %H:%M')
                
                # Ensure IsDeleted is int and Status is set
                user['IsDeleted'] = int(user.get('IsDeleted', 0))
                user['Status'] = 'Inactive' if user['IsDeleted'] == 1 else 'Active'
                
                # Encrypt the UserID for safe URL usage
                if user.get('UserID'):
                    user['EncryptedUserID'] = encrypt_id(user['UserID'])
                
                users.append(user)
            
            total_count = users[0].get('TotalCount', 0) if users else 0

    except Exception as e:
        logger.error(f"Error fetching user list via procedure: {e}")
        if not is_ajax:
            messages.error(request, "Error loading user data. Please try again.")

    # Use Django Paginator
    paginator = Paginator(users, page_size)
    try:
        users_page = paginator.page(page_number)
    except PageNotAnInteger:
        users_page = paginator.page(1)
    except EmptyPage:
        users_page = paginator.page(paginator.num_pages)

    # Calculate active users
    active_users = sum(1 for user in users if user.get('Status') == 'Active')

    # If it's an AJAX request, return JSON
    if is_ajax:
        serializable_users = []
        for user in users_page:
            user_copy = user.copy()
            if 'CreatedAt' in user_copy and user_copy['CreatedAt']:
                user_copy['CreatedAt'] = user_copy['CreatedAt'].isoformat()
            if 'UserPhoto' in user_copy:
                del user_copy['UserPhoto']
            serializable_users.append(user_copy)
        return JsonResponse({
            'users': serializable_users,
            'total_count': total_count,
            'start_index': users_page.start_index(),
            'end_index': users_page.end_index(),
            'has_next': users_page.has_next()
        })

    # Update context with user list specific data
    context = {
        'users': users_page,
        'per_page': page_size,
        'search_name': search_term,
        'user_code': user_code,
        'email': email,
        'phone': phone,
        'profile_id': profile,
        'school_id': school,
        'status': status,
        'from_date': from_date,
        'to_date': to_date,
        'order_by': order_by,
        'order_direction': order_direction,
        'total_users': total_count,
        'active_users': active_users,
        'start_index': users_page.start_index(),
        'end_index': users_page.end_index(),
        'profiles': profiles,
        'schools': schools,
        'current_profile_name': current_profile_name,
    }
    
    # Merge global context (header, theme, etc.)
    context.update(get_context(request))
    
    return render(request, 'user_list.html', context)

@custom_login_required
def load_more_users(request):
    if request.method == 'GET' and request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        # Get pagination and filter parameters
        page = safe_int(request.GET.get('page', 1))
        page_size = safe_int(request.GET.get('per_page', 10))
        search_name = request.GET.get('search_name', '')
        user_code = request.GET.get('user_code', '')
        user_name = request.GET.get('user_name', '')
        email = request.GET.get('email', '')
        phone = request.GET.get('phone', '')
        profile = request.GET.get('profile', '')
        school = request.GET.get('school', '')
        from_date = request.GET.get('from_date', '')
        to_date = request.GET.get('to_date', '')
        status = request.GET.get('status', '')
        order_by = request.GET.get('order_by', 'UserCode')
        order_direction = request.GET.get('order_direction', 'ASC')
        
        # Get session information
        user_id = request.session.get('UserId')
        profile_id = request.session.get('ProfileID')
        school_id = request.session.get('SchoolID')
        
        # Fetch Profile Name for logic
        try:
            current_profile_name = ProfileMaster.objects.get(profile_id=profile_id).profile_name
        except ProfileMaster.DoesNotExist:
            current_profile_name = "Unknown"
        
        # Convert date strings to datetime objects if provided
        from_date_obj = None
        to_date_obj = None
        if from_date:
            try:
                from_date_obj = safe_strptime(from_date, '%Y-%m-%d')
            except ValueError:
                from_date = ''
        if to_date:
            try:
                to_date_obj = safe_strptime(to_date, '%Y-%m-%d')
            except ValueError:
                to_date = ''
        
        try:
            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT * FROM "Proc_UserList_Get"(
                        %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s
                    )
                """, [
                    page, 
                    page_size, 
                    user_id, 
                    safe_int(profile) if profile else (profile_id if (current_profile_name not in ['Super Admin', 'School Admin']) else None),
                    safe_int(school) if school else school_id,
                    user_name or search_name,
                    from_date_obj,
                    to_date_obj,
                    status,
                    order_by,
                    order_direction
                ])
                
                columns = [col[0] for col in cursor.description]
                rows = cursor.fetchall()
                
                users = []
                for row in rows:
                    user_data = dict(zip(columns, row))
                    if user_data.get('UserPhoto'):
                        try:
                            user_data['UserPhotoBase64'] = base64.b64encode(user_data['UserPhoto']).decode('utf-8')
                        except Exception as e:
                            logger.error(f"Error encoding user photo: {e}")
                            user_data['UserPhotoBase64'] = None
                    else:
                        user_data['UserPhotoBase64'] = None
                        
                    if user_data.get('CreatedAt'):
                        user_data['CreatedAtFormatted'] = user_data['CreatedAt'].strftime('%Y-%m-%d %H:%M')
                    
                    # Ensure formatting for JSON
                    user_data['IsDeleted'] = int(user_data.get('IsDeleted', 0))
                    user_data['Status'] = 'Inactive' if user_data['IsDeleted'] == 1 else 'Active'
                    
                    # Encrypt UserID for safe URL generation in JS
                    if user_data.get('UserID'):
                        user_data['EncryptedUserID'] = encrypt_id(user_data['UserID'])
                    
                    if 'UserPhoto' in user_data:
                        del user_data['UserPhoto']
                    if 'CreatedAt' in user_data and user_data['CreatedAt']:
                        user_data['CreatedAt'] = user_data['CreatedAt'].isoformat()
                        
                    users.append(user_data)
                
                total_count = users[0]['TotalCount'] if users and 'TotalCount' in users[0] else 0
                start_index = (page - 1) * page_size + 1
                end_index = min(page * page_size, total_count)
                has_next = end_index < total_count
                
                # Convert datetime objects for JSON
                serializable_users = []
                for user in users:
                    user_copy = user.copy()
                    if 'CreatedAt' in user_copy:
                        user_copy['CreatedAt'] = user_copy['CreatedAt'].isoformat() if user_copy['CreatedAt'] else None
                    serializable_users.append(user_copy)
                
                return JsonResponse({
                    'users': serializable_users,
                    'total_count': total_count,
                    'start_index': start_index,
                    'end_index': end_index,
                    'has_next': has_next
                })
                
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    
    return JsonResponse({'error': 'Invalid request'}, status=400)

@custom_login_required
def create_user(request):
    # Restrict User Creation to Super Admin only
    logged_in_profile_id = request.session.get('ProfileID')
    try:
        current_profile_name = ProfileMaster.objects.get(profile_id=logged_in_profile_id).profile_name
    except ProfileMaster.DoesNotExist:
        current_profile_name = "Unknown"

    if current_profile_name != 'Super Admin':
        messages.error(request, "Access Denied: Only Super Admins can create new users.")
        return redirect('user_list')

    context = {}
    profiles = []
    schools = []
    
    if request.method == 'POST':
        # Extract form data
        user_name = request.POST.get('user_name')
        password = request.POST.get('password')
        email = request.POST.get('email')
        phone = request.POST.get('phone')
        profile_id = request.POST.get('profile')
        school_id = request.POST.get('school') or None
        
        user_photo = None
        if 'user_photo' in request.FILES:
            file = request.FILES['user_photo']
            from .utils import validate_uploaded_file
            is_valid, error_msg = validate_uploaded_file(file)
            if not is_valid:
                return JsonResponse({'status': 'ERROR', 'message': f'Profile photo validation failed: {error_msg}'}, status=400)
            is_valid, message = validate_uploaded_file(file, ALLOWED_IMAGE_TYPES)
            if not is_valid:
                return render(request, 'create_user.html', context)
            user_photo = file.read()

        # Validate required fields
        if not all([user_name, password, profile_id]):
            return render(request, 'create_user.html', context)

        # School selection rule - Super Admin can create users without school
        if str(profile_id) == "1":  # Creating a new Super Admin user
            school_id = None
        else:
            # Super Admin can create users with or without school association
            school_id = school_id if school_id else None

        # Get logged in UserId from session
        created_by = request.session.get('UserId')
        
        if not created_by:
            messages.error(request, "Session expired. Please login again.")
            return redirect('login')

        # Insert into UserMaster using new PostgreSQL function
        try:
            with connection.cursor() as cursor:
                cursor.execute(
                    """
                    SELECT "Status", "Message", "NewUserID" 
                    FROM "Proc_User_Manage"(
                        p_Action := 'INSERT',
                        p_UserName := %s,
                        p_PasswordHash := %s,
                        p_Email := %s,
                        p_Phone := %s,
                        p_ProfileID := %s,
                        p_SchoolID := %s,
                        p_UserPhoto := %s,
                        p_CreatedBy := %s
                    )
                    """,
                    [
                        user_name,
                        make_password(password),
                        email or None,
                        phone or None,
                        profile_id,
                        school_id,
                        user_photo,
                        created_by
                    ]
                )
                res = cursor.fetchone()
                status = res[0] if res else 'FAILED'
                message = res[1] if res else 'Unknown error'
                
                if status != 'SUCCESS':
                    messages.error(request, message)
                    return render(request, 'create_user.html', context)
        except Exception as e:
            logger.error(f"Error creating user: {e}")
            messages.error(request, f"Error: {e}")
            return render(request, 'create_user.html', context)

        messages.success(request, f"User '{user_name}' created successfully.")
        return redirect('dashboard')

    else:
        # Fetch ProfileMaster and SchoolMaster (Only Super Admin reaches here)
        profiles = list(ProfileMaster.objects.filter(is_deleted=False).order_by('profile_name').values(ProfileID=F('profile_id'), ProfileName=F('profile_name')))
        
        # Use get_school_dropdown for [Code] Name format
        from .utils import get_school_dropdown
        schools = get_school_dropdown()

        # Get logged in school id for context
        logged_in_school_id = request.session.get('SchoolID')

        # Update context with create user specific data
        context = {
            'profiles': profiles,
            'schools': schools,
            'user_school_id': logged_in_school_id,
        }
        return render(request, 'create_user.html', context)

@custom_login_required
def update_user(request, encrypted_id):
    # Decrypt the encrypted user ID
    user_id = decrypt_id_int(encrypted_id)
    if not user_id:
        messages.error(request, "Invalid or expired link. Please try again.")
        return redirect('user_list')
    
    context = {}
    profiles = []
    schools = []
    """Update user information"""
    # Get user context for header
    
    if request.method == 'POST':
        # Extract form data
        user_name = request.POST.get('user_name')
        email = request.POST.get('email')
        phone = request.POST.get('phone')
        profile_id = request.POST.get('profile')
        school_id = request.POST.get('school') or None
        new_password = request.POST.get('new_password')
        
        user_photo = None
        if 'user_photo' in request.FILES:
            file = request.FILES['user_photo']
            from .utils import validate_uploaded_file
            is_valid, error_msg = validate_uploaded_file(file)
            if not is_valid:
                return JsonResponse({'status': 'ERROR', 'message': f'Profile photo validation failed: {error_msg}'}, status=400)
            is_valid, message = validate_uploaded_file(file, ALLOWED_IMAGE_TYPES)
            if not is_valid:
                return render(request, 'update_user.html', context)
            user_photo = file.read()

        # Validate required fields
        if not all([user_name, profile_id]):
            return render(request, 'update_user.html', context)

        # Validate password if provided
        if new_password and new_password.strip():
            if len(new_password) < 8:
                return render(request, 'update_user.html', context)
            
            # Check if password contains at least one letter and one number
            import re
            if not re.search(r'[A-Za-z]', new_password) or not re.search(r'[0-9]', new_password):
                return render(request, 'update_user.html', context)

        # School selection rule
        if str(profile_id) == "1":  # Super Admin
            school_id = None
        else:
            if not school_id:
                return render(request, 'update_user.html', context)

        # Get logged in UserId from session
        modified_by = request.session.get('UserId')
        
        if not modified_by:
            messages.error(request, "Session expired. Please login again.")
            return redirect('login')

        # Call stored procedure to update user
        try:
            with connection.cursor() as cursor:
                # Hash the password if provided
                hashed_password = None
                if new_password and new_password.strip():
                    hashed_password = make_password(new_password)
                
                cursor.execute("""
                    SELECT "Status", "Message"
                    FROM "Proc_User_Manage"(
                        p_Action := 'UPDATE',
                        p_UserID := %s,
                        p_UserName := %s,
                        p_Email := %s,
                        p_Phone := %s,
                        p_ProfileID := %s,
                        p_SchoolID := %s,
                        p_UserPhoto := %s,
                        p_PasswordHash := %s,
                        p_ModifiedBy := %s
                    )
                """, [
                    user_id,
                    user_name,
                    email or None,
                    phone or None,
                    profile_id,
                    school_id,
                    user_photo,
                    hashed_password,
                    modified_by
                ])
                
                result = cursor.fetchone()
                status = result[0] if result else 'FAILED'
                error_message = result[1] if result else 'Unknown error occurred'
                
                if status == 'SUCCESS':
                    success_message = f"User '{user_name}' updated successfully."
                    if new_password and new_password.strip():
                        success_message += " Password has been updated."
                    messages.success(request, success_message)
                    from django.urls import reverse
                    redirect_url = reverse('user_list')
                    if school_id:
                        redirect_url += f'?school={school_id}'
                    return redirect(redirect_url)
                else:
                    messages.error(request, error_message)
                    return render(request, 'update_user.html', context)
                    
        except Exception as e:
            logger.error(f"Error updating user {user_id}: {e}")
            messages.error(request, f"Error: {e}")
            return render(request, 'update_user.html', context)

    else:
        # GET request - fetch user data and render form
        try:
            # Fetch user data using ORM
            try:
                user_obj = UserMaster.objects.get(user_id=user_id, is_deleted=False)
            except UserMaster.DoesNotExist:
                messages.error(request, "User not found.")
                return redirect('user_list')
            
            # Get logged in user info
            logged_in_user_id = request.session.get('UserId')
            logged_in_profile_id = request.session.get('ProfileID')
            logged_in_school_id = request.session.get('SchoolID')
            
            # Fetch current user's profile name
            try:
                current_profile = ProfileMaster.objects.get(profile_id=logged_in_profile_id)
                current_profile_name = current_profile.profile_name
            except ProfileMaster.DoesNotExist:
                current_profile_name = "Unknown"
            
            is_super_admin = current_profile_name == 'Super Admin'
            is_school_admin = current_profile_name == 'School Admin'
            is_own_profile = (logged_in_user_id == user_id)
            
            # === ACCESS CONTROL ===
            # Super Admin: can update anyone
            # School Admin: can update users in their school only
            # Others: can only update their own profile
            
            if is_super_admin:
                can_edit = True
            elif is_school_admin:
                # School Admin can only edit users in their school
                can_edit = (user_obj.school_id == logged_in_school_id)
            else:
                # Regular users can only edit their own profile
                can_edit = is_own_profile
            
            if not can_edit:
                messages.error(request, "You don't have permission to update this user.")
                return redirect('user_list')
            
            # Convert to dictionary for template compatibility
            user = {
                'UserID': user_obj.user_id,
                'UserCode': user_obj.user_code,
                'UserName': user_obj.user_name,
                'Email': user_obj.email,
                'Phone': user_obj.phone,
                'ProfileID': user_obj.profile_id,
                'SchoolID': user_obj.school_id,
            }
            
            # Process photo data
            if user_obj.user_photo:
                try:
                    user['UserPhotoBase64'] = base64.b64encode(user_obj.user_photo).decode('utf-8')
                except Exception as e:
                    logger.error(f"Error encoding user photo: {e}")
                    user['UserPhotoBase64'] = None
            else:
                user['UserPhotoBase64'] = None

            # === PROFILE AND SCHOOL DROPDOWN BASED ON ROLE ===
            from .utils import get_school_dropdown
            
            if is_super_admin:
                # Super Admin: can see all profiles and all schools
                profiles = list(ProfileMaster.objects.filter(is_deleted=False).order_by('profile_name').values(ProfileID=F('profile_id'), ProfileName=F('profile_name')))
                schools = get_school_dropdown()
                can_change_profile = True
                can_change_school = True
            elif is_school_admin:
                # School Admin: can see all profiles EXCEPT Super Admin, only their school
                profiles = list(ProfileMaster.objects.filter(is_deleted=False).exclude(profile_name='Super Admin').order_by('profile_name').values(ProfileID=F('profile_id'), ProfileName=F('profile_name')))
                all_schools = get_school_dropdown()
                schools = [s for s in all_schools if s['SchoolID'] == logged_in_school_id]
                can_change_profile = True
                can_change_school = False
            else:
                # Regular users: cannot change profile or school (editing own profile)
                profiles = []  # Empty - will be hidden in template
                all_schools = get_school_dropdown()
                schools = [s for s in all_schools if s['SchoolID'] == logged_in_school_id]
                can_change_profile = False
                can_change_school = False

            # Update context with update user specific data
            context = {
                'user': user,
                'profiles': profiles,
                'schools': schools,
                'user_school_id': logged_in_school_id,
                'can_change_profile': can_change_profile,
                'can_change_school': can_change_school,
                'is_super_admin': is_super_admin,
                'is_school_admin': is_school_admin,
                'is_own_profile': is_own_profile,
            }
            return render(request, 'update_user.html', context)
            
        except Exception as e:
            logger.error(f"Error fetching user data for update: {e}")
            messages.error(request, "Error loading user data. Please try again.")
            return redirect('user_list')

@custom_login_required
def user_soft_delete(request, user_id):
    """Soft delete a user using the PostgreSQL function"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method'})
    
    try:
        current_user_id = request.session.get('UserId')
        if not current_user_id:
            return JsonResponse({'success': False, 'message': 'Session expired. Please login again.'})
        
        # Ensure user_id is integer
        user_id = safe_int(user_id)
        
        with connection.cursor() as cursor:
            # Call PostgreSQL function
            cursor.execute(
                "SELECT * FROM \"Proc_User_DeleteRestore\"(%s, %s, %s)",
                [user_id, 'DELETE', current_user_id]
            )
            result = cursor.fetchone()
            
            if result and result[0].get('success'):
                return JsonResponse({'success': True, 'message': result[0].get('message', 'User deleted successfully')})
            else:
                msg = result[0].get('message') if result else 'Operation failed'
                return JsonResponse({'success': False, 'message': msg})
        
    except Exception as e:
        logger.error(f"Error deleting user {user_id}: {e}")
        return JsonResponse({'success': False, 'message': f'An error occurred: {str(e)}'})

@custom_login_required
def user_restore(request, user_id):
    """Restore a soft deleted user using the PostgreSQL function"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method'})
    
    try:
        current_user_id = request.session.get('UserId')
        if not current_user_id:
            return JsonResponse({'success': False, 'message': 'Session expired. Please login again.'})
        
        # Ensure user_id is integer
        user_id = safe_int(user_id)
        
        with connection.cursor() as cursor:
            # Call PostgreSQL function
            cursor.execute(
                "SELECT * FROM \"Proc_User_DeleteRestore\"(%s, %s, %s)",
                [user_id, 'RESTORE', current_user_id]
            )
            result = cursor.fetchone()
            
            if result and result[0].get('success'):
                return JsonResponse({'success': True, 'message': result[0].get('message', 'User restored successfully')})
            else:
                msg = result[0].get('message') if result else 'Operation failed'
                return JsonResponse({'success': False, 'message': msg})
        
    except Exception as e:
        logger.error(f"Error restoring user {user_id}: {e}")
        return JsonResponse({'success': False, 'message': f'An error occurred: {str(e)}'})

@custom_login_required
def get_user_password(request, user_id):
    """Retrieve user password (admin only)"""
    # This view is usually for super admins to see passwords in plain text if needed
    try:
        # Check if the requester is a super admin
        if request.session.get('ProfileID') != 1:
            return JsonResponse({'success': False, 'message': 'Unauthorized access'})

        with connection.cursor() as cursor:
            cursor.execute("SELECT \"PasswordHash\" FROM \"UserMaster\" WHERE \"UserID\" = %s", [user_id])
            row = cursor.fetchone()
            if row:
                # Note: In a real hashed system, we can't get plain text password
                # But if system stores plain text or reversible (not recommended)
                # Here we return the hash as it is a common request in this legacy migration
                return JsonResponse({'success': True, 'password': row[0]})
        return JsonResponse({'success': False, 'message': 'User not found'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})

def validate_user_api(request):
    """Validate if user exists in system for face authentication"""
    try:
        identifier = request.POST.get('identifier', '').strip()
        
        if not identifier:
            return JsonResponse({
                'success': False,
                'error': 'Email or username is required'
            })
        
        # Check if user exists using the same query pattern as login
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT "UserCode", "Email", "UserName", "IsActive"
                FROM "UserMaster" 
                WHERE ("UserCode" = %s OR "Email" = %s OR "UserName" = %s) 
                  AND COALESCE("IsDeleted", FALSE) = FALSE
            """, [identifier, identifier, identifier])
            
            user_row = cursor.fetchone()
            
            if user_row:
                user_code, email, user_name, is_active = user_row
                
                if not is_active:
                    return JsonResponse({
                        'success': False,
                        'error': 'User account is inactive. Please contact administrator.'
                    })
                
                return JsonResponse({
                    'success': True,
                    'message': f'User found: {user_name}',
                    'user_code': user_code,
                    'email': email
                })
            else:
                return JsonResponse({
                    'success': False,
                    'error': 'User not found. Please check your email or username.'
                })
                
    except Exception as e:
        logger.error(f"User validation error: {e}")
        return JsonResponse({
            'success': False,
            'error': 'System error. Please try again.'
        })

@custom_login_required
def users_export(request):
    """Export users as Excel or CSV"""
    format_type = request.GET.get('format', 'excel').lower()
    
    # Get session data for filtering
    user_id = request.session.get('UserId')
    profile_id = request.session.get('ProfileID')
    school_id = request.session.get('SchoolID')

    # Get filter parameters (same as user_list)
    search_term = request.GET.get('search_name', '')
    status = request.GET.get('status', '')
    profile = request.GET.get('profile', '')
    school = request.GET.get('school', '')
    order_by = request.GET.get('order_by', 'UserCode')
    order_direction = request.GET.get('order_direction', 'ASC')
    from_date = request.GET.get('from_date', '')
    to_date = request.GET.get('to_date', '')
    
    # Convert date strings to datetime objects
    from_date_obj = safe_strptime(from_date, '%Y-%m-%d') if from_date else None
    to_date_obj = safe_strptime(to_date, '%Y-%m-%d') if to_date else None
    
    # Fetch Profile Name for logic
    try:
        current_profile_name = ProfileMaster.objects.get(profile_id=profile_id).profile_name
    except ProfileMaster.DoesNotExist:
        current_profile_name = "Unknown"

    # Fetch users using stored procedure (Fetch a large number for export)
    users = []
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT * FROM "Proc_UserList_Get"(
                    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s
                )
            """, [
                1, 10000, user_id,
                safe_int(profile) if profile else (profile_id if (current_profile_name not in ['Super Admin', 'School Admin']) else None),
                safe_int(school) if school else school_id,
                search_term, from_date_obj, to_date_obj,
                status,
                order_by, order_direction
            ])
            
            columns = [col[0] for col in cursor.description]
            rows = cursor.fetchall()
            
            for row in rows:
                user = dict(zip(columns, row))
                # Clean up and format
                if user.get('CreatedAt'):
                    user['CreatedAtFormatted'] = user['CreatedAt'].strftime('%Y-%m-%d %H:%M')
                
                # Status based on IsDeleted
                user['Status'] = 'Inactive' if int(user.get('IsDeleted', 0)) == 1 else 'Active'
                
                # Remove internal/blob columns or columns that shouldn't be in Excel/CSV
                columns_to_remove = ['UserPhoto', 'TotalCount', 'IsDeleted', 'UserPhotoBase64', 'CreatedAt']
                for col in columns_to_remove:
                    if col in user: del user[col]
                
                users.append(user)

    except Exception as e:
        logger.error(f"Error fetching users for export: {e}")
        return HttpResponse("Error generating export file", status=500)

    filename = f"users_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}"

    if format_type == 'csv':
        delimiter_type = request.GET.get('delimiter', 'comma').lower()
        delimiters = {'comma': ',', 'pipe': '|', 'tab': '\t'}
        actual_delimiter = delimiters.get(delimiter_type, ',')
        
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
        
        if users:
            writer = csv.DictWriter(response, fieldnames=users[0].keys(), delimiter=actual_delimiter)
            writer.writeheader()
            writer.writerows(users)
        return response

    else:  # Excel
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Users"
        
        if users:
            headers = list(users[0].keys())
            # Header styling
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="004aad", end_color="004aad", fill_type="solid")
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            # Data rows
            for row_num, user in enumerate(users, 2):
                for col_num, (key, value) in enumerate(user.items(), 1):
                    ws.cell(row=row_num, column=col_num, value=str(value) if value is not None else "")
            
            # Auto-adjust column width
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except: pass
                ws.column_dimensions[column_letter].width = max_length + 2

        wb.save(response)
        return response

@custom_login_required
def get_themes_api(request):
    """API to fetch all active themes"""
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT "ThemeID", "ThemeName", "PrimaryColor", "PrimaryHover" 
                FROM "ThemeMaster" 
                WHERE "IsActive" = TRUE 
                ORDER BY "DisplayOrder" ASC
            """)
            themes = []
            for row in cursor.fetchall():
                themes.append({
                    'id': row[0],
                    'name': row[1],
                    'primary_color': row[2],
                    'primary_hover': row[3]
                })
            
            # Get current selected theme ID for the user
            user_id = request.session.get('UserId')
            current_theme_id = None
            if user_id:
                cursor.execute('SELECT "ThemeID" FROM "UserMaster" WHERE "UserID" = %s', [user_id])
                row = cursor.fetchone()
                if row:
                    current_theme_id = row[0]

            return JsonResponse({
                'success': True, 
                'themes': themes,
                'current_theme_id': current_theme_id
            })
    except Exception as e:
        logger.error(f"Error fetching themes: {e}")
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@custom_login_required
def update_theme_api(request):
    """API to update user and optionally school theme"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid method'}, status=405)
    
    try:
        data = json.loads(request.body)
        theme_id = data.get('theme_id')
        apply_to_school = data.get('apply_to_school', False)
        user_id = request.session.get('UserId')
        school_id = request.session.get('SchoolID')
        profile_id = request.session.get('ProfileID')

        if not theme_id:
            return JsonResponse({'success': False, 'error': 'Theme ID is required'}, status=400)

        with connection.cursor() as cursor:
            # 1. Update User preference
            cursor.execute('UPDATE "UserMaster" SET "ThemeID" = %s WHERE "UserID" = %s', [theme_id, user_id])
            
            # 2. If apply_to_school and user is Admin (Profile 1 or 2)
            if apply_to_school and profile_id in [1, 2] and school_id:
                cursor.execute('UPDATE "SchoolMaster" SET "ThemeID" = %s WHERE "SchoolID" = %s', [theme_id, school_id])
                message = "Theme updated personally and for the entire school."
            else:
                message = "Theme updated successfully."

        return JsonResponse({'success': True, 'message': message})
    except Exception as e:
        logger.error(f"Error updating theme: {e}")
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@custom_login_required
def blocked_users_list(request):
    """List users who are currently blocked or have failed login attempts"""
    user_id = request.session.get('UserId')
    profile_id = request.session.get('ProfileID')
    school_id = request.session.get('SchoolID')
    
    try:
        current_profile_name = ProfileMaster.objects.get(profile_id=profile_id).profile_name
    except ProfileMaster.DoesNotExist:
        current_profile_name = "Unknown"
        
    is_super_admin = current_profile_name == 'Super Admin'
    is_school_admin = current_profile_name == 'School Admin'
    
    if not (is_super_admin or is_school_admin):
        messages.error(request, "Access Denied: You don't have permission to view blocked users.")
        return redirect('dashboard')
        
    blocked_users = []
    try:
        with connection.cursor() as cursor:
            # Query users who have a block or failed attempts
            # Super Admin sees everyone. School Admin sees school-specific profiles.
            query = '''
                SELECT u."UserID", u."UserCode", u."UserName", u."Email", u."Phone", 
                       p."ProfileName", s."SchoolName", u."BlockedUntil", u."FailedLoginAttempts",
                       u."ProfileID", u."SchoolID"
                FROM "UserMaster" u
                JOIN "ProfileMaster" p ON u."ProfileID" = p."ProfileID"
                LEFT JOIN "SchoolMaster" s ON u."SchoolID" = s."SchoolID"
                WHERE (u."BlockedUntil" > NOW() OR u."FailedLoginAttempts" > 0)
                AND u."IsDeleted" = FALSE
            '''
            params = []
            
            if is_school_admin:
                # Filter by SchoolID and allowed profiles: Teacher(3), Student(4), Parent(5), Driver(6), Librarian(7), Accountant(8)
                # Note: User mentioned "Employee" - Teacher and Accountant are employees.
                query += ' AND u."SchoolID" = %s AND u."ProfileID" IN (3, 4, 5, 6, 7, 8)'
                params.append(school_id)
            
            query += ' ORDER BY u."BlockedUntil" DESC NULLS LAST, u."UserName" ASC'
            
            cursor.execute(query, params)
            columns = [col[0] for col in cursor.description]
            for row in cursor.fetchall():
                user_data = dict(zip(columns, row))
                # Encrypt UserID for URL safety
                user_data['EncryptedUserID'] = encrypt_id(user_data['UserID'])
                # Format block message
                from django.utils import timezone
                if user_data['BlockedUntil'] and user_data['BlockedUntil'] > timezone.now():
                    user_data['BlockStatus'] = 'Temporarily Blocked'
                else:
                    user_data['BlockStatus'] = 'Risky (Failed Attempts)'
                
                blocked_users.append(user_data)
                
    except Exception as e:
        logger.error(f"Error fetching blocked users: {e}")
        messages.error(request, "Error loading blocked users list.")

    context = {
        'blocked_users': blocked_users,
        'is_super_admin': is_super_admin,
        'is_school_admin': is_school_admin
    }
    context.update(get_context(request))
    return render(request, 'blocked_users.html', context)

@custom_login_required
def unblock_user(request, user_id):
    """Action to unblock a user immediately"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method'})
        
    logged_in_user_id = request.session.get('UserId')
    logged_in_profile_id = request.session.get('ProfileID')
    logged_in_school_id = request.session.get('SchoolID')
    
    try:
        # Check permission
        try:
            current_profile_name = ProfileMaster.objects.get(profile_id=logged_in_profile_id).profile_name
        except ProfileMaster.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Profile error'})

        if not (current_profile_name == 'Super Admin' or current_profile_name == 'School Admin'):
            return JsonResponse({'success': False, 'message': 'Unauthorized'})

        # Validation: School Admin cannot unblock self
        if int(user_id) == int(logged_in_user_id) and current_profile_name == 'School Admin':
             return JsonResponse({'success': False, 'message': 'Access Denied: You cannot unblock your own account.'})

        # Fetch target user to check school/profile constraints for School Admin
        with connection.cursor() as cursor:
            cursor.execute('SELECT "SchoolID", "ProfileID" FROM "UserMaster" WHERE "UserID" = %s', [user_id])
            target = cursor.fetchone()
            if not target:
                return JsonResponse({'success': False, 'message': 'User not found'})
            
            target_school_id, target_profile_id = target
            
            if current_profile_name == 'School Admin':
                if target_school_id != logged_in_school_id:
                    return JsonResponse({'success': False, 'message': 'Access Denied: Target user belongs to another school.'})
                if target_profile_id not in [3, 4, 5, 6, 7, 8]:
                    return JsonResponse({'success': False, 'message': 'Access Denied: You do not have permission to manage this profile type.'})

            # Perform unblock
            cursor.execute('''
                UPDATE "UserMaster" 
                SET "BlockedUntil" = NULL, 
                    "FailedLoginAttempts" = 0,
                    "IsActive" = TRUE
                WHERE "UserID" = %s
            ''', [user_id])
            connection.commit()
            
            return JsonResponse({'success': True, 'message': 'User has been unblocked successfully.'})
            
    except Exception as e:
        logger.error(f"Error unblocking user {user_id}: {e}")
        return JsonResponse({'success': False, 'message': str(e)})

@custom_login_required
def admin_password_reset(request):
    """Admin-initiated password reset without validation"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method'})
        
    user_id = request.POST.get('user_id')
    new_password = request.POST.get('new_password')
    
    if not user_id or not new_password:
        return JsonResponse({'success': False, 'message': 'User ID and new password are required.'})
        
    logged_in_profile_id = request.session.get('ProfileID')
    logged_in_school_id = request.session.get('SchoolID')
    
    try:
        # Check permission
        try:
            current_profile_name = ProfileMaster.objects.get(profile_id=logged_in_profile_id).profile_name
        except ProfileMaster.DoesNotExist:
             return JsonResponse({'success': False, 'message': 'Profile error'})
             
        if not (current_profile_name == 'Super Admin' or current_profile_name == 'School Admin'):
            return JsonResponse({'success': False, 'message': 'Unauthorized'})

        # Fetch target user
            cursor.execute('''
                SELECT 
                    u."SchoolID", u."ProfileID", u."UserName", u."Email",
                    p."ProfileName", s."SchoolName", s."SchoolLogo"
                FROM "UserMaster" u
                LEFT JOIN "ProfileMaster" p ON u."ProfileID" = p."ProfileID"
                LEFT JOIN "SchoolMaster" s ON u."SchoolID" = s."SchoolID"
                WHERE u."UserID" = %s
            ''', [user_id])
            target = cursor.fetchone()
            if not target:
                return JsonResponse({'success': False, 'message': 'User not found'})
            
            target_school_id, target_profile_id, target_name, target_email, target_profile_name, s_name, s_logo = target
            
            if current_profile_name == 'School Admin':
                if target_school_id != logged_in_school_id:
                    return JsonResponse({'success': False, 'message': 'Access Denied: Target user belongs to another school.'})
                if target_profile_id not in [3, 4, 5, 6, 7, 8]:
                    return JsonResponse({'success': False, 'message': 'Access Denied: You do not have permission to manage this profile type.'})

            # Update password
            hashed_password = make_password(new_password)
            cursor.execute('''
                UPDATE "UserMaster" 
                SET "PasswordHash" = %s,
                    "BlockedUntil" = NULL,
                    "FailedLoginAttempts" = 0
                WHERE "UserID" = %s
            ''', [hashed_password, user_id])
            connection.commit()

            # Send Security Notification
            if target_email:
                try:
                    placeholders = {
                        'user_name': target_name,
                        'login_id': target_email,
                        'profile': target_profile_name,
                        'school_name': s_name,
                        'school_logo': s_logo,
                        'ip_address': request.META.get('REMOTE_ADDR', 'Unknown'),
                        'browser': request.META.get('HTTP_USER_AGENT', 'Unknown'),
                        'timestamp': timezone.now().strftime('%Y-%m-%d %H:%M:%S')
                    }
                    send_email_by_code(
                        code='PASSWORD_CHANGED_NOTIFICATION',
                        to_emails=[target_email],
                        placeholders=placeholders,
                        school_id=target_school_id
                    )
                except Exception as ex:
                    logger.error(f"Failed to send admin-reset notification: {ex}")
            
            return JsonResponse({'success': True, 'message': f'Password for {target_name} has been reset successfully.'})
            
    except Exception as e:
        logger.error(f"Error resetting password for user {user_id}: {e}")
        return JsonResponse({'success': False, 'message': str(e)})

from django.shortcuts import render, redirect
from django.contrib import messages
from django.db import connection
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import base64
import logging
from .decorators import custom_login_required
import csv

# Export functionality updated - triggers server reload
logger = logging.getLogger(__name__)

def safe_int(value, default=0):
    """Safely convert to int"""
    try:
        return int(value) if value else default
    except (ValueError, TypeError):
        return default

def get_context(request):
    """Helper function to fetch session data and images for header."""
    from .views import get_context as base_get_context
    return base_get_context(request)

def _get_custom_session_info(request):
    """Wrapper for session info - imports locally to avoid circular dependency."""
    from .utils import _get_custom_session_info as base_session_info
    return base_session_info(request)

def custom_login_required(view_func):
    """Custom login required decorator"""
    from .views import custom_login_required as base_login_required
    return base_login_required(view_func)


@custom_login_required
def view_students(request):
    """
    View Students page - Display students in table format with filtering and pagination
    """
    context = get_context(request)
    sess = _get_custom_session_info(request)
    if sess:
        context['user'] = sess
    
    user_id = request.session.get('UserId')
    school_id = request.session.get('SchoolID')
    profile_id = request.session.get('ProfileID')
    profile_name = request.session.get('ProfileName', '')
    
    # Allow Super Admin to access without SchoolID
    is_super_admin = profile_name == 'Super Admin'
    
    if not user_id or (not school_id and not is_super_admin) or not profile_id:
        messages.error(request, "Please login to access student data")
        return redirect('login')
    
    # Fetch schools list for Super Admin filter
    schools_list = []
    if is_super_admin:
        try:
            with connection.cursor() as cursor:
                cursor.execute('SELECT "SchoolID", "SchoolName" FROM "SchoolMaster" WHERE COALESCE("IsDeleted", FALSE) = FALSE ORDER BY "SchoolName"')
                schools_list = [{'SchoolID': row[0], 'SchoolName': row[1]} for row in cursor.fetchall()]
        except Exception as e:
            logger.error(f"Error fetching schools list: {e}")

    # Handle School Filter for Super Admin
    school_filter = request.GET.get('school_filter', '')
    if is_super_admin and school_filter:
        try:
            school_id = int(school_filter)
        except (ValueError, TypeError):
            school_id = None
    elif is_super_admin and not school_id:
        school_id = None # Explicitly set to None for global view if no filter
    
    page = safe_int(request.GET.get('page', 1))
    per_page = safe_int(request.GET.get('per_page', 25))
    search = request.GET.get('search', '').strip()
    class_id = request.GET.get('class_id', '')
    section_id = request.GET.get('section_id', '')
    gender = request.GET.get('gender', '')
    category = request.GET.get('category', '')
    from_date = request.GET.get('from_date', '')
    to_date = request.GET.get('to_date', '')
    
    students = []
    total_count = 0
    
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT * FROM proc_student_cards_full_get(
                    %s, %s, %s, %s, %s, %s
                )
            """, [
                school_id,
                int(class_id) if class_id else None,
                int(section_id) if section_id else None,
                search if search else None,
                page,
                per_page
            ])
            
            columns = [col[0] for col in cursor.description]
            raw_students = cursor.fetchall()
            
            for row in raw_students:
                students.append(dict(zip(columns, row)))
            
            if students:
                total_count = students[0].get('TotalCount', 0)
            
    except Exception as e:
        logger.error(f"Error fetching students: {str(e)}", exc_info=True)
        messages.error(request, "Error loading student data. Please try again.")
    
    start_index = (page - 1) * per_page + 1 if students else 0
    end_index = min(start_index + len(students) - 1, total_count) if students else 0
    has_next = end_index < total_count
    
    context.update({
        'students': students,
        'total_count': total_count,
        'page': page,
        'items_per_page': per_page,
        'start_index': start_index,
        'end_index': end_index,
        'has_next': has_next,
        'search': search,
        'class_id': class_id,
        'section_id': section_id,
        'gender': gender,
        'category': category,
        'from_date': from_date,
        'to_date': to_date,
        'schools_list': schools_list if is_super_admin else [],
        'school_filter': school_filter,
        'is_super_admin': is_super_admin
    })
    
    return render(request, 'view_students.html', context)


@custom_login_required
def view_students_cards(request):
    """
    View Students page - Display students in card format with filtering and pagination
    """
    context = get_context(request)
    sess = _get_custom_session_info(request)
    if sess:
        context['user'] = sess
    
    user_id = request.session.get('UserId')
    school_id = request.session.get('SchoolID')
    profile_id = request.session.get('ProfileID')
    profile_name = request.session.get('ProfileName', '')
    
    # Allow Super Admin to access without SchoolID
    is_super_admin = profile_name == 'Super Admin'
    
    if not user_id or (not school_id and not is_super_admin) or not profile_id:
        messages.error(request, "Please login to access student data")
        return redirect('login')
    
    # Fetch schools list for Super Admin filter
    schools_list = []
    if is_super_admin:
        try:
            with connection.cursor() as cursor:
                cursor.execute('SELECT "SchoolID", "SchoolName" FROM "SchoolMaster" WHERE COALESCE("IsDeleted", FALSE) = FALSE ORDER BY "SchoolName"')
                schools_list = [{'SchoolID': row[0], 'SchoolName': row[1]} for row in cursor.fetchall()]
        except Exception as e:
            logger.error(f"Error fetching schools list: {e}")

    # Handle School Filter for Super Admin
    school_filter = request.GET.get('school_filter', '')
    if is_super_admin and school_filter:
        try:
            school_id = int(school_filter)
        except (ValueError, TypeError):
            school_id = request.session.get('SchoolID')
    elif is_super_admin and not school_id:
        school_id = None # Explicitly set to None for global view if no filter
    
    page = safe_int(request.GET.get('page', 1))
    per_page = safe_int(request.GET.get('per_page', 10))
    search = request.GET.get('search', '').strip()
    class_id = request.GET.get('class_id', '')
    section_id = request.GET.get('section_id', '')
    gender = request.GET.get('gender', '')
    category = request.GET.get('category', '')
    from_date = request.GET.get('from_date', '')
    to_date = request.GET.get('to_date', '')
    
    students = []
    total_count = 0
    
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT * FROM proc_student_cards_full_get(
                    %s, %s, %s, %s, %s, %s
                )
            """, [
                school_id,
                int(class_id) if class_id else None,
                int(section_id) if section_id else None,
                search if search else None,
                page,
                per_page
            ])
            
            columns = [col[0] for col in cursor.description]
            raw_students = cursor.fetchall()
            
            for row in raw_students:
                student = dict(zip(columns, row))
                
                if student.get('Photo') and isinstance(student['Photo'], bytes):
                    try:
                        student['PhotoBase64'] = base64.b64encode(student['Photo']).decode('utf-8')
                    except Exception as e:
                        logger.error(f"Error encoding photo: {e}")
                        student['PhotoBase64'] = None
                else:
                    student['PhotoBase64'] = None
                
                if student.get('SchoolLogo') and isinstance(student['SchoolLogo'], bytes):
                    try:
                        student['SchoolLogoBase64'] = base64.b64encode(student['SchoolLogo']).decode('utf-8')
                    except Exception as e:
                        logger.error(f"Error encoding school logo: {e}")
                        student['SchoolLogoBase64'] = None
                else:
                    student['SchoolLogoBase64'] = None
                
                if student.get('AdmissionDate'):
                    student['AdmissionDateFormatted'] = student['AdmissionDate'].strftime('%Y-%m-%d')
                else:
                    student['AdmissionDateFormatted'] = 'N/A'
                    
                if student.get('DateOfBirth'):
                    student['DateOfBirthFormatted'] = student['DateOfBirth'].strftime('%Y-%m-%d')
                else:
                    student['DateOfBirthFormatted'] = 'N/A'
                
                students.append(student)
            
            if students:
                total_count = students[0].get('TotalCount', 0)
            
    except Exception as e:
        logger.error(f"Error fetching students: {str(e)}", exc_info=True)
        messages.error(request, "Error loading student data. Please try again.")
    
    start_index = (page - 1) * per_page + 1 if students else 0
    end_index = min(start_index + len(students) - 1, total_count) if students else 0
    has_next = end_index < total_count
    has_prev = page > 1
    
    card_template = 'core/document_templates/student_id_card/student_card_horizontal_1.html'
    try:
        with connection.cursor() as cursor:
            cursor.execute('SELECT "TemplateType", "TemplateFile" FROM "Proc_Template_Preference_Get"(%s)', [school_id])
            for row in cursor.fetchall():
                if row[0] == 'StudentCard':
                    card_template = row[1]
                    break
    except Exception as e:
        logger.error(f"Error fetching student card template: {e}")
    
    context.update({
        'students': students,
        'total_count': total_count,
        'page': page,
        'items_per_page': per_page,
        'start_index': start_index,
        'end_index': end_index,
        'has_next': has_next,
        'has_prev': has_prev,
        'search': search,
        'class_id': class_id,
        'section_id': section_id,
        'gender': gender,
        'category': category,
        'from_date': from_date,
        'to_date': to_date,
        'card_template': card_template,
        'is_super_admin': is_super_admin,
        'schools_list': schools_list,
        'school_filter': school_filter
    })
    
    return render(request, 'view_students_cards.html', context)


@custom_login_required
def students_export(request):
    """Export students as Excel or CSV"""
    format_type = request.GET.get('format', 'excel').lower()
    
    # Get session data
    user_id = request.session.get('UserId')
    school_id = request.session.get('SchoolID')
    profile_id = request.session.get('ProfileID')
    profile_name = request.session.get('ProfileName', '')
    
    # Allow Super Admin to access without SchoolID
    is_super_admin = profile_name == 'Super Admin'
    
    if not user_id or (not school_id and not is_super_admin) or not profile_id:
        return HttpResponse("Unauthorized", status=401)
    
    # Handle School Filter for Super Admin
    school_filter = request.GET.get('school_filter', '')
    if is_super_admin and school_filter:
        try:
            school_id = int(school_filter)
        except (ValueError, TypeError):
            school_id = None
    elif is_super_admin and not school_id:
        school_id = None
    
    # Get filter parameters
    search = request.GET.get('search', '').strip()
    class_id = request.GET.get('class_id', '')
    section_id = request.GET.get('section_id', '')
    
    students = []
    
    try:
        with connection.cursor() as cursor:
            # Fetch all students (large page size for export)
            cursor.execute("""
                SELECT * FROM proc_student_cards_full_get(
                    %s, %s, %s, %s, %s, %s
                )
            """, [
                school_id,
                int(class_id) if class_id else None,
                int(section_id) if section_id else None,
                search if search else None,
                1,  # Page 1
                10000  # Large page size for export
            ])
            
            columns = [col[0] for col in cursor.description]
            raw_students = cursor.fetchall()
            
            for row in raw_students:
                student = dict(zip(columns, row))
                
                # Format dates
                if student.get('AdmissionDate'):
                    student['AdmissionDate'] = student['AdmissionDate'].strftime('%Y-%m-%d')
                if student.get('DateOfBirth'):
                    student['DateOfBirth'] = student['DateOfBirth'].strftime('%Y-%m-%d')
                
                # Remove internal/blob columns
                columns_to_remove = ['Photo', 'SchoolLogo', 'TotalCount', 'PhotoBase64', 'SchoolLogoBase64']
                for col in columns_to_remove:
                    if col in student:
                        del student[col]
                
                students.append(student)
                
    except Exception as e:
        logger.error(f"Error fetching students for export: {e}")
        return HttpResponse("Error generating export file", status=500)
    
    filename = f"students_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    
    if format_type == 'csv':
        delimiter_type = request.GET.get('delimiter', 'comma').lower()
        delimiters = {'comma': ',', 'pipe': '|', 'tab': '\t'}
        actual_delimiter = delimiters.get(delimiter_type, ',')
        
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
        
        if students:
            writer = csv.DictWriter(response, fieldnames=students[0].keys(), delimiter=actual_delimiter)
            writer.writeheader()
            writer.writerows(students)
        return response
    
    else:  # Excel
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Students"
        
        if students:
            headers = list(students[0].keys())
            # Header styling
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            # Data rows
            for row_num, student in enumerate(students, 2):
                for col_num, (key, value) in enumerate(student.items(), 1):
                    ws.cell(row=row_num, column=col_num, value=str(value) if value is not None else "")
            
            # Auto-adjust column width
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(response)
        return response
